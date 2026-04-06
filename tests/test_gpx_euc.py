import pytest

from gopro_overlay import gpx
from gopro_overlay.gpx import _preprocess_gpx, load_xml
from gopro_overlay.units import units


def test_preprocess_gpx_euc_world_speed_relocated():
    xml = """<gpx xmlns="http://www.topografix.com/GPX/1/1" version="1.1">
  <trk>
    <trkseg>
      <trkpt lat="51.5" lon="-0.14">
        <ele>100.0</ele>
        <time>2024-01-01T12:00:00Z</time>
        <speed>36</speed>
      </trkpt>
    </trkseg>
  </trk>
</gpx>"""

    result = _preprocess_gpx(xml)
    # Speed should be relocated into extensions/TrackPointExtension
    assert "gpxtpx:speed" in result
    assert "TrackPointExtension" in result
    # 36 km/h = 10.0 m/s
    assert "10.00" in result


def test_preprocess_gpx_speed_conversion_kmh_to_mps():
    xml = """<gpx xmlns="http://www.topografix.com/GPX/1/1" version="1.1">
  <trk>
    <trkseg>
      <trkpt lat="51.5" lon="-0.14">
        <ele>100.0</ele>
        <time>2024-01-01T12:00:00Z</time>
        <speed>72</speed>
      </trkpt>
    </trkseg>
  </trk>
</gpx>"""

    result = _preprocess_gpx(xml)
    # 72 km/h / 3.6 = 20.0 m/s
    assert "20.00" in result


def test_preprocess_gpx_no_change_when_already_in_extensions():
    xml = """<gpx
    xmlns="http://www.topografix.com/GPX/1/1"
    xmlns:gpxtpx="http://www.garmin.com/xmlschemas/TrackPointExtension/v1"
    version="1.1">
  <trk>
    <trkseg>
      <trkpt lat="51.5" lon="-0.14">
        <ele>100.0</ele>
        <time>2024-01-01T12:00:00Z</time>
        <extensions>
          <gpxtpx:TrackPointExtension>
            <gpxtpx:speed>6.86</gpxtpx:speed>
          </gpxtpx:TrackPointExtension>
        </extensions>
      </trkpt>
    </trkseg>
  </trk>
</gpx>"""

    result = _preprocess_gpx(xml)
    # Should be unchanged — speed is already in TrackPointExtension
    assert result == xml


def test_preprocess_gpx_no_speed_element_unchanged():
    xml = """<gpx xmlns="http://www.topografix.com/GPX/1/1" version="1.1">
  <trk>
    <trkseg>
      <trkpt lat="51.5" lon="-0.14">
        <ele>100.0</ele>
        <time>2024-01-01T12:00:00Z</time>
      </trkpt>
    </trkseg>
  </trk>
</gpx>"""

    result = _preprocess_gpx(xml)
    assert result == xml


def test_preprocess_gpx_zero_speed():
    xml = """<gpx xmlns="http://www.topografix.com/GPX/1/1" version="1.1">
  <trk>
    <trkseg>
      <trkpt lat="51.5" lon="-0.14">
        <ele>100.0</ele>
        <time>2024-01-01T12:00:00Z</time>
        <speed>0</speed>
      </trkpt>
    </trkseg>
  </trk>
</gpx>"""

    result = _preprocess_gpx(xml)
    assert "0.00" in result


def test_load_xml_battery_voltage_current():
    xml = """<gpx
    xmlns="http://www.topografix.com/GPX/1/1"
    xmlns:gpxtpx="http://www.garmin.com/xmlschemas/TrackPointExtension/v1"
    version="1.1">
  <trk>
    <trkseg>
      <trkpt lat="51.5" lon="-0.14">
        <ele>100.0</ele>
        <time>2024-01-01T12:00:00Z</time>
        <extensions>
          <gpxtpx:TrackPointExtension>
            <gpxtpx:speed>10.0</gpxtpx:speed>
            <gpxtpx:battery>85</gpxtpx:battery>
            <gpxtpx:voltage>67.2</gpxtpx:voltage>
            <gpxtpx:current>5.1</gpxtpx:current>
          </gpxtpx:TrackPointExtension>
        </extensions>
      </trkpt>
    </trkseg>
  </trk>
</gpx>"""

    entries = list(load_xml(xml, units))
    assert len(entries) == 1
    assert entries[0].battery == units.Quantity(85, units.percent)
    assert entries[0].voltage == units.Quantity(67.2, units.volt)
    assert entries[0].current == units.Quantity(5.1, units.ampere)
    assert entries[0].speed == units.Quantity(10.0, units.mps)


def test_load_xml_battery_none_when_absent():
    xml = """<gpx
    xmlns="http://www.topografix.com/GPX/1/1"
    xmlns:gpxtpx="http://www.garmin.com/xmlschemas/TrackPointExtension/v1"
    version="1.1">
  <trk>
    <trkseg>
      <trkpt lat="51.5" lon="-0.14">
        <ele>100.0</ele>
        <time>2024-01-01T12:00:00Z</time>
        <extensions>
          <gpxtpx:TrackPointExtension>
            <gpxtpx:speed>6.86</gpxtpx:speed>
          </gpxtpx:TrackPointExtension>
        </extensions>
      </trkpt>
    </trkseg>
  </trk>
</gpx>"""

    entries = list(load_xml(xml, units))
    assert len(entries) == 1
    assert entries[0].battery is None
    assert entries[0].voltage is None
    assert entries[0].current is None


def test_load_xlsx_basic(tmp_path):
    import openpyxl
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.append(["GPS Latitude [°]", "GPS Longitude [°]", "Date & Time",
               "GPS Altitude [m]", "Speed [km/h]", "Battery [%]",
               "Voltage [V]", "Current [A]", "Power [W]"])
    ws.append([51.5, -0.14, "2024-01-01 12:00:00", 100.0, 36.0, 85.0, 67.2, 5.1, 340.0])
    ws.append([51.501, -0.141, "2024-01-01 12:00:01", 101.0, 72.0, 84.0, 66.8, 10.2, 680.0])
    path = tmp_path / "test.xlsx"
    wb.save(path)

    from gopro_overlay.gpx import load_xlsx
    entries = load_xlsx(path, units)

    assert len(entries) == 2
    assert entries[0].lat == 51.5
    assert entries[0].lon == -0.14
    assert entries[0].battery == units.Quantity(85.0, units.percent)
    assert entries[0].voltage == units.Quantity(67.2, units.volt)
    assert entries[0].current == units.Quantity(5.1, units.ampere)


def test_load_xlsx_speed_converted_to_mps(tmp_path):
    import openpyxl
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.append(["GPS Latitude [°]", "GPS Longitude [°]", "Date & Time",
               "Speed [km/h]"])
    ws.append([51.5, -0.14, "2024-01-01 12:00:00", 36.0])
    path = tmp_path / "test.xlsx"
    wb.save(path)

    from gopro_overlay.gpx import load_xlsx
    entries = load_xlsx(path, units)

    assert len(entries) == 1
    # 36 km/h / 3.6 = 10.0 m/s
    assert entries[0].speed.to("mps").magnitude == pytest.approx(10.0)


def test_load_xlsx_missing_required_columns(tmp_path):
    import openpyxl
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.append(["Some Column", "Another Column"])
    ws.append([1, 2])
    path = tmp_path / "test.xlsx"
    wb.save(path)

    from gopro_overlay.gpx import load_xlsx
    with pytest.raises(IOError, match="missing required columns"):
        load_xlsx(path, units)
