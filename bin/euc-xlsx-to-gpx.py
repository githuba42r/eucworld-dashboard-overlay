#!/usr/bin/env python3
"""
Convert EUC World XLSX export to GPX with extended telemetry.

Extracts GPS data plus EUC-specific metrics (battery, voltage, current,
power, temperature) and writes a standard GPX 1.1 file with Garmin
TrackPointExtension for speed and custom euc: namespace extensions for
the additional telemetry fields.

Usage:
    euc-xlsx-to-gpx.py input.xlsx [output.gpx]

If output is omitted, writes to input stem + .gpx in the same directory.
"""

import argparse
import sys
from datetime import datetime, timezone, timedelta
from pathlib import Path

try:
    import openpyxl
except ImportError:
    sys.exit("openpyxl is required: pip install openpyxl")


EUC_NS = "https://euc.world/gpx/1/0"
TPX_NS = "http://www.garmin.com/xmlschemas/TrackPointExtension/v1"


def convert(xlsx_path: Path, gpx_path: Path):
    wb = openpyxl.load_workbook(xlsx_path, read_only=True)
    ws = wb.active

    # Read headers
    headers = []
    for row in ws.iter_rows(min_row=1, max_row=1, values_only=True):
        headers = [h.strip() if h else "" for h in row]

    # Map header names to column indices
    col = {}
    for i, h in enumerate(headers):
        col[h] = i

    # Required columns
    lat_col = col.get("GPS Latitude [°]")
    lon_col = col.get("GPS Longitude [°]")
    time_col = col.get("Date & Time")
    if lat_col is None or lon_col is None or time_col is None:
        sys.exit(f"Missing required columns. Found: {headers}")

    # Optional columns
    alt_col = col.get("GPS Altitude [m]")
    gps_speed_col = col.get("GPS Speed [km/h]")
    speed_col = col.get("Speed [km/h]")
    battery_col = col.get("Battery [%]")
    voltage_col = col.get("Voltage [V]")
    current_col = col.get("Current [A]")
    power_col = col.get("Power [W]")
    temp_col = col.get("Temperature [°C]")
    bearing_col = col.get("GPS Bearing [°]")

    points = []
    for row in ws.iter_rows(min_row=2, values_only=True):
        dt_val = row[time_col]
        lat = row[lat_col]
        lon = row[lon_col]

        if dt_val is None or lat is None or lon is None:
            continue

        # Parse datetime — XLSX gives naive datetime in local time
        # EUC World exports in device local time without timezone
        if isinstance(dt_val, datetime):
            dt = dt_val
        else:
            dt = datetime.strptime(str(dt_val), "%Y-%m-%d %H:%M:%S")

        point = {
            "lat": float(lat),
            "lon": float(lon),
            "time": dt,
        }

        def safe_float(idx):
            if idx is not None and idx < len(row) and row[idx] is not None:
                try:
                    return float(row[idx])
                except (ValueError, TypeError):
                    pass
            return None

        point["alt"] = safe_float(alt_col)
        point["gps_speed"] = safe_float(gps_speed_col)
        point["speed"] = safe_float(speed_col)
        point["battery"] = safe_float(battery_col)
        point["voltage"] = safe_float(voltage_col)
        point["current"] = safe_float(current_col)
        point["power"] = safe_float(power_col)
        point["temp"] = safe_float(temp_col)
        point["bearing"] = safe_float(bearing_col)

        points.append(point)

    wb.close()

    if not points:
        sys.exit("No valid data points found in XLSX")

    # Detect timezone: EUC World XLSX has local time but GPX needs UTC.
    # We infer from the GPX file if it exists alongside, otherwise assume local system tz.
    # For now, use the system's local timezone offset.
    local_tz = datetime.now().astimezone().tzinfo

    # Write GPX
    with open(gpx_path, "w") as out:
        out.write('<?xml version="1.0" encoding="UTF-8"?>\n')
        out.write('<gpx version="1.1" creator="EUC World XLSX Converter"\n')
        out.write(f'  xmlns="http://www.topografix.com/GPX/1/1"\n')
        out.write(f'  xmlns:gpxtpx="{TPX_NS}"\n')
        out.write(f'  xmlns:euc="{EUC_NS}">\n')
        out.write('  <trk>\n')
        out.write(f'    <name>{xlsx_path.stem}</name>\n')
        out.write('    <trkseg>\n')

        for pt in points:
            dt = pt["time"]
            if dt.tzinfo is None:
                dt = dt.replace(tzinfo=local_tz)
            utc_dt = dt.astimezone(timezone.utc)
            time_str = utc_dt.strftime("%Y-%m-%dT%H:%M:%SZ")

            out.write(f'      <trkpt lat="{pt["lat"]}" lon="{pt["lon"]}">\n')
            if pt["alt"] is not None:
                out.write(f'        <ele>{pt["alt"]}</ele>\n')
            out.write(f'        <time>{time_str}</time>\n')

            # Extensions
            has_ext = any(pt.get(k) is not None for k in
                         ("speed", "gps_speed", "battery", "voltage", "current", "power", "temp"))
            if has_ext:
                out.write('        <extensions>\n')

                # Speed in standard TrackPointExtension (m/s)
                spd = pt.get("speed") or pt.get("gps_speed")
                if spd is not None:
                    spd_mps = spd / 3.6
                    out.write(f'          <gpxtpx:TrackPointExtension>\n')
                    out.write(f'            <gpxtpx:speed>{spd_mps:.2f}</gpxtpx:speed>\n')
                    if pt.get("temp") is not None:
                        out.write(f'            <gpxtpx:atemp>{pt["temp"]:.1f}</gpxtpx:atemp>\n')
                    out.write(f'          </gpxtpx:TrackPointExtension>\n')

                # EUC-specific extensions
                euc_fields = []
                if pt.get("battery") is not None:
                    euc_fields.append(("battery", f'{pt["battery"]:.1f}'))
                if pt.get("voltage") is not None:
                    euc_fields.append(("voltage", f'{pt["voltage"]:.2f}'))
                if pt.get("current") is not None:
                    euc_fields.append(("current", f'{pt["current"]:.2f}'))
                if pt.get("power") is not None:
                    euc_fields.append(("power", f'{pt["power"]:.1f}'))
                if pt.get("speed") is not None and pt.get("gps_speed") is not None:
                    # Include wheel speed separately if both exist
                    euc_fields.append(("wheel_speed", f'{pt["speed"] / 3.6:.2f}'))

                if euc_fields:
                    out.write(f'          <euc:TelemetryExtension>\n')
                    for tag, val in euc_fields:
                        out.write(f'            <euc:{tag}>{val}</euc:{tag}>\n')
                    out.write(f'          </euc:TelemetryExtension>\n')

                out.write('        </extensions>\n')

            out.write('      </trkpt>\n')

        out.write('    </trkseg>\n')
        out.write('  </trk>\n')
        out.write('</gpx>\n')

    print(f"Converted {len(points)} points")
    print(f"Output: {gpx_path}")


def main():
    parser = argparse.ArgumentParser(description="Convert EUC World XLSX export to GPX with telemetry extensions")
    parser.add_argument("input", type=Path, help="Input XLSX file from EUC World")
    parser.add_argument("output", type=Path, nargs="?", help="Output GPX file (default: input.gpx)")
    args = parser.parse_args()

    if not args.input.exists():
        sys.exit(f"File not found: {args.input}")

    output = args.output or args.input.with_suffix(".gpx")
    convert(args.input, output)


if __name__ == "__main__":
    main()
