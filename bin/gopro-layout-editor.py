#!/usr/bin/env python3
"""
GoPro Dashboard Overlay — Visual Layout Editor

A tkinter GUI for visually positioning overlay components on video frames,
generating XML layouts, and launching encoding with gopro-dashboard-overlay.
"""

import io
import json
import math
import os
import queue
import subprocess
import sys
import threading
import tkinter as tk
from dataclasses import dataclass, field
from pathlib import Path
from tkinter import filedialog, messagebox, ttk
from typing import Optional

from PIL import Image, ImageTk


# ---------------------------------------------------------------------------
# Data model
# ---------------------------------------------------------------------------

@dataclass
class VideoEntry:
    path: Path
    raw_width: int
    raw_height: int
    rotation: int
    eff_width: int
    eff_height: int
    duration_seconds: float
    creation_time: str  # ISO string

    @property
    def basename(self):
        return self.path.name


@dataclass
class OverlayComponent:
    name: str
    label: str
    x: int  # full-res coords
    y: int
    width: int  # full-res bounding box
    height: int
    enabled: bool = True
    color: str = "#4488ff"
    canvas_rect_id: Optional[int] = None
    canvas_text_id: Optional[int] = None
    canvas_handle_ids: dict = field(default_factory=dict)  # edge -> canvas item id
    xml_template: str = ""
    # Per-component custom properties (e.g. date/time formats)
    custom_props: dict = field(default_factory=dict)


# Component definitions: (name, label, ref_w, ref_h, colour, xml_template)
# Positions are set dynamically based on video dimensions
COMPONENT_DEFS = [
    ("date_and_time", "Date + Time", 200, 60, "#44aa88",
     '''    <composite x="{x}" y="{y}" width="{w}" height="{h}" name="date_and_time">
        <component type="datetime" x="0" y="0" format="{date_format}" size="{date_size}" align="left"/>
        <component type="datetime" x="0" y="{time_y}" format="{time_format}" {time_truncate}size="{time_size}" align="left"/>
    </composite>'''),

    ("date_only", "Date Only", 200, 24, "#44aa66",
     '''    <composite x="{x}" y="{y}" width="{w}" height="{h}" name="date_only">
        <component type="datetime" x="0" y="0" format="{date_format}" size="{date_size}" align="right"/>
    </composite>'''),

    ("time_only", "Time Only", 200, 36, "#44aacc",
     '''    <composite x="{x}" y="{y}" width="{w}" height="{h}" name="time_only">
        <component type="datetime" x="0" y="0" format="{time_format}" {time_truncate}size="{time_size}" align="right"/>
    </composite>'''),

    ("gps_info", "GPS (Full)", 280, 80, "#88aa44",
     '''    <composite x="{x}" y="{y}" width="{w}" height="{h}" name="gps_info">
        <frame name="gps-lock" x="226" y="24" width="32" height="32" bg="0,0,0,128" cr="5" opacity="0.4">
            <component type="gps-lock-icon" size="32"/>
         </frame>
        <composite y="36">
            <component type="text" x="0" y="0" size="{text_size}" align="left">GPS INFO</component>
            <component type="text" x="0" y="24" size="{text_size}" align="left">Lat: </component>
            <component type="text" x="128" y="24" size="{text_size}" align="left">Lon: </component>
            <component type="metric" x="118" y="24" metric="lat" dp="6" size="{text_size}" align="right" cache="False"/>
            <component type="metric" x="256" y="24" metric="lon" dp="6" size="{text_size}" align="right" cache="False"/>
        </composite>
    </composite>'''),

    ("gps_coords", "GPS Coordinates", 280, 24, "#88cc44",
     '''    <composite x="{x}" y="{y}" width="{w}" height="{h}" name="gps_coords">
        <component type="text" x="0" y="0" size="{text_size}" align="left">Lat: </component>
        <component type="text" x="128" y="0" size="{text_size}" align="left">Lon: </component>
        <component type="metric" x="118" y="0" metric="lat" dp="6" size="{text_size}" align="right" cache="False"/>
        <component type="metric" x="256" y="0" metric="lon" dp="6" size="{text_size}" align="right" cache="False"/>
    </composite>'''),

    ("gps_lock", "GPS Lock", 40, 40, "#88aa88",
     '''    <frame name="gps_lock" x="{x}" y="{y}" width="32" height="32" bg="0,0,0,128" cr="5" opacity="0.4">
        <component type="gps-lock-icon" size="32"/>
    </frame>'''),

    ("big_mph", "Speed", 208, 160, "#ff6644",
     '''    <composite x="{x}" y="{y}" width="{w}" height="{h}" name="big_mph">
        <component type="metric_unit" metric="speed" units="speed" size="{text_size}" rgb="{label_rgb}">{{:~c}}</component>
        <component type="metric" x="0" y="0" metric="speed" units="speed" dp="0" size="{speed_font_size}" rgb="{value_rgb}" />
    </composite>'''),

    ("gradient_chart", "Alt Chart", 600, 80, "#aa88cc",
     '''    <composite x="{x}" y="{y}" name="gradient_chart">
        <component type="chart" x="0" y="0" units="alt" width="{w}" height="{h}" filled="{chart_filled}" seconds="{chart_seconds}" bg="{chart_bg}" fill="{chart_fill}" line="{chart_line}"/>
{chart_title_line}
    </composite>'''),

    ("gradient", "Gradient", 150, 70, "#ccaa44",
     '''    <composite x="{x}" y="{y}" width="{w}" height="{h}" name="gradient">
        <component type="text" x="{icon_size}" y="0" size="{text_size}" rgb="{label_rgb}">SLOPE(%)</component>
        <component type="icon" x="0" y="0" file="slope-triangle.png" size="{icon_size}"/>
        <component type="metric" x="{icon_size}" y="{value_y}" metric="gradient" dp="0" size="{value_font_size}" rgb="{value_rgb}" />
    </composite>'''),

    ("altitude", "Altitude", 150, 70, "#44ccaa",
     '''    <composite x="{x}" y="{y}" width="{w}" height="{h}" name="altitude">
        <component type="metric_unit" x="{icon_size}" y="0" metric="alt" units="alt" size="{text_size}" rgb="{label_rgb}">ALT({{:~C}})</component>
        <component type="icon" x="0" y="0" file="mountain.png" size="{icon_size}"/>
        <component type="metric" x="{icon_size}" y="{value_y}" metric="alt" units="alt" dp="0" size="{value_font_size}" rgb="{value_rgb}" />
    </composite>'''),

    ("temperature", "Temp", 150, 70, "#cc4444",
     '''    <composite x="{x}" y="{y}" width="{w}" height="{h}" name="temperature">
        <component type="metric_unit" x="-70" y="0" size="{text_size}" align="right" metric="temp" units="temp">TEMP({{:~C}})</component>
        <component type="icon" x="-64" y="0" file="thermometer.png" size="64"/>
        <component type="metric" x="-70" y="18" dp="0" size="{value_font_size}" align="right" metric="temp" units="temp"/>
    </composite>'''),

    ("cadence", "Cadence", 150, 70, "#44cc44",
     '''    <composite x="{x}" y="{y}" width="{w}" height="{h}" name="cadence">
        <component type="text" x="-70" y="0" size="{text_size}" align="right">RPM</component>
        <component type="icon" x="-64" y="0" file="gauge.png" size="64"/>
        <component type="metric" x="-70" y="18" metric="cadence" dp="0" size="{value_font_size}" align="right"/>
    </composite>'''),

    ("heartbeat", "Heart Rate", 150, 70, "#cc4488",
     '''    <composite x="{x}" y="{y}" width="{w}" height="{h}" name="heartbeat">
        <component type="text" x="-70" y="0" size="{text_size}" align="right">BPM</component>
        <component type="icon" x="-64" y="0" file="heartbeat.png" size="64"/>
        <component type="metric" x="-70" y="18" metric="hr" dp="0" size="{value_font_size}" align="right"/>
    </composite>'''),

    ("moving_map", "Moving Map", 256, 256, "#6688cc",
     '''    <component type="moving_map" name="moving_map" x="{x}" y="{y}" size="{map_size}" zoom="{map_zoom}" corner_radius="{map_corner_radius}" opacity="{map_opacity}" rotate="{map_rotate}"/>'''),

    ("journey_map", "Journey Map", 256, 256, "#8866cc",
     '''    <component type="journey_map" name="journey_map" x="{x}" y="{y}" size="{map_size}" corner_radius="{map_corner_radius}" opacity="{map_opacity}" fill="{jm_track_colour}" line-width="{jm_line_width}" loc-fill="{jm_loc_colour}" loc-outline="{jm_loc_outline}" loc-size="{jm_loc_size}"/>'''),

    # -- EUC telemetry components --
    ("battery_value", "Battery %", 150, 70, "#44bb44",
     '''    <composite x="{x}" y="{y}" width="{w}" height="{h}" name="battery_value">
        <component type="text" x="0" y="0" size="{text_size}" rgb="{label_rgb}">BATTERY</component>
        <component type="metric" x="0" y="{value_y}" metric="battery" dp="0" size="{value_font_size}" rgb="{value_rgb}" />
        <component type="text" x="{w}" y="{value_y}" size="{value_font_size}" rgb="{value_rgb}" align="right">%</component>
    </composite>'''),

    ("voltage_value", "Voltage", 150, 70, "#bbbb44",
     '''    <composite x="{x}" y="{y}" width="{w}" height="{h}" name="voltage_value">
        <component type="text" x="0" y="0" size="{text_size}" rgb="{label_rgb}">VOLTAGE</component>
        <component type="metric" x="0" y="{value_y}" metric="voltage" dp="1" size="{value_font_size}" rgb="{value_rgb}" />
        <component type="text" x="{w}" y="{value_y}" size="{value_font_size}" rgb="{value_rgb}" align="right">V</component>
    </composite>'''),

    ("current_value", "Current", 150, 70, "#bb8844",
     '''    <composite x="{x}" y="{y}" width="{w}" height="{h}" name="current_value">
        <component type="text" x="0" y="0" size="{text_size}" rgb="{label_rgb}">CURRENT</component>
        <component type="metric" x="0" y="{value_y}" metric="current" dp="1" size="{value_font_size}" rgb="{value_rgb}" />
        <component type="text" x="{w}" y="{value_y}" size="{value_font_size}" rgb="{value_rgb}" align="right">A</component>
    </composite>'''),

    ("power_value", "Power", 150, 70, "#bb4488",
     '''    <composite x="{x}" y="{y}" width="{w}" height="{h}" name="power_value">
        <component type="text" x="0" y="0" size="{text_size}" rgb="{label_rgb}">POWER</component>
        <component type="metric" x="0" y="{value_y}" metric="power" dp="0" size="{value_font_size}" rgb="{value_rgb}" />
        <component type="text" x="{w}" y="{value_y}" size="{value_font_size}" rgb="{value_rgb}" align="right">W</component>
    </composite>'''),

    ("battery_chart", "Battery Chart", 300, 80, "#44bb44",
     '''    <composite x="{x}" y="{y}" name="battery_chart">
        <component type="chart" x="0" y="0" metric="battery" units="percent" width="{w}" height="{h}" filled="true" seconds="300" bg="0,0,0,170" fill="68,187,68,170" line="255,255,255,170"/>
{chart_title_line}
    </composite>'''),

    ("voltage_chart", "Voltage Chart", 300, 80, "#bbbb44",
     '''    <composite x="{x}" y="{y}" name="voltage_chart">
        <component type="chart" x="0" y="0" metric="voltage" units="volt" width="{w}" height="{h}" filled="true" seconds="300" bg="0,0,0,170" fill="187,187,68,170" line="255,255,255,170"/>
{chart_title_line}
    </composite>'''),

    ("power_chart", "Power Chart", 300, 80, "#bb4488",
     '''    <composite x="{x}" y="{y}" name="power_chart">
        <component type="chart" x="0" y="0" metric="power" units="watt" width="{w}" height="{h}" filled="true" seconds="300" bg="0,0,0,170" fill="187,68,136,170" line="255,255,255,170"/>
{chart_title_line}
    </composite>'''),

    # -- Gauges --
    ("speed_gauge", "Speed Gauge", 256, 256, "#ff6644",
     '''    <composite x="{x}" y="{y}" width="{w}" height="{h}" name="speed_gauge">
        <component type="{gauge_style}" size="{gauge_size}" metric="speed" units="kph" max="{gauge_max}" min="{gauge_min}" start="{gauge_start}" length="{gauge_length}" sectors="{gauge_sectors}"{gauge_colour_attrs}/>
{gauge_title_line}
    </composite>'''),

    ("battery_gauge", "Battery Gauge", 256, 256, "#44bb44",
     '''    <composite x="{x}" y="{y}" width="{w}" height="{h}" name="battery_gauge">
        <component type="{gauge_style}" size="{gauge_size}" metric="battery" units="percent" max="{gauge_max}" min="{gauge_min}" start="{gauge_start}" length="{gauge_length}" sectors="{gauge_sectors}"{gauge_colour_attrs}/>
{gauge_title_line}
    </composite>'''),

    ("power_gauge", "Power Gauge", 256, 256, "#bb4488",
     '''    <composite x="{x}" y="{y}" width="{w}" height="{h}" name="power_gauge">
        <component type="{gauge_style}" size="{gauge_size}" metric="power" units="watt" max="{gauge_max}" min="{gauge_min}" start="{gauge_start}" length="{gauge_length}" sectors="{gauge_sectors}"{gauge_colour_attrs}/>
{gauge_title_line}
    </composite>'''),

    ("voltage_gauge", "Voltage Gauge", 256, 256, "#bbbb44",
     '''    <composite x="{x}" y="{y}" width="{w}" height="{h}" name="voltage_gauge">
        <component type="{gauge_style}" size="{gauge_size}" metric="voltage" units="volt" max="{gauge_max}" min="{gauge_min}" start="{gauge_start}" length="{gauge_length}" sectors="{gauge_sectors}"{gauge_colour_attrs}/>
{gauge_title_line}
    </composite>'''),

    ("compass_display", "Compass", 256, 256, "#6688cc",
     '''    <composite x="{x}" y="{y}" width="{w}" height="{h}" name="compass_display">
        <component type="compass" size="{gauge_size}"/>
    </composite>'''),

    ("compass_arrow_display", "Compass Arrow", 256, 256, "#4466aa",
     '''    <composite x="{x}" y="{y}" width="{w}" height="{h}" name="compass_arrow_display">
        <component type="compass-arrow" size="{gauge_size}"/>
    </composite>'''),

    ("circuit_map", "Circuit Map", 256, 256, "#66aa88",
     '''    <composite x="{x}" y="{y}" width="{w}" height="{h}" name="circuit_map">
        <component type="cairo-circuit-map" size="{gauge_size}"/>
    </composite>'''),

    # -- Bars --
    ("speed_bar", "Speed Bar", 400, 30, "#ff8844",
     '''    <composite x="{x}" y="{y}" width="{w}" height="{h}" name="speed_bar">
        <component type="bar" width="{w}" height="{h}" metric="speed" units="kph" max="{gauge_max}" min="0" outline="255,255,255,128" fill="255,136,68,200"/>
{bar_scale_line}
{chart_title_line}
    </composite>'''),

    ("battery_bar", "Battery Bar", 400, 30, "#44bb44",
     '''    <composite x="{x}" y="{y}" width="{w}" height="{h}" name="battery_bar">
        <component type="bar" width="{w}" height="{h}" metric="battery" units="percent" max="100" min="0" outline="255,255,255,128" fill="68,187,68,200"/>
{bar_scale_line}
{chart_title_line}
    </composite>'''),

    # -- Additional maps --
    ("moving_journey_map", "Moving Journey Map", 256, 256, "#7766cc",
     '''    <component type="moving-journey-map" name="moving_journey_map" x="{x}" y="{y}" size="{map_size}" zoom="{map_zoom}" corner_radius="{map_corner_radius}" opacity="{map_opacity}"/>'''),

    # -- ASI / MSI gauges --
    ("asi_gauge", "Airspeed Indicator", 256, 256, "#448888",
     '''    <composite x="{x}" y="{y}" width="{w}" height="{h}" name="asi_gauge">
        <component type="asi" size="{gauge_size}" metric="speed" units="kph"/>
    </composite>'''),

    ("msi_gauge", "Motor Speed Indicator", 256, 256, "#884488",
     '''    <composite x="{x}" y="{y}" width="{w}" height="{h}" name="msi_gauge">
        <component type="msi2" size="{gauge_size}" metric="speed" units="kph" green="{msi_green}" yellow="{msi_yellow}" end="{msi_end}" rotate="{msi_rotate}" textsize="{msi_textsize}"/>
{gauge_title_line}
    </composite>'''),

    # -- Average speed components --
    ("avg_speed_value", "Avg Speed", 150, 70, "#ff9944",
     '''    <composite x="{x}" y="{y}" width="{w}" height="{h}" name="avg_speed_value">
        <component type="text" x="0" y="0" size="{text_size}" rgb="{label_rgb}">AVG SPEED</component>
        <component type="metric" x="0" y="{value_y}" metric="avg-speed" units="speed" dp="1" size="{value_font_size}" rgb="{value_rgb}" />
    </composite>'''),

    ("avg_speed_moving_value", "Avg Moving Speed", 150, 70, "#ffbb44",
     '''    <composite x="{x}" y="{y}" width="{w}" height="{h}" name="avg_speed_moving_value">
        <component type="text" x="0" y="0" size="{text_size}" rgb="{label_rgb}">AVG MOVING</component>
        <component type="metric" x="0" y="{value_y}" metric="avg-speed-moving" units="speed" dp="1" size="{value_font_size}" rgb="{value_rgb}" />
    </composite>'''),

    ("avg_speed_gauge", "Avg Speed Gauge", 256, 256, "#ff9944",
     '''    <composite x="{x}" y="{y}" width="{w}" height="{h}" name="avg_speed_gauge">
        <component type="{gauge_style}" size="{gauge_size}" metric="avg-speed" units="kph" max="{gauge_max}" min="{gauge_min}" start="{gauge_start}" length="{gauge_length}" sectors="{gauge_sectors}"{gauge_colour_attrs}/>
{gauge_title_line}
    </composite>'''),

    ("avg_speed_moving_gauge", "Avg Moving Gauge", 256, 256, "#ffbb44",
     '''    <composite x="{x}" y="{y}" width="{w}" height="{h}" name="avg_speed_moving_gauge">
        <component type="{gauge_style}" size="{gauge_size}" metric="avg-speed-moving" units="kph" max="{gauge_max}" min="{gauge_min}" start="{gauge_start}" length="{gauge_length}" sectors="{gauge_sectors}"{gauge_colour_attrs}/>
{gauge_title_line}
    </composite>'''),

    ("avg_speed_bar", "Avg Speed Bar", 400, 30, "#ff9944",
     '''    <composite x="{x}" y="{y}" width="{w}" height="{h}" name="avg_speed_bar">
        <component type="bar" width="{w}" height="{h}" metric="avg-speed" units="kph" max="{gauge_max}" min="0" outline="255,255,255,128" fill="255,153,68,200"/>
{bar_scale_line}
{chart_title_line}
    </composite>'''),

    ("avg_speed_moving_bar", "Avg Moving Bar", 400, 30, "#ffbb44",
     '''    <composite x="{x}" y="{y}" width="{w}" height="{h}" name="avg_speed_moving_bar">
        <component type="bar" width="{w}" height="{h}" metric="avg-speed-moving" units="kph" max="{gauge_max}" min="0" outline="255,255,255,128" fill="255,187,68,200"/>
{bar_scale_line}
{chart_title_line}
    </composite>'''),

    # -- Heading tape --
    ("heading_tape", "Heading Tape", 400, 60, "#cc6644",
     '''    <composite x="{x}" y="{y}" width="{w}" height="{h}" name="heading_tape">
        <component type="heading-tape" width="{w}" height="{h}" size="{ht_font_size}" tick-interval="{ht_tick_interval}" label-interval="{ht_label_interval}" visible-range="{ht_visible_range}" show-values="{ht_show_values}" show-border="{ht_show_border}" bg="{ht_bg}" fg="{ht_fg}" marker-rgb="{ht_marker}" opacity="{ht_opacity}"/>
    </composite>'''),
]

MAP_STYLES = [
    "osm",
    "cyclosm",
    "tf-cycle",
    "tf-outdoors",
    "tf-landscape",
    "tf-transport",
    "tf-transport-dark",
    "tf-atlas",
    "tf-pioneer",
    "tf-neighbourhood",
    "tf-spinal-map",
    "tf-mobile-atlas",
    "geo-osm-carto",
    "geo-osm-bright",
    "geo-osm-bright-grey",
    "geo-osm-bright-smooth",
    "geo-osm-liberty",
    "geo-positron",
    "geo-positron-blue",
    "geo-positron-red",
    "geo-dark-matter",
    "geo-dark-matter-brown",
    "geo-dark-matter-dark-grey",
    "geo-dark-matter-dark-purple",
    "geo-dark-matter-purple-roads",
    "geo-dark-matter-yellow-roads",
    "geo-toner",
    "geo-toner-grey",
    "geo-klokantech-basic",
    "geo-maptiler-3d",
]

# Map style descriptions and tile URL templates
# Sample tile: zoom=13, x=4823, y=3218 (approx Brisbane area)
MAP_STYLE_INFO = {
    "osm":            ("OpenStreetMap Standard", "free", "http://a.tile.openstreetmap.org/13/4823/3218.png"),
    "cyclosm":        ("CyclOSM - Cycling emphasis", "free", "https://a.tile-cyclosm.openstreetmap.fr/cyclosm/13/4823/3218.png"),
    "tf-cycle":       ("Thunderforest OpenCycleMap", "key:thunderforest", None),
    "tf-outdoors":    ("Thunderforest Outdoors - Hiking/trails", "key:thunderforest", None),
    "tf-landscape":   ("Thunderforest Landscape - Terrain", "key:thunderforest", None),
    "tf-transport":   ("Thunderforest Transport", "key:thunderforest", None),
    "tf-transport-dark": ("Thunderforest Transport Dark", "key:thunderforest", None),
    "tf-atlas":       ("Thunderforest Atlas", "key:thunderforest", None),
    "tf-pioneer":     ("Thunderforest Pioneer - Vintage", "key:thunderforest", None),
    "tf-neighbourhood": ("Thunderforest Neighbourhood", "key:thunderforest", None),
    "tf-spinal-map":  ("Thunderforest Spinal Map", "key:thunderforest", None),
    "tf-mobile-atlas": ("Thunderforest Mobile Atlas", "key:thunderforest", None),
    "geo-osm-carto":  ("Geoapify OSM Carto", "key:geoapify", None),
    "geo-osm-bright": ("Geoapify OSM Bright", "key:geoapify", None),
    "geo-osm-bright-grey": ("Geoapify OSM Bright Grey", "key:geoapify", None),
    "geo-osm-bright-smooth": ("Geoapify OSM Bright Smooth", "key:geoapify", None),
    "geo-osm-liberty": ("Geoapify OSM Liberty", "key:geoapify", None),
    "geo-positron":   ("Geoapify Positron - Light minimal", "key:geoapify", None),
    "geo-positron-blue": ("Geoapify Positron Blue", "key:geoapify", None),
    "geo-positron-red": ("Geoapify Positron Red", "key:geoapify", None),
    "geo-dark-matter": ("Geoapify Dark Matter", "key:geoapify", None),
    "geo-dark-matter-brown": ("Geoapify Dark Matter Brown", "key:geoapify", None),
    "geo-dark-matter-dark-grey": ("Geoapify Dark Matter Grey", "key:geoapify", None),
    "geo-dark-matter-dark-purple": ("Geoapify Dark Matter Purple", "key:geoapify", None),
    "geo-dark-matter-purple-roads": ("Geoapify Dark Purple Roads", "key:geoapify", None),
    "geo-dark-matter-yellow-roads": ("Geoapify Dark Yellow Roads", "key:geoapify", None),
    "geo-toner":      ("Geoapify Toner - B&W high contrast", "key:geoapify", None),
    "geo-toner-grey": ("Geoapify Toner Grey", "key:geoapify", None),
    "geo-klokantech-basic": ("Geoapify Klokantech Basic", "key:geoapify", None),
    "geo-maptiler-3d": ("Geoapify MapTiler 3D", "key:geoapify", None),
}

# Tile coordinate conversion
def _latlon_to_tile(lat: float, lon: float, zoom: int) -> tuple[int, int]:
    """Convert lat/lon to tile x/y coordinates at a given zoom level."""
    n = 2 ** zoom
    x = int((lon + 180.0) / 360.0 * n)
    lat_rad = math.radians(lat)
    y = int((1.0 - math.log(math.tan(lat_rad) + 1.0 / math.cos(lat_rad)) / math.pi) / 2.0 * n)
    return x, y


def _parse_gpx_points(gpx_path: Path) -> list[tuple[float, float, str]]:
    """Parse all track points from a GPX file. Returns list of (lat, lon, time_iso)."""
    import xml.etree.ElementTree as ET
    points = []
    try:
        tree = ET.parse(gpx_path)
        root = tree.getroot()
        ns = "http://www.topografix.com/GPX/1/1"
        for trkpt in root.iter(f"{{{ns}}}trkpt"):
            lat = float(trkpt.get("lat"))
            lon = float(trkpt.get("lon"))
            time_elem = trkpt.find(f"{{{ns}}}time")
            time_str = time_elem.text if time_elem is not None else ""
            points.append((lat, lon, time_str))
        if not points:
            for trkpt in root.iter("trkpt"):
                lat = float(trkpt.get("lat"))
                lon = float(trkpt.get("lon"))
                time_elem = trkpt.find("time")
                time_str = time_elem.text if time_elem is not None else ""
                points.append((lat, lon, time_str))
    except Exception:
        pass
    return points


def _scan_route_ranges(filepath: Path) -> dict:
    """Scan a route file (GPX or XLSX) and return min/max for key metrics.

    Returns dict with keys like 'speed_max', 'battery_max', 'voltage_max', etc.
    Values are in the source file's units (km/h for speed, native for others).
    """
    ranges: dict = {}
    suffix = filepath.suffix.lower()

    try:
        if suffix == ".xlsx":
            import openpyxl
            wb = openpyxl.load_workbook(filepath, read_only=True)
            ws = wb.active
            headers = []
            for row in ws.iter_rows(min_row=1, max_row=1, values_only=True):
                headers = [h.strip() if h else "" for h in row]
            col = {h: i for i, h in enumerate(headers)}

            metrics = {
                "speed": col.get("Speed [km/h]"),
                "battery": col.get("Battery [%]"),
                "voltage": col.get("Voltage [V]"),
                "current": col.get("Current [A]"),
                "power": col.get("Power [W]"),
            }
            vals: dict[str, list] = {k: [] for k in metrics}

            for row in ws.iter_rows(min_row=2, values_only=True):
                for name, idx in metrics.items():
                    if idx is not None and idx < len(row) and row[idx] is not None:
                        try:
                            vals[name].append(float(row[idx]))
                        except (ValueError, TypeError):
                            pass
            wb.close()

            for name, data in vals.items():
                if data:
                    ranges[f"{name}_min"] = min(data)
                    ranges[f"{name}_max"] = max(data)

        elif suffix == ".gpx":
            import xml.etree.ElementTree as ET
            tree = ET.parse(filepath)
            root = tree.getroot()
            ns = "http://www.topografix.com/GPX/1/1"
            euc_ns = "https://euc.world/gpx/1/0"
            tpx_ns = "http://www.garmin.com/xmlschemas/TrackPointExtension/v1"

            speeds, batteries, voltages, currents, powers = [], [], [], [], []

            for trkpt in root.iter(f"{{{ns}}}trkpt"):
                # Speed from extensions or direct child
                for elem in trkpt.iter():
                    tag = elem.tag.split("}")[-1] if "}" in elem.tag else elem.tag
                    if tag == "speed" and elem.text:
                        try:
                            speeds.append(float(elem.text))
                        except ValueError:
                            pass
                    elif tag == "battery" and elem.text:
                        try:
                            batteries.append(float(elem.text))
                        except ValueError:
                            pass
                    elif tag == "voltage" and elem.text:
                        try:
                            voltages.append(float(elem.text))
                        except ValueError:
                            pass
                    elif tag == "current" and elem.text:
                        try:
                            currents.append(float(elem.text))
                        except ValueError:
                            pass
                    elif tag == "power" and elem.text:
                        try:
                            powers.append(float(elem.text))
                        except ValueError:
                            pass

            if speeds:
                # GPX speeds may be m/s or km/h depending on source
                ranges["speed_max"] = max(speeds)
            if batteries:
                ranges["battery_min"] = min(batteries)
                ranges["battery_max"] = max(batteries)
            if voltages:
                ranges["voltage_min"] = min(voltages)
                ranges["voltage_max"] = max(voltages)
            if currents:
                ranges["current_min"] = min(currents)
                ranges["current_max"] = max(currents)
            if powers:
                ranges["power_min"] = min(powers)
                ranges["power_max"] = max(powers)

    except Exception:
        pass

    return ranges


def _parse_gpx_elevations(gpx_path: Path) -> list[tuple[float, float]]:
    """Parse GPX file and return list of (seconds_from_start, elevation_metres)."""
    import xml.etree.ElementTree as ET
    from datetime import datetime, timezone
    points = []
    try:
        tree = ET.parse(gpx_path)
        root = tree.getroot()
        ns = "http://www.topografix.com/GPX/1/1"

        first_time = None
        for trkpt in root.iter(f"{{{ns}}}trkpt"):
            ele_elem = trkpt.find(f"{{{ns}}}ele")
            time_elem = trkpt.find(f"{{{ns}}}time")
            if ele_elem is None or time_elem is None:
                continue
            ele = float(ele_elem.text)
            ts = time_elem.text.replace("Z", "+00:00")
            dt = datetime.fromisoformat(ts)
            if dt.tzinfo is None:
                dt = dt.replace(tzinfo=timezone.utc)
            if first_time is None:
                first_time = dt
            secs = (dt - first_time).total_seconds()
            points.append((secs, ele))

        # Fall back to no-namespace parsing
        if not points:
            for trkpt in root.iter("trkpt"):
                ele_elem = trkpt.find("ele")
                time_elem = trkpt.find("time")
                if ele_elem is None or time_elem is None:
                    continue
                ele = float(ele_elem.text)
                ts = time_elem.text.replace("Z", "+00:00")
                dt = datetime.fromisoformat(ts)
                if dt.tzinfo is None:
                    dt = dt.replace(tzinfo=timezone.utc)
                if first_time is None:
                    first_time = dt
                secs = (dt - first_time).total_seconds()
                points.append((secs, ele))
    except Exception:
        pass
    return points


def _gpx_start_time(gpx_path: Path) -> Optional[datetime]:
    """Get the first timestamp from a GPX file as a datetime."""
    from datetime import datetime, timezone
    import xml.etree.ElementTree as ET
    try:
        tree = ET.parse(gpx_path)
        root = tree.getroot()
        ns = "http://www.topografix.com/GPX/1/1"
        for trkpt in root.iter(f"{{{ns}}}trkpt"):
            time_elem = trkpt.find(f"{{{ns}}}time")
            if time_elem is not None:
                ts = time_elem.text.replace("Z", "+00:00")
                dt = datetime.fromisoformat(ts)
                if dt.tzinfo is None:
                    dt = dt.replace(tzinfo=timezone.utc)
                return dt
        for trkpt in root.iter("trkpt"):
            time_elem = trkpt.find("time")
            if time_elem is not None:
                ts = time_elem.text.replace("Z", "+00:00")
                dt = datetime.fromisoformat(ts)
                if dt.tzinfo is None:
                    dt = dt.replace(tzinfo=timezone.utc)
                return dt
    except Exception:
        pass
    return None


def _extract_gpx_first_point(gpx_path: Path) -> Optional[tuple[float, float]]:
    """Extract the first lat/lon from a GPX file."""
    points = _parse_gpx_points(gpx_path)
    if points:
        return (points[0][0], points[0][1])
    return None


def _find_gpx_point_at_time(gpx_path: Path, video_creation_time: str) -> Optional[tuple[float, float]]:
    """Find the GPX point closest to a video's creation time."""
    from datetime import datetime, timezone
    points = _parse_gpx_points(gpx_path)
    if not points:
        return None

    # Parse video creation time
    try:
        # Handle formats like "2026-04-03T07:40:50.000000Z"
        vt = video_creation_time.replace("Z", "+00:00")
        video_dt = datetime.fromisoformat(vt)
        if video_dt.tzinfo is None:
            video_dt = video_dt.replace(tzinfo=timezone.utc)
    except Exception:
        # Can't parse video time, return first point
        return (points[0][0], points[0][1])

    # Find closest point by time
    best = None
    best_diff = None
    for lat, lon, time_str in points:
        if not time_str:
            continue
        try:
            pt = time_str.replace("Z", "+00:00")
            point_dt = datetime.fromisoformat(pt)
            if point_dt.tzinfo is None:
                point_dt = point_dt.replace(tzinfo=timezone.utc)
            diff = abs((point_dt - video_dt).total_seconds())
            if best_diff is None or diff < best_diff:
                best_diff = diff
                best = (lat, lon)
        except Exception:
            continue

    return best if best else (points[0][0], points[0][1])


# Map tile URL templates by provider
_TF_STYLES = {
    "tf-cycle": "cycle", "tf-outdoors": "outdoors", "tf-landscape": "landscape",
    "tf-transport": "transport", "tf-transport-dark": "transport-dark",
    "tf-atlas": "atlas", "tf-pioneer": "pioneer", "tf-neighbourhood": "neighbourhood",
    "tf-spinal-map": "spinal-map", "tf-mobile-atlas": "mobile-atlas",
}
_GEO_STYLES = {
    "geo-osm-carto": "osm-carto", "geo-osm-bright": "osm-bright",
    "geo-osm-bright-grey": "osm-bright-grey", "geo-osm-bright-smooth": "osm-bright-smooth",
    "geo-osm-liberty": "osm-liberty", "geo-positron": "positron",
    "geo-positron-blue": "positron-blue", "geo-positron-red": "positron-red",
    "geo-dark-matter": "dark-matter", "geo-dark-matter-brown": "dark-matter-brown",
    "geo-dark-matter-dark-grey": "dark-matter-dark-grey",
    "geo-dark-matter-dark-purple": "dark-matter-dark-purple",
    "geo-dark-matter-purple-roads": "dark-matter-purple-roads",
    "geo-dark-matter-yellow-roads": "dark-matter-yellow-roads",
    "geo-toner": "toner", "geo-toner-grey": "toner-grey",
    "geo-klokantech-basic": "klokantech-basic", "geo-maptiler-3d": "maptiler-3d",
}


def _get_api_keys() -> dict[str, str]:
    """Load API keys from config file and environment."""
    import json as _json
    keys = {}
    keys_file = Path.home() / ".gopro-graphics" / "map-api-keys.json"
    if keys_file.exists():
        try:
            keys = _json.loads(keys_file.read_text())
        except Exception:
            pass
    for ref in ("thunderforest", "geoapify"):
        env_val = os.environ.get(f"API_KEY_{ref.upper()}", "")
        if env_val:
            keys[ref] = env_val
    return keys


def _build_tile_url(style: str, lat: float, lon: float, zoom: int = 14) -> Optional[str]:
    """Build a tile preview URL for a given style centered on lat/lon."""
    tx, ty = _latlon_to_tile(lat, lon, zoom)
    keys = _get_api_keys()

    if style == "osm":
        return f"http://a.tile.openstreetmap.org/{zoom}/{tx}/{ty}.png"
    elif style == "cyclosm":
        return f"https://a.tile-cyclosm.openstreetmap.fr/cyclosm/{zoom}/{tx}/{ty}.png"
    elif style in _TF_STYLES:
        key = keys.get("thunderforest", "")
        if key:
            return f"https://a.tile.thunderforest.com/{_TF_STYLES[style]}/{zoom}/{tx}/{ty}.png?apikey={key}"
    elif style in _GEO_STYLES:
        key = keys.get("geoapify", "")
        if key:
            return f"https://maps.geoapify.com/v1/tile/{_GEO_STYLES[style]}/{zoom}/{tx}/{ty}.png?apiKey={key}"
    return None

# Date format presets (date only)
DATE_FORMATS = [
    "%Y/%m/%d",
    "%d %b %Y",
    "%d %B %Y",
    "%d/%m/%Y",
    "%m/%d/%Y",
    "%Y-%m-%d",
    "%a %d %b %Y",
    "%a %d %b",
    "%d %b",
]

# Time format presets (time only)
TIME_FORMATS = [
    "%H:%M:%S.%f",
    "%H:%M:%S",
    "%H:%M",
    "%I:%M:%S %p",
    "%I:%M %p",
]


# ---------------------------------------------------------------------------
# Video analysis via ffprobe
# ---------------------------------------------------------------------------

def analyse_video(video_path: Path) -> VideoEntry:
    """Analyse a video file using ffprobe, returning dimensions, rotation, duration."""
    cmd = [
        "ffprobe", "-v", "quiet", "-print_format", "json",
        "-show_format", "-show_streams", "-show_entries",
        "stream=width,height,duration,codec_type:stream_side_data=rotation:format=duration:format_tags=creation_time",
        str(video_path)
    ]
    result = subprocess.run(cmd, capture_output=True, text=True)
    data = json.loads(result.stdout)

    raw_w = raw_h = 0
    rotation = 0
    duration = 0.0
    creation_time = ""

    for stream in data.get("streams", []):
        if stream.get("codec_type") == "video" and raw_w == 0:
            raw_w = int(stream.get("width", 0))
            raw_h = int(stream.get("height", 0))
            if "side_data_list" in stream:
                for sd in stream["side_data_list"]:
                    if "rotation" in sd:
                        rotation = int(float(sd["rotation"]))

    fmt = data.get("format", {})
    duration = float(fmt.get("duration", 0))
    tags = fmt.get("tags", {})
    creation_time = tags.get("creation_time", "")

    # Effective dimensions after rotation
    if rotation in (-90, 90, -270, 270):
        eff_w, eff_h = raw_h, raw_w
    else:
        eff_w, eff_h = raw_w, raw_h

    return VideoEntry(
        path=video_path,
        raw_width=raw_w, raw_height=raw_h,
        rotation=rotation,
        eff_width=eff_w, eff_height=eff_h,
        duration_seconds=duration,
        creation_time=creation_time,
    )


def extract_frame(video_path: Path, time_seconds: float, eff_w: int, eff_h: int) -> Optional[Image.Image]:
    """Extract a single frame from a video at a given timestamp, auto-rotating."""
    cmd = [
        "ffmpeg", "-hide_banner", "-loglevel", "error",
        "-ss", str(time_seconds),
        "-i", str(video_path),
        "-frames:v", "1",
        "-vf", f"scale={eff_w}:{eff_h}",
        "-f", "rawvideo", "-pix_fmt", "rgba",
        "-"
    ]
    result = subprocess.run(cmd, capture_output=True, timeout=15)
    expected = eff_w * eff_h * 4
    if len(result.stdout) != expected:
        return None
    return Image.frombytes("RGBA", (eff_w, eff_h), result.stdout)


def detect_gpu() -> Optional[str]:
    """Check for NVIDIA GPU encoding support."""
    try:
        subprocess.run(["nvidia-smi"], capture_output=True, check=True)
        result = subprocess.run(["ffmpeg", "-hide_banner", "-encoders"],
                                capture_output=True, text=True)
        if "h264_nvenc" in result.stdout:
            return "nvgpu"
    except (FileNotFoundError, subprocess.CalledProcessError):
        pass
    return None


def detect_font() -> str:
    """Find a usable font."""
    candidates = [
        "/usr/share/fonts/TTF/DejaVuSans.ttf",
        "/usr/share/fonts/truetype/dejavu/DejaVuSans.ttf",
        "/usr/share/fonts/TTF/Roboto-Medium.ttf",
        "/usr/share/fonts/truetype/roboto/Roboto-Medium.ttf",
        "/usr/share/fonts/liberation/LiberationSans-Regular.ttf",
        "/usr/share/fonts/truetype/liberation/LiberationSans-Regular.ttf",
    ]
    for c in candidates:
        if os.path.isfile(c):
            return c
    try:
        result = subprocess.run(["fc-match", "DejaVu Sans", "--format=%{file}"],
                                capture_output=True, text=True)
        if os.path.isfile(result.stdout.strip()):
            return result.stdout.strip()
    except FileNotFoundError:
        pass
    return "Roboto-Medium.ttf"


# ---------------------------------------------------------------------------
# Default component positions for a given resolution
# ---------------------------------------------------------------------------

def default_component_positions(eff_w: int, eff_h: int) -> list[OverlayComponent]:
    """Create overlay components with sensible default positions for the resolution."""
    # Scale component sizes relative to 1920x1080 baseline
    sx = eff_w / 1920.0
    sy = eff_h / 1080.0
    s = min(sx, sy)

    right_x = eff_w - int(280 * s)
    bottom_y = eff_h - int(100 * s)
    speed_y = eff_h - int(280 * s)

    defaults = {
        "date_and_time": (int(260 * s), 30),
        "date_only": (int(260 * s), 30),
        "time_only": (int(260 * s), 30),
        "gps_info": (right_x, 0),
        "gps_coords": (right_x, 0),
        "gps_lock": (right_x + int(226 * s), int(24 * s)),
        "big_mph": (16, speed_y),
        "gradient_chart": (int(400 * s), bottom_y),
        "gradient": (int(220 * s), bottom_y),
        "altitude": (16, bottom_y),
        "temperature": (eff_w - 20, speed_y),
        "cadence": (eff_w - 20, speed_y + int(80 * s)),
        "heartbeat": (eff_w - 20, speed_y + int(160 * s)),
        "moving_map": (right_x, int(100 * s)),
        "journey_map": (right_x, int(100 * s) + int(276 * s)),
        # EUC telemetry defaults
        "battery_value": (16, int(100 * s)),
        "voltage_value": (16, int(180 * s)),
        "current_value": (16, int(260 * s)),
        "power_value": (16, int(340 * s)),
        "battery_chart": (int(200 * s), int(100 * s)),
        "voltage_chart": (int(200 * s), int(190 * s)),
        "power_chart": (int(200 * s), int(280 * s)),
        # Gauges
        "speed_gauge": (int(400 * s), int(200 * s)),
        "battery_gauge": (int(400 * s), int(480 * s)),
        "power_gauge": (int(680 * s), int(200 * s)),
        "voltage_gauge": (int(680 * s), int(480 * s)),
        "compass_display": (right_x, int(600 * s)),
        "compass_arrow_display": (right_x, int(600 * s)),
        "speed_bar": (16, speed_y + int(180 * s)),
        "battery_bar": (16, speed_y + int(220 * s)),
        "moving_journey_map": (right_x, int(100 * s) + int(276 * s)),
        "circuit_map": (right_x, int(600 * s)),
        "asi_gauge": (int(400 * s), int(200 * s)),
        "msi_gauge": (int(400 * s), int(200 * s)),
        "avg_speed_value": (16, int(420 * s)),
        "avg_speed_moving_value": (16, int(500 * s)),
        "avg_speed_gauge": (int(400 * s), int(760 * s)),
        "avg_speed_moving_gauge": (int(680 * s), int(760 * s)),
        "avg_speed_bar": (16, speed_y + int(260 * s)),
        "avg_speed_moving_bar": (16, speed_y + int(300 * s)),
        "heading_tape": ((eff_w - 400) // 2, 30),
    }

    # Default enabled/disabled
    disabled = {"heartbeat", "temperature", "cadence", "date_only", "time_only",
                 "gps_coords", "gps_lock",
                 "battery_value", "voltage_value", "current_value", "power_value",
                 "battery_chart", "voltage_chart", "power_chart",
                 "speed_gauge", "battery_gauge", "power_gauge", "voltage_gauge",
                 "compass_display", "compass_arrow_display",
                 "speed_bar", "battery_bar",
                 "moving_journey_map", "circuit_map", "asi_gauge", "msi_gauge",
                 "avg_speed_value", "avg_speed_moving_value",
                 "avg_speed_gauge", "avg_speed_moving_gauge",
                 "avg_speed_bar", "avg_speed_moving_bar",
                 "heading_tape"}

    components = []
    for name, label, ref_w, ref_h, color, xml_tmpl in COMPONENT_DEFS:
        # Use ref_w/ref_h as actual pixel sizes — they match the fixed pixel
        # values in the XML templates (font sizes, icon sizes, map sizes).
        # Don't scale by s; scaling is only for default positions.
        w = ref_w
        h = ref_h
        x, y = defaults.get(name, (0, 0))
        components.append(OverlayComponent(
            name=name, label=label,
            x=x, y=y, width=w, height=h,
            enabled=name not in disabled,
            color=color,
            xml_template=xml_tmpl,
        ))
    return components


# ---------------------------------------------------------------------------
# XML layout generation
# ---------------------------------------------------------------------------

def load_layout_xml(xml_path: Path, components: list[OverlayComponent]):
    """Load component positions and sizes from a layout XML file into existing components."""
    import xml.etree.ElementTree as ET
    tree = ET.parse(xml_path)
    root = tree.getroot()

    # Map XML attribute names to template variable names
    _XML_TO_PROP = {
        # Chart attributes
        "filled": "chart_filled", "seconds": "chart_seconds",
        # Gauge attributes
        "max": "gauge_max", "min": "gauge_min",
        "start": "gauge_start", "length": "gauge_length",
        "sectors": "gauge_sectors",
        "background-rgb": "gauge_bg", "needle-rgb": "gauge_needle",
        "major-tick-rgb": "gauge_tick", "minor-tick-rgb": "gauge_tick",
        "major-ann-rgb": "gauge_ann", "minor-ann-rgb": "gauge_ann",
        "tick-rgb": "gauge_tick", "gauge-rgb": "gauge_fill",
        "arc-inner-rgb": "regen_colour",
        # MSI attributes
        "green": "msi_green", "yellow": "msi_yellow",
        "end": "msi_end", "rotate": "msi_rotate",
        "textsize": "msi_textsize",
        # Map attributes
        "zoom": "map_zoom", "corner_radius": "map_corner_radius",
        "opacity": "map_opacity",
        # Heading tape
        "tick-interval": "ht_tick_interval",
        # Text colours
        "rgb": "value_rgb",
        # Title
        "align": None,  # skip — internal positioning
    }
    # Attributes to skip (used for positioning, not properties)
    _SKIP = {"x", "y", "type", "name", "metric", "units", "dp", "format",
             "truncate", "cache", "file", "width", "height", "size", "align"}

    comp_data: dict[str, dict] = {}

    for elem in root:
        name = elem.get("name", "")
        if not name:
            continue
        data = {
            "x": int(elem.get("x", "0")),
            "y": int(elem.get("y", "0")),
        }

        # Extract size from the element itself
        if elem.get("size"):
            s = int(elem.get("size"))
            data["width"] = s
            data["height"] = s
        if elem.get("width"):
            data["width"] = int(elem.get("width"))
        if elem.get("height"):
            data["height"] = int(elem.get("height"))

        # Scan children and extract all attributes as custom_props
        for child in elem:
            child_type = child.get("type", "")

            # Chart: width/height from chart component (only if composite didn't set them)
            if child_type == "chart":
                if child.get("width") and "width" not in data:
                    data["width"] = int(child.get("width"))
                if child.get("height") and "height" not in data:
                    data["height"] = int(child.get("height"))

            # Gauge style from type attribute
            if child_type.startswith("cairo-gauge"):
                data.setdefault("custom_props", {})["gauge_style"] = child_type

            # Map/gauge component size — only set width/height if composite
            # didn't already specify them (composite dimensions are the preview size)
            if child_type in ("moving_map", "journey_map", "moving-journey-map",
                              "compass", "compass-arrow", "cairo-circuit-map",
                              "asi", "msi2") and child.get("size"):
                s = int(child.get("size"))
                if "width" not in data:
                    data["width"] = s
                if "height" not in data:
                    data["height"] = s
                data.setdefault("custom_props", {})["gauge_size"] = s

            # Speed font size
            if child_type == "metric" and child.get("size"):
                data.setdefault("custom_props", {})["speed_font_size"] = int(child.get("size"))
                data.setdefault("custom_props", {})["value_font_size"] = int(child.get("size"))

            # Label font size
            if child_type in ("text", "datetime", "metric_unit") and child.get("size"):
                data.setdefault("custom_props", {})["text_size"] = int(child.get("size"))

            # Datetime format strings
            if child_type == "datetime" and child.get("format"):
                fmt = child.get("format")
                if name == "time_only":
                    # time_only has a single datetime child — it's the time format
                    data.setdefault("custom_props", {})["time_format"] = fmt
                elif "date_format" not in data.get("custom_props", {}):
                    # First datetime in composite = date
                    data.setdefault("custom_props", {})["date_format"] = fmt
                else:
                    # Second datetime = time
                    data.setdefault("custom_props", {})["time_format"] = fmt
                # Also extract size for date/time
                if child.get("size"):
                    s = int(child.get("size"))
                    if name == "time_only":
                        data.setdefault("custom_props", {})["time_size"] = s
                    elif "date_size" not in data.get("custom_props", {}):
                        data.setdefault("custom_props", {})["date_size"] = s
                    else:
                        data.setdefault("custom_props", {})["time_size"] = s
                # Extract align
                if child.get("align"):
                    data.setdefault("custom_props", {})["dt_align"] = child.get("align")

            # Title text content
            if child_type == "text" and child.text and child.get("align") in ("center", "centre"):
                data.setdefault("custom_props", {})["comp_title"] = child.text
                data.setdefault("custom_props", {})["show_title"] = "true"
                if child.get("size"):
                    data["custom_props"]["title_size"] = int(child.get("size"))
                if child.get("rgb"):
                    data["custom_props"]["title_rgb"] = child.get("rgb")
            elif child_type == "text" and child.text and child.get("x") == "4" and child.get("y") == "2":
                data.setdefault("custom_props", {})["comp_title"] = child.text
                data.setdefault("custom_props", {})["show_title"] = "true"
                if child.get("size"):
                    data["custom_props"]["title_size"] = int(child.get("size"))
                if child.get("rgb"):
                    data["custom_props"]["title_rgb"] = child.get("rgb")

            # Extract all mapped attributes from children
            for attr, val in child.attrib.items():
                if attr in _SKIP:
                    continue
                prop_name = _XML_TO_PROP.get(attr)
                if prop_name is None:
                    continue
                # Try to parse as int, fall back to string
                try:
                    val = int(val)
                except (ValueError, TypeError):
                    try:
                        val = float(val)
                    except (ValueError, TypeError):
                        pass
                data.setdefault("custom_props", {})[prop_name] = val

            # Chart-specific: prefix with chart_
            if child_type == "chart":
                for attr in ("filled", "seconds", "bg", "fill", "line"):
                    val = child.get(attr)
                    if val is not None:
                        data.setdefault("custom_props", {})[f"chart_{attr}"] = val

            # Colour attributes from text components
            if child_type == "text" and child.get("rgb"):
                # Only set label_rgb from non-title text components
                if child.text and child.text not in (
                    data.get("custom_props", {}).get("comp_title", ""),
                ):
                    data.setdefault("custom_props", {})["label_rgb"] = child.get("rgb")

        comp_data[name] = data

    # Apply parsed data to components
    for comp in components:
        if comp.name in comp_data:
            d = comp_data[comp.name]
            comp.x = d["x"]
            comp.y = d["y"]
            if "width" in d:
                comp.width = d["width"]
            if "height" in d:
                comp.height = d["height"]
            for k, v in d.get("custom_props", {}).items():
                comp.custom_props[k] = v
            comp.enabled = True
        else:
            comp.enabled = False


def _auto_font_sizes(comp: OverlayComponent) -> dict:
    """Derive font sizes from component dimensions for text-based components.

    Returns a dict of template variable overrides. Components that don't
    have scalable text return an empty dict, preserving manual settings.
    """
    h = comp.height
    w = comp.width
    name = comp.name

    _gauge_names = {"speed_gauge", "battery_gauge", "power_gauge", "voltage_gauge",
                     "compass_display", "compass_arrow_display", "asi_gauge", "msi_gauge",
                     "circuit_map", "avg_speed_gauge", "avg_speed_moving_gauge"}
    if name in _gauge_names:
        size = min(w, h)
        return {"gauge_size": size,
                "title_x": size // 2,
                "title_gauge_y": int(size * 0.82)}

    if name == "big_mph":
        # Speed number fills the component height; unit label is small
        return {"speed_font_size": h, "text_size": max(8, h // 10)}
    elif name in ("gradient", "altitude", "temperature", "cadence", "heartbeat",
                   "battery_value", "voltage_value", "current_value", "power_value",
                   "avg_speed_value", "avg_speed_moving_value"):
        # Icon is ~91% of height; value text ~46%, label ~23%
        icon_size = max(16, int(h * 0.91))
        text_size = max(8, int(h * 0.23))
        value_font_size = max(12, int(h * 0.46))
        value_y = text_size + value_font_size // 2
        return {"value_font_size": value_font_size,
                "text_size": text_size,
                "icon_size": icon_size,
                "value_y": value_y}
    elif name == "date_and_time":
        # Two lines: date ~33% of height, time ~67%
        date_size = max(8, int(h * 0.33))
        time_size = max(8, int(h * 0.67))
        time_y = date_size + max(2, date_size // 4)  # gap after date line
        return {"date_size": date_size, "time_size": time_size, "time_y": time_y}
    elif name == "date_only":
        return {"date_size": max(8, h)}
    elif name == "time_only":
        return {"time_size": max(8, int(h * 0.89))}

    return {}


def _auto_gauge_ranges(comp: OverlayComponent, route_ranges: dict) -> dict:
    """Derive gauge max/min from route data for each gauge component.

    Returns template variable overrides. Only sets values not already in
    custom_props (manual overrides take precedence).
    """
    import math
    name = comp.name

    def _round_up(val, step):
        return int(math.ceil(val / step) * step)

    def _round_down(val, step):
        return int(math.floor(val / step) * step)

    if name in ("speed_gauge", "speed_bar", "avg_speed_gauge", "avg_speed_bar",
                 "avg_speed_moving_gauge", "avg_speed_moving_bar"):
        max_spd = route_ranges.get("speed_max", 60)
        return {"gauge_max": _round_up(max_spd * 1.05, 10), "gauge_min": 0}

    elif name in ("battery_gauge", "battery_bar"):
        return {"gauge_max": 100, "gauge_min": 0}

    elif name == "voltage_gauge":
        v_max = route_ranges.get("voltage_max", 100)
        v_min = route_ranges.get("voltage_min", 60)
        return {"gauge_max": _round_up(v_max * 1.05, 5),
                "gauge_min": _round_down(v_min * 0.95, 5)}

    elif name == "power_gauge":
        p_max = route_ranges.get("power_max", 2000)
        p_min = route_ranges.get("power_min", -500)
        return {"gauge_style": "cairo-gauge-arc-annotated",
                "gauge_max": _round_up(max(p_max, 100), 100),
                "gauge_min": _round_down(min(p_min, 0), 100)}

    return {}


def _gauge_colour_attrs(comp: OverlayComponent, fmt_vars: dict) -> str:
    """Build the colour XML attributes string for the selected gauge style."""
    style = fmt_vars.get("gauge_style", "cairo-gauge-round-annotated")
    bg = fmt_vars.get("gauge_bg", "255,255,255,150")
    needle = fmt_vars.get("gauge_needle", "255,0,0")
    tick = fmt_vars.get("gauge_tick", "0,0,0")
    ann = fmt_vars.get("gauge_ann", "0,0,0")
    fill = fmt_vars.get("gauge_fill", "0,191,255")
    regen = fmt_vars.get("regen_colour", "68,187,68,180")
    gauge_min = fmt_vars.get("gauge_min", 0)

    # Arc/regen zone attributes — only for styles that support them
    arc_attrs = ""
    if comp.name == "power_gauge" and style in ("cairo-gauge-arc-annotated", "cairo-gauge-donut"):
        arc_attrs = (f' arc-value-lower="{gauge_min}" arc-value-upper="0"'
                     f' arc-inner-rgb="{regen}" arc-outer-rgb="{regen}"')

    if style == "cairo-gauge-marker":
        return (f' tick-rgb="{tick}" background-rgb="{bg}"'
                f' gauge-rgb="{fill}"')
    elif style == "cairo-gauge-donut":
        return (f' needle-rgb="{needle}"'
                f' major-tick-rgb="{tick}" minor-tick-rgb="{tick}"'
                f' major-ann-rgb="{ann}" minor-ann-rgb="{ann}"'
                f'{arc_attrs}')
    elif style == "cairo-gauge-arc-annotated":
        return (f' background-rgb="{bg}" needle-rgb="{needle}"'
                f' major-tick-rgb="{tick}" minor-tick-rgb="{tick}"'
                f' major-ann-rgb="{ann}" minor-ann-rgb="{ann}"'
                f'{arc_attrs}')
    else:
        # round-annotated — no arc attributes
        return (f' background-rgb="{bg}" needle-rgb="{needle}"'
                f' major-tick-rgb="{tick}" minor-tick-rgb="{tick}"'
                f' major-ann-rgb="{ann}" minor-ann-rgb="{ann}"')


def generate_layout_xml(components: list[OverlayComponent], map_props: dict,
                        route_ranges: Optional[dict] = None) -> str:
    """Generate layout XML from current component positions."""
    if route_ranges is None:
        route_ranges = {}
    lines = ["<layout>"]
    for comp in components:
        if comp.enabled:
            # Merge global map_props with per-component overrides
            fmt_vars = dict(map_props)
            fmt_vars["x"] = comp.x
            fmt_vars["y"] = comp.y
            fmt_vars["w"] = comp.width
            fmt_vars["h"] = comp.height
            # Auto-derive font sizes from component dimensions
            fmt_vars.update(_auto_font_sizes(comp))
            # Auto-derive gauge ranges from route data
            fmt_vars.update(_auto_gauge_ranges(comp, route_ranges))
            # Per-component overrides (manual settings take precedence)
            for k, v in comp.custom_props.items():
                fmt_vars[k] = v
            # Only truncate time output when format includes microseconds (%f)
            time_fmt = fmt_vars.get("time_format", "")
            fmt_vars["time_truncate"] = 'truncate="5" ' if "%f" in time_fmt else ""
            # Build gauge colour attributes based on selected style
            fmt_vars["gauge_colour_attrs"] = _gauge_colour_attrs(comp, fmt_vars)
            # Build title lines — only when show_title is enabled and title text is set
            show_title = str(fmt_vars.get("show_title", "false")).lower() in ("true", "1", "yes")
            title = str(fmt_vars.get("comp_title", "")).strip()
            if show_title and title:
                ts = fmt_vars.get("title_size", 14)
                tr = fmt_vars.get("title_rgb", "255,255,255")
                tx = fmt_vars.get("title_x", 128)
                tgy = fmt_vars.get("title_gauge_y", 200)
                fmt_vars["chart_title_line"] = f'        <component type="text" x="4" y="2" size="{ts}" rgb="{tr}">{title}</component>'
                fmt_vars["gauge_title_line"] = f'        <component type="text" x="{tx}" y="{tgy}" size="{ts}" rgb="{tr}" align="centre">{title}</component>'
            else:
                fmt_vars["chart_title_line"] = ""
                fmt_vars["gauge_title_line"] = ""
            # Build bar scale labels — min/max at ends of bar
            show_scale = str(fmt_vars.get("show_bar_scale", "true")).lower() in ("true", "1", "yes")
            if show_scale and comp.name.endswith("_bar"):
                bw = comp.width
                bh = comp.height
                g_max = fmt_vars.get("gauge_max", 100)
                scale_size = max(8, min(14, bh - 4))
                sr = fmt_vars.get("title_rgb", "255,255,255")
                fmt_vars["bar_scale_line"] = (
                    f'        <component type="text" x="2" y="0" size="{scale_size}" rgb="{sr}">0</component>\n'
                    f'        <component type="text" x="{bw - 2}" y="0" size="{scale_size}" rgb="{sr}" align="right">{g_max}</component>'
                )
            else:
                fmt_vars["bar_scale_line"] = ""
            xml = comp.xml_template.format(**fmt_vars)
            lines.append(xml)
            lines.append("")
    lines.append("</layout>")
    return "\n".join(lines)


# ---------------------------------------------------------------------------
# Shell script generation
# ---------------------------------------------------------------------------

def generate_shell_script(
    videos: list[VideoEntry],
    gpx_path: Path,
    layout_xml: str,
    speed_unit: str,
    font: str,
    gpu_profile: Optional[str],
    dashboard_script: Path,
    map_style: str = "osm",
    gpx_time_offset: float = 0.0,
    sample_duration: Optional[float] = None,
) -> str:
    """Generate a self-contained bash script with embedded layout XML."""
    lines = ["#!/usr/bin/env bash", "set -euo pipefail", ""]

    # Write layout XML to a temp file and clean up on exit
    lines.append("# Embedded layout XML")
    lines.append('LAYOUT_XML="$(mktemp /tmp/layout-XXXXXX.xml)"')
    lines.append('trap \'rm -f "$LAYOUT_XML"\' EXIT')
    lines.append(f'cat > "$LAYOUT_XML" << \'__LAYOUT_EOF__\'')
    lines.append(layout_xml)
    lines.append("__LAYOUT_EOF__")
    lines.append("")

    for video in videos:
        out_name = video.path.stem + "_overlay" + video.path.suffix
        out_path = video.path.parent / out_name

        # Use --xlsx or --gpx depending on input file type
        gpx_arg = "--xlsx" if gpx_path.suffix.lower() == ".xlsx" else "--gpx"

        cmd_parts = [
            f'python "{dashboard_script}"',
            f'--font "{font}"',
            f'{gpx_arg} "{gpx_path}"',
            "--use-gpx-only",
            "--video-time-start video-created",
            f"--overlay-size {video.eff_width}x{video.eff_height}",
            f"--units-speed {speed_unit}",
            "--layout xml",
            '--layout-xml "$LAYOUT_XML"',
            f"--map-style {map_style}",
        ]
        if gpx_time_offset:
            cmd_parts.append(f"--gpx-time-offset {gpx_time_offset}")
        if sample_duration:
            cmd_parts.append(f"--sample-duration {sample_duration}")
        if gpu_profile:
            cmd_parts.append(f"--profile {gpu_profile}")
        cmd_parts.append(f'"{video.path}"')
        cmd_parts.append(f'"{out_path}"')

        lines.append(f'echo "Processing {video.basename}..."')
        lines.append(" \\\n    ".join(cmd_parts))
        lines.append("")

    lines.append('echo "Done."')
    return "\n".join(lines)


# ---------------------------------------------------------------------------
# Main GUI Application
# ---------------------------------------------------------------------------

class LayoutEditorApp(tk.Tk):

    CANVAS_MAX_W = 960
    CANVAS_MAX_H = 800

    def __init__(self):
        super().__init__()
        self.title("GoPro Dashboard — Layout Editor")
        self.geometry("1400x900")

        self.gpx_path: Optional[Path] = None
        self.gpx_location: Optional[tuple[float, float]] = None  # (lat, lon) from GPX
        self.route_ranges: dict = {}  # metric ranges from route data
        self.videos: list[VideoEntry] = []
        self.components: list[OverlayComponent] = []
        self.active_video_idx: int = -1
        self.scale_factor: float = 1.0
        self.current_photo: Optional[ImageTk.PhotoImage] = None
        self.speed_unit = tk.StringVar(value="kph")
        self.map_style = tk.StringVar(value="osm")
        self.map_size = tk.IntVar(value=256)
        self.map_zoom = tk.IntVar(value=16)
        self.map_corner_radius = tk.IntVar(value=35)
        self.map_opacity = tk.DoubleVar(value=0.7)
        self.map_rotate = tk.BooleanVar(value=True)  # rotate map to heading (north pointer)
        # Journey map specific
        self.jm_track_colour = tk.StringVar(value="255,0,0")
        self.jm_line_width = tk.IntVar(value=4)
        self.jm_loc_colour = tk.StringVar(value="0,0,255")
        self.jm_loc_outline = tk.StringVar(value="0,0,0")
        self.jm_loc_size = tk.IntVar(value=6)
        self.date_format = tk.StringVar(value="%Y/%m/%d")
        self.time_format = tk.StringVar(value="%H:%M:%S.%f")
        self.gpx_time_offset = tk.DoubleVar(value=0.0)
        self.sample_duration = tk.StringVar(value="")  # empty = full video
        self.gpu_profile = detect_gpu()
        self.use_gpu = tk.BooleanVar(value=self.gpu_profile is not None)
        self.font_path = detect_font()
        self.workdir: Optional[str] = None  # initial directory for file dialogs
        self.snap_enabled = tk.BooleanVar(value=False)
        self.snap_size = tk.IntVar(value=20)
        self.frame_queue: queue.Queue = queue.Queue()

        # Drag state
        self._drag_component: Optional[OverlayComponent] = None
        self._drag_offset_x = 0
        self._drag_offset_y = 0
        self._resize_component: Optional[OverlayComponent] = None
        self._resize_edge: str = ""  # "br", "r", "b" etc.

        # Scrub debounce
        self._scrub_after_id = None

        self._script_dir = Path(__file__).resolve().parent
        self._dashboard_script = self._script_dir / "gopro-dashboard.py"

        self._build_ui()

    def _build_ui(self):
        # Menu bar
        menubar = tk.Menu(self)
        file_menu = tk.Menu(menubar, tearoff=0)
        file_menu.add_command(label="Open Route File...", command=self._open_gpx, accelerator="Ctrl+G")
        file_menu.add_command(label="Clear Route File", command=self._clear_gpx)
        file_menu.add_separator()
        file_menu.add_command(label="Add Video Files...", command=self._add_videos, accelerator="Ctrl+O")
        file_menu.add_command(label="Remove Selected Video", command=self._remove_video, accelerator="Delete")
        file_menu.add_command(label="Clear All Videos", command=self._clear_videos)
        file_menu.add_separator()
        file_menu.add_command(label="Load Layout XML...", command=self._load_layout_xml)
        file_menu.add_command(label="Save Layout XML...", command=self._save_layout_xml)
        file_menu.add_command(label="Export Shell Script...", command=self._export_shell_script)
        file_menu.add_separator()
        file_menu.add_command(label="Select Font...", command=self._select_font)
        file_menu.add_separator()
        file_menu.add_command(label="Quit", command=self.quit, accelerator="Ctrl+Q")
        menubar.add_cascade(label="File", menu=file_menu)

        settings_menu = tk.Menu(menubar, tearoff=0)
        settings_menu.add_command(label="GPX/Video Time Sync...", command=self._show_sync_dialog)
        settings_menu.add_separator()
        settings_menu.add_command(label="API Keys...", command=self._show_api_keys_dialog)
        menubar.add_cascade(label="Settings", menu=settings_menu)

        self.config(menu=menubar)

        self.bind_all("<Control-g>", lambda e: self._open_gpx())
        self.bind_all("<Control-o>", lambda e: self._add_videos())
        self.bind_all("<Control-q>", lambda e: self.quit())
        # Delete key disabled — was accidentally clearing videos

        # Main layout: left panel | canvas + slider | right panel
        main_pane = ttk.PanedWindow(self, orient=tk.HORIZONTAL)
        main_pane.pack(fill=tk.BOTH, expand=True, padx=4, pady=4)

        # -- Left panel: videos + components --
        left_frame = ttk.Frame(main_pane, width=220)
        main_pane.add(left_frame, weight=0)

        # GPX display
        gpx_frame = ttk.LabelFrame(left_frame, text="Route File")
        gpx_frame.pack(fill=tk.X, padx=4, pady=2)
        self.gpx_label = ttk.Label(gpx_frame, text="(none)", wraplength=200)
        self.gpx_label.pack(padx=4, pady=2)
        gpx_btn_frame = ttk.Frame(gpx_frame)
        gpx_btn_frame.pack(padx=4, pady=2)
        ttk.Button(gpx_btn_frame, text="Browse...", command=self._open_gpx).pack(side=tk.LEFT, padx=2)
        ttk.Button(gpx_btn_frame, text="Clear", command=self._clear_gpx).pack(side=tk.LEFT, padx=2)

        # Video list
        vid_frame = ttk.LabelFrame(left_frame, text="Videos")
        vid_frame.pack(fill=tk.BOTH, expand=True, padx=4, pady=2)
        self.video_listbox = tk.Listbox(vid_frame, height=6, exportselection=False)
        self.video_listbox.pack(fill=tk.BOTH, expand=True, padx=4, pady=2)
        self.video_listbox.bind("<<ListboxSelect>>", self._on_video_select)
        vid_btn_frame = ttk.Frame(vid_frame)
        vid_btn_frame.pack(fill=tk.X, padx=4, pady=2)
        ttk.Button(vid_btn_frame, text="Add...", command=self._add_videos).pack(side=tk.LEFT, padx=2)
        ttk.Button(vid_btn_frame, text="Remove", command=self._remove_video).pack(side=tk.LEFT, padx=2)
        ttk.Button(vid_btn_frame, text="Clear All", command=self._clear_videos).pack(side=tk.LEFT, padx=2)

        # Video info
        self.video_info_label = ttk.Label(vid_frame, text="", wraplength=200, justify=tk.LEFT)
        self.video_info_label.pack(padx=4, pady=2)

        # Components
        comp_frame = ttk.LabelFrame(left_frame, text="Components")
        comp_frame.pack(fill=tk.BOTH, expand=True, padx=4, pady=2)
        self.comp_checks_frame = ttk.Frame(comp_frame)
        self.comp_checks_frame.pack(fill=tk.BOTH, expand=True, padx=4, pady=2)
        self.comp_vars: dict[str, tk.BooleanVar] = {}

        # -- Centre: canvas + slider --
        centre_frame = ttk.Frame(main_pane)
        main_pane.add(centre_frame, weight=1)

        self.canvas = tk.Canvas(centre_frame, bg="#1a1a1a", highlightthickness=0)
        self.canvas.pack(fill=tk.BOTH, expand=True, padx=4, pady=4)
        self.canvas.bind("<Configure>", self._on_canvas_resize)

        slider_frame = ttk.Frame(centre_frame)
        slider_frame.pack(fill=tk.X, padx=4, pady=2)
        self.scrub_slider = ttk.Scale(slider_frame, from_=0, to=100, orient=tk.HORIZONTAL,
                                       command=self._on_scrub)
        self.scrub_slider.pack(side=tk.LEFT, fill=tk.X, expand=True)
        self.time_label = ttk.Label(slider_frame, text="00:00:00", width=10)
        self.time_label.pack(side=tk.RIGHT, padx=4)

        # -- Right panel: settings --
        right_frame = ttk.Frame(main_pane, width=200)
        main_pane.add(right_frame, weight=0)

        settings_frame = ttk.LabelFrame(right_frame, text="Settings")
        settings_frame.pack(fill=tk.X, padx=4, pady=2)

        ttk.Label(settings_frame, text="Speed Unit:").pack(padx=4, pady=(4, 0), anchor=tk.W)
        speed_combo = ttk.Combobox(settings_frame, textvariable=self.speed_unit,
                                    values=["kph", "mph", "knots"], state="readonly", width=10)
        speed_combo.pack(padx=4, pady=2, anchor=tk.W)

        ttk.Label(settings_frame, text="GPX Time Offset (s):").pack(padx=4, pady=(4, 0), anchor=tk.W)
        offset_frame = ttk.Frame(settings_frame)
        offset_frame.pack(fill=tk.X, padx=4, pady=2)
        self.gpx_offset_entry = ttk.Entry(offset_frame, width=7)
        self.gpx_offset_entry.insert(0, "0.0")
        self.gpx_offset_entry.pack(side=tk.RIGHT, padx=(4, 0))
        self.gpx_offset_entry.bind("<Return>", self._on_gpx_offset_entry)
        self.gpx_offset_entry.bind("<FocusOut>", self._on_gpx_offset_entry)
        gpx_offset_scale = ttk.Scale(offset_frame, from_=-3600, to=3600, orient=tk.HORIZONTAL,
                                      variable=self.gpx_time_offset,
                                      command=self._on_gpx_offset_change)
        gpx_offset_scale.pack(side=tk.LEFT, fill=tk.X, expand=True)

        ttk.Label(settings_frame, text="Sample Duration (s):").pack(padx=4, pady=(4, 0), anchor=tk.W)
        sample_frame = ttk.Frame(settings_frame)
        sample_frame.pack(fill=tk.X, padx=4, pady=2)
        self.sample_entry = ttk.Entry(sample_frame, textvariable=self.sample_duration, width=10)
        self.sample_entry.pack(side=tk.LEFT)
        ttk.Label(sample_frame, text="(blank = full)").pack(side=tk.LEFT, padx=4)

        gpu_text = f"GPU: {self.gpu_profile or 'not detected'}"
        ttk.Label(settings_frame, text=gpu_text).pack(padx=4, pady=(4, 0), anchor=tk.W)
        ttk.Checkbutton(settings_frame, text="Use GPU encoding",
                         variable=self.use_gpu,
                         state=tk.NORMAL if self.gpu_profile else tk.DISABLED).pack(padx=4, pady=2, anchor=tk.W)

        font_row = ttk.Frame(settings_frame)
        font_row.pack(fill=tk.X, padx=4, pady=(4, 0))
        ttk.Label(font_row, text="Font:").pack(side=tk.LEFT)
        self.font_label = ttk.Label(font_row, text=Path(self.font_path).name, foreground="#4488ff")
        self.font_label.pack(side=tk.LEFT, padx=4)
        ttk.Button(font_row, text="...", width=3, command=self._select_font).pack(side=tk.LEFT)

        # Snap grid
        snap_frame = ttk.LabelFrame(right_frame, text="Snap Grid")
        snap_frame.pack(fill=tk.X, padx=4, pady=4)

        ttk.Checkbutton(snap_frame, text="Enable snap to grid",
                         variable=self.snap_enabled,
                         command=self._on_snap_toggle).pack(padx=4, pady=2, anchor=tk.W)

        grid_row = ttk.Frame(snap_frame)
        grid_row.pack(fill=tk.X, padx=4, pady=2)
        ttk.Label(grid_row, text="Grid size:").pack(side=tk.LEFT)
        self.snap_size_label = ttk.Label(grid_row, text="20", width=4)
        self.snap_size_label.pack(side=tk.RIGHT)
        self.snap_size_scale = ttk.Scale(grid_row, from_=5, to=100, orient=tk.HORIZONTAL,
                                          variable=self.snap_size,
                                          command=self._on_snap_size_change)
        self.snap_size_scale.pack(side=tk.LEFT, fill=tk.X, expand=True, padx=4)

        # Hidden labels for map settings (updated by component dialogs)
        self.map_size_label = ttk.Label(right_frame)
        self.map_zoom_label = ttk.Label(right_frame)

        # Action buttons
        action_frame = ttk.LabelFrame(right_frame, text="Actions")
        action_frame.pack(fill=tk.X, padx=4, pady=8)
        ttk.Button(action_frame, text="Save Layout XML...", command=self._save_layout_xml).pack(
            fill=tk.X, padx=8, pady=4)
        ttk.Button(action_frame, text="Export Shell Script...", command=self._export_shell_script).pack(
            fill=tk.X, padx=8, pady=4)
        ttk.Button(action_frame, text="Run Encoding...", command=self._run_encoding).pack(
            fill=tk.X, padx=8, pady=4)

        # Status bar
        self.status_var = tk.StringVar(value="Ready. Open a GPX file and add videos to begin.")
        status_bar = ttk.Label(self, textvariable=self.status_var, relief=tk.SUNKEN, anchor=tk.W)
        status_bar.pack(fill=tk.X, side=tk.BOTTOM, padx=2, pady=2)

    # -- File operations --

    def _open_gpx(self):
        path = filedialog.askopenfilename(
            title="Select Track File",
            initialdir=self.workdir,
            filetypes=[("GPX files", "*.gpx"), ("FIT files", "*.fit"),
                       ("EUC World XLSX", "*.xlsx"), ("All files", "*.*")])
        if path:
            self.gpx_path = Path(path)
            suffix = self.gpx_path.suffix.lower()
            type_label = {".gpx": "GPX", ".fit": "FIT", ".xlsx": "XLSX"}.get(suffix, suffix.upper())
            self.gpx_label.config(text=f"[{type_label}] {self.gpx_path.name}")
            # Scan route data for metric ranges
            self.route_ranges = _scan_route_ranges(self.gpx_path)
            if suffix == ".xlsx":
                self.status_var.set(f"Route: {self.gpx_path.name} (EUC World XLSX)")
            else:
                loc = _extract_gpx_first_point(self.gpx_path)
                if loc:
                    self.gpx_location = loc
                    self.status_var.set(f"Route: {self.gpx_path.name} ({type_label}, lat={loc[0]:.4f}, lon={loc[1]:.4f})")
                else:
                    self.status_var.set(f"Route: {self.gpx_path.name} ({type_label})")

    def _clear_gpx(self):
        self.gpx_path = None
        self.gpx_location = None
        self.gpx_label.config(text="(none)")
        self.status_var.set("Route file cleared.")

    def _remove_video(self):
        sel = self.video_listbox.curselection()
        if not sel:
            return
        idx = sel[0]
        removed = self.videos.pop(idx)
        self.video_listbox.delete(idx)

        # Reset canvas if the active video was removed
        if idx == self.active_video_idx:
            self.active_video_idx = -1
            self.canvas.delete("all")
            self.current_photo = None
            self.components = []
            self._rebuild_component_checkboxes()
            self.video_info_label.config(text="")
            # Select next available video
            if self.videos:
                new_idx = min(idx, len(self.videos) - 1)
                self.video_listbox.selection_set(new_idx)
                self._select_video(new_idx)
        elif idx < self.active_video_idx:
            self.active_video_idx -= 1

        self.status_var.set(f"Removed {removed.basename}. {len(self.videos)} video(s) remaining.")

    def _clear_videos(self):
        if not self.videos:
            return
        if not messagebox.askyesno("Clear Videos", f"Remove all {len(self.videos)} video(s)?"):
            return
        self.videos.clear()
        self.video_listbox.delete(0, tk.END)
        self.active_video_idx = -1
        self.canvas.delete("all")
        self.current_photo = None
        self.components = []
        self._rebuild_component_checkboxes()
        self.video_info_label.config(text="")
        self.status_var.set("All videos cleared.")

    def _select_font(self):
        FontSelectorDialog(self)

    def _on_gpx_offset_change(self, value):
        v = round(float(value), 1)
        self.gpx_time_offset.set(v)
        self.gpx_offset_entry.delete(0, tk.END)
        self.gpx_offset_entry.insert(0, str(v))

    def _on_gpx_offset_entry(self, event=None):
        try:
            v = float(self.gpx_offset_entry.get())
            self.gpx_time_offset.set(v)
        except ValueError:
            self.gpx_offset_entry.delete(0, tk.END)
            self.gpx_offset_entry.insert(0, str(self.gpx_time_offset.get()))

    def _on_map_size_change(self, value):
        size = int(float(value))
        self.map_size.set(size)
        self.map_size_label.config(text=str(size))
        # Update map component bounding boxes on canvas
        for comp in self.components:
            if comp.name in ("moving_map", "journey_map"):
                comp.width = size
                comp.height = size
        self._redraw_components()

    def _on_map_zoom_change(self, value):
        zoom = int(float(value))
        self.map_zoom.set(zoom)
        self.map_zoom_label.config(text=str(zoom))

    def _get_map_props(self) -> dict:
        """Build the template variables dict for XML generation."""
        return {
            "map_size": self.map_size.get(),
            "map_zoom": self.map_zoom.get(),
            "map_corner_radius": self.map_corner_radius.get(),
            "map_opacity": self.map_opacity.get(),
            "map_rotate": str(self.map_rotate.get()).lower(),
            "jm_track_colour": self.jm_track_colour.get(),
            "jm_line_width": self.jm_line_width.get(),
            "jm_loc_colour": self.jm_loc_colour.get(),
            "jm_loc_outline": self.jm_loc_outline.get(),
            "jm_loc_size": self.jm_loc_size.get(),
            "date_format": self.date_format.get(),
            "time_format": self.time_format.get(),
            # Text defaults (overridable per-component via custom_props and auto-sizing)
            "text_size": 16,
            "value_font_size": 32,
            "speed_font_size": 160,
            "date_size": 16,
            "time_size": 32,
            "time_y": 24,
            "icon_size": 64,
            "value_y": 18,
            "text_rgb": "255,255,255",
            # Chart defaults
            "chart_height": 64,
            "chart_filled": "true",
            "chart_seconds": 300,
            "chart_corner_radius": 0,
            "chart_outline": "0,0,0,0",
            "chart_bg": "0,0,0,170",
            "chart_fill": "91,113,146,170",
            "chart_line": "255,255,255,170",
            # Gauge defaults — these are overridden per-component by
            # _auto_gauge_ranges() when route data is available
            "gauge_size": 256,
            "gauge_style": "cairo-gauge-round-annotated",
            "gauge_max": 100,
            "gauge_min": 0,
            "gauge_start": 143,
            "gauge_length": 254,
            "gauge_sectors": 6,
            "gauge_bg": "255,255,255,150",
            "gauge_needle": "255,0,0",
            "gauge_tick": "0,0,0",
            "gauge_ann": "0,0,0",
            "gauge_fill": "0,191,255",
            "regen_colour": "68,187,68,180",
            # MSI defaults
            "msi_green": 0,
            "msi_yellow": 40,
            "msi_end": 60,
            "msi_rotate": 180,
            "msi_textsize": 16,
            # Sub-component colour defaults
            "label_rgb": "255,255,255",
            "value_rgb": "255,255,255",
            # Component title defaults
            "show_title": "false",
            "comp_title": "",
            "title_size": 14,
            "title_rgb": "255,255,255",
            "title_x": 128,
            "title_gauge_y": 200,
            # Heading tape defaults
            "ht_font_size": 16,
            "ht_tick_interval": 10,
            "ht_label_interval": 30,
            "ht_visible_range": 90,
            "ht_show_values": "true",
            "ht_show_border": "true",
            "ht_bg": "0,0,0",
            "ht_fg": "255,255,255",
            "ht_marker": "255,0,0",
            "ht_opacity": 180,
        }

    def _on_snap_toggle(self):
        self._draw_snap_grid()
        self._redraw_components()

    def _on_snap_size_change(self, value):
        size = int(float(value))
        self.snap_size.set(size)
        self.snap_size_label.config(text=str(size))
        if self.snap_enabled.get():
            self._draw_snap_grid()

    def _snap_value(self, val: int) -> int:
        """Snap a value to the nearest grid point."""
        grid = self.snap_size.get()
        return round(val / grid) * grid

    def _draw_snap_grid(self):
        """Draw or clear the snap grid overlay on the canvas."""
        self.canvas.delete("snap_grid")
        if not self.snap_enabled.get() or self.active_video_idx < 0:
            return

        video = self.videos[self.active_video_idx]
        s = self.scale_factor
        grid = self.snap_size.get()
        w = int(video.eff_width * s)
        h = int(video.eff_height * s)

        # Vertical lines
        x = 0
        while x <= video.eff_width:
            cx = int(x * s)
            self.canvas.create_line(cx, 0, cx, h,
                                     fill="#444444", dash=(2, 4),
                                     tags="snap_grid")
            x += grid

        # Horizontal lines
        y = 0
        while y <= video.eff_height:
            cy = int(y * s)
            self.canvas.create_line(0, cy, w, cy,
                                     fill="#444444", dash=(2, 4),
                                     tags="snap_grid")
            y += grid

        # Ensure grid is above background but below components
        self.canvas.tag_raise("snap_grid", "bg_image")
        self.canvas.tag_raise("overlay_comp", "snap_grid")

    def _show_sync_dialog(self):
        if not self.gpx_path:
            messagebox.showinfo("Info", "Load a GPX file first.")
            return
        if not self.videos:
            messagebox.showinfo("Info", "Add a video first.")
            return
        video = self.videos[max(0, self.active_video_idx)]
        SyncDialog(self, video, self.gpx_path, self.gpx_time_offset)

    def _show_api_keys_dialog(self):
        ApiKeysDialog(self)

    def _add_videos(self):
        paths = filedialog.askopenfilenames(
            title="Select Video Files",
            initialdir=self.workdir,
            filetypes=[("Video files", "*.mp4 *.MP4 *.mov *.MOV *.avi"), ("All files", "*.*")])
        if not paths:
            return
        for p in paths:
            self.status_var.set(f"Analysing {Path(p).name}...")
            self.update_idletasks()
            try:
                entry = analyse_video(Path(p))
                self.videos.append(entry)
                self.video_listbox.insert(tk.END, entry.basename)
            except Exception as e:
                messagebox.showerror("Error", f"Failed to analyse {Path(p).name}:\n{e}")

        if self.videos and self.active_video_idx < 0:
            self.video_listbox.selection_set(0)
            self._select_video(0)

        self.status_var.set(f"{len(self.videos)} video(s) loaded.")

    def _on_video_select(self, event):
        sel = self.video_listbox.curselection()
        if sel:
            self._select_video(sel[0])

    def _select_video(self, idx: int):
        self.active_video_idx = idx
        video = self.videos[idx]

        # Update GPX location to match video start time
        if self.gpx_path and video.creation_time:
            loc = _find_gpx_point_at_time(self.gpx_path, video.creation_time)
            if loc:
                self.gpx_location = loc

        # Update info label
        orient = "Portrait" if video.eff_height > video.eff_width else "Landscape"
        dur = _format_time(video.duration_seconds)
        self.video_info_label.config(
            text=f"{video.eff_width}x{video.eff_height} ({orient})\n"
                 f"Rotation: {video.rotation}\u00b0\nDuration: {dur}")

        # Update slider range
        self.scrub_slider.config(to=max(1, int(video.duration_seconds)))

        # Reset components for this resolution
        self.components = default_component_positions(video.eff_width, video.eff_height)
        # Reapply layout XML — either a pending CLI layout or a previously loaded one
        layout = getattr(self, "_pending_layout_xml", None) or getattr(self, "_loaded_layout_path", None)
        if layout:
            self._load_layout_xml(layout)
            self._pending_layout_xml = None
        self._rebuild_component_checkboxes()

        # Calculate scale
        self._update_scale()
        self._draw_snap_grid()

        # Extract and show first frame
        self._extract_and_show(0.0)

    def _update_scale(self):
        if self.active_video_idx < 0:
            return
        video = self.videos[self.active_video_idx]
        cw = self.canvas.winfo_width() or self.CANVAS_MAX_W
        ch = self.canvas.winfo_height() or self.CANVAS_MAX_H
        self.scale_factor = min(cw / video.eff_width, ch / video.eff_height)

    def _on_canvas_resize(self, event):
        self._update_scale()
        self._redraw_components()
        self._draw_snap_grid()
        # Re-show current frame if we have one
        if self.current_photo:
            self._show_current_frame()

    # -- Frame extraction --

    def _extract_and_show(self, time_seconds: float):
        if self.active_video_idx < 0:
            return
        video = self.videos[self.active_video_idx]
        self.status_var.set(f"Extracting frame at {_format_time(time_seconds)}...")
        self.update_idletasks()

        # Run extraction in background thread
        def _extract():
            img = extract_frame(video.path, time_seconds, video.eff_width, video.eff_height)
            self.frame_queue.put(img)

        threading.Thread(target=_extract, daemon=True).start()
        self._poll_frame_queue()

    def _poll_frame_queue(self):
        try:
            img = self.frame_queue.get_nowait()
            if img:
                self._display_frame(img)
            else:
                self.status_var.set("Failed to extract frame.")
        except queue.Empty:
            self.after(50, self._poll_frame_queue)

    def _display_frame(self, img: Image.Image):
        self._update_scale()
        scaled_w = int(img.width * self.scale_factor)
        scaled_h = int(img.height * self.scale_factor)
        scaled = img.resize((scaled_w, scaled_h), Image.Resampling.LANCZOS)
        self.current_photo = ImageTk.PhotoImage(scaled)
        self._show_current_frame()
        self._redraw_components()
        self.status_var.set("Ready.")

    def _show_current_frame(self):
        self.canvas.delete("bg_image")
        if self.current_photo:
            self.canvas.create_image(0, 0, anchor=tk.NW, image=self.current_photo, tags="bg_image")
            self.canvas.tag_lower("bg_image")

    # -- Scrub slider --

    def _on_scrub(self, value):
        t = float(value)
        self.time_label.config(text=_format_time(t))
        # Debounce frame extraction
        if self._scrub_after_id:
            self.after_cancel(self._scrub_after_id)
        self._scrub_after_id = self.after(300, lambda: self._extract_and_show(t))

    # -- Component checkboxes --

    def _rebuild_component_checkboxes(self):
        for widget in self.comp_checks_frame.winfo_children():
            widget.destroy()
        self.comp_vars.clear()

        for comp in self.components:
            var = tk.BooleanVar(value=comp.enabled)
            self.comp_vars[comp.name] = var
            cb = ttk.Checkbutton(
                self.comp_checks_frame, text=comp.label, variable=var,
                command=lambda c=comp, v=var: self._toggle_component(c, v))
            cb.pack(anchor=tk.W, padx=2, pady=1)

    def _toggle_component(self, comp: OverlayComponent, var: tk.BooleanVar):
        comp.enabled = var.get()
        self._redraw_components()

    # -- Component drawing and dragging --

    def _redraw_components(self):
        self.canvas.delete("overlay_comp")
        for comp in self.components:
            comp.canvas_rect_id = None
            comp.canvas_text_id = None
            comp.canvas_handle_ids = {}
            if not comp.enabled:
                continue
            self._draw_component(comp)

    _HANDLE_SIZE = 8  # pixels, half-size of resize handle squares

    def _draw_component(self, comp: OverlayComponent):
        s = self.scale_factor
        cx = int(comp.x * s)
        cy = int(comp.y * s)
        cw = int(comp.width * s)
        ch = int(comp.height * s)
        hs = self._HANDLE_SIZE

        rect_id = self.canvas.create_rectangle(
            cx, cy, cx + cw, cy + ch,
            outline=comp.color, fill=comp.color, stipple="gray25",
            width=2, tags="overlay_comp")

        text_id = self.canvas.create_text(
            cx + cw // 2, cy + ch // 2,
            text=comp.label, fill="white", font=("sans-serif", 9, "bold"),
            tags="overlay_comp")

        comp.canvas_rect_id = rect_id
        comp.canvas_text_id = text_id

        # Resize handles: bottom-right corner, right edge midpoint, bottom edge midpoint
        comp.canvas_handle_ids = {}
        handles = {
            "br": (cx + cw, cy + ch),
            "r":  (cx + cw, cy + ch // 2),
            "b":  (cx + cw // 2, cy + ch),
        }
        for edge, (hx, hy) in handles.items():
            handle_id = self.canvas.create_rectangle(
                hx - hs, hy - hs, hx + hs, hy + hs,
                outline=comp.color, fill="white", width=1,
                tags="overlay_comp")
            comp.canvas_handle_ids[edge] = handle_id
            self.canvas.tag_bind(handle_id, "<ButtonPress-1>",
                                  lambda e, c=comp, ed=edge: self._resize_start(e, c, ed))
            self.canvas.tag_bind(handle_id, "<B1-Motion>",
                                  lambda e, c=comp: self._resize_motion(e, c))
            self.canvas.tag_bind(handle_id, "<ButtonRelease-1>",
                                  lambda e, c=comp: self._resize_end(e, c))
            # Set cursor for resize handles
            cursor = {"br": "bottom_right_corner", "r": "right_side", "b": "bottom_side"}[edge]
            self.canvas.tag_bind(handle_id, "<Enter>",
                                  lambda e, cur=cursor: self.canvas.config(cursor=cur))
            self.canvas.tag_bind(handle_id, "<Leave>",
                                  lambda e: self.canvas.config(cursor=""))

        # Bind drag events and right-click on main rect and label
        for item_id in (rect_id, text_id):
            self.canvas.tag_bind(item_id, "<ButtonPress-1>",
                                  lambda e, c=comp: self._drag_start(e, c))
            self.canvas.tag_bind(item_id, "<B1-Motion>",
                                  lambda e, c=comp: self._drag_motion(e, c))
            self.canvas.tag_bind(item_id, "<ButtonRelease-1>",
                                  lambda e, c=comp: self._drag_end(e, c))
            self.canvas.tag_bind(item_id, "<ButtonPress-3>",
                                  lambda e, c=comp: self._component_context_menu(e, c))

    def _update_handles(self, comp: OverlayComponent):
        """Move resize handles to match current component rect."""
        s = self.scale_factor
        cx = int(comp.x * s)
        cy = int(comp.y * s)
        cw = int(comp.width * s)
        ch = int(comp.height * s)
        hs = self._HANDLE_SIZE
        positions = {
            "br": (cx + cw, cy + ch),
            "r":  (cx + cw, cy + ch // 2),
            "b":  (cx + cw // 2, cy + ch),
        }
        for edge, (hx, hy) in positions.items():
            hid = comp.canvas_handle_ids.get(edge)
            if hid:
                self.canvas.coords(hid, hx - hs, hy - hs, hx + hs, hy + hs)

    def _drag_start(self, event, comp: OverlayComponent):
        self._drag_component = comp
        s = self.scale_factor
        self._drag_offset_x = event.x - int(comp.x * s)
        self._drag_offset_y = event.y - int(comp.y * s)

    def _drag_motion(self, event, comp: OverlayComponent):
        if self._drag_component is not comp:
            return
        s = self.scale_factor
        new_x = (event.x - self._drag_offset_x) / s
        new_y = (event.y - self._drag_offset_y) / s

        # Clamp to video bounds
        if self.active_video_idx >= 0:
            video = self.videos[self.active_video_idx]
            new_x = max(0, min(new_x, video.eff_width - comp.width))
            new_y = max(0, min(new_y, video.eff_height - comp.height))

        comp.x = int(new_x)
        comp.y = int(new_y)

        # Snap to grid if enabled
        if self.snap_enabled.get():
            comp.x = self._snap_value(comp.x)
            comp.y = self._snap_value(comp.y)

        # Move canvas items and handles
        cx = int(comp.x * s)
        cy = int(comp.y * s)
        cw = int(comp.width * s)
        ch = int(comp.height * s)
        self.canvas.coords(comp.canvas_rect_id, cx, cy, cx + cw, cy + ch)
        self.canvas.coords(comp.canvas_text_id, cx + cw // 2, cy + ch // 2)
        self._update_handles(comp)

    def _drag_end(self, event, comp: OverlayComponent):
        self._drag_component = None
        if self.active_video_idx >= 0:
            video = self.videos[self.active_video_idx]
            self.status_var.set(
                f"{comp.label}: ({comp.x}, {comp.y}) on {video.eff_width}x{video.eff_height}")

    def _resize_start(self, event, comp: OverlayComponent, edge: str):
        self._resize_component = comp
        self._resize_edge = edge
        # Prevent drag handler from also activating
        self._drag_component = None

    def _resize_motion(self, event, comp: OverlayComponent):
        if self._resize_component is not comp:
            return
        s = self.scale_factor
        mouse_x = event.x / s
        mouse_y = event.y / s
        min_size = 20

        if self._resize_edge in ("r", "br"):
            comp.width = max(min_size, int(mouse_x - comp.x))
        if self._resize_edge in ("b", "br"):
            comp.height = max(min_size, int(mouse_y - comp.y))

        if self.snap_enabled.get():
            comp.width = self._snap_value(comp.width)
            comp.height = self._snap_value(comp.height)

        # Update rect, label, and handles in place — no full redraw
        cx = int(comp.x * s)
        cy = int(comp.y * s)
        cw = int(comp.width * s)
        ch = int(comp.height * s)
        self.canvas.coords(comp.canvas_rect_id, cx, cy, cx + cw, cy + ch)
        self.canvas.coords(comp.canvas_text_id, cx + cw // 2, cy + ch // 2)
        self._update_handles(comp)
        self.status_var.set(
            f"{comp.label}: ({comp.x}, {comp.y}) {comp.width}x{comp.height}")

    def _resize_end(self, event, comp: OverlayComponent):
        self._resize_component = None
        self._resize_edge = ""
        # Redraw once to reposition handles to new size
        self._redraw_components()
        self.status_var.set(
            f"{comp.label}: ({comp.x}, {comp.y}) {comp.width}x{comp.height}")

    def _component_context_menu(self, event, comp: OverlayComponent):
        """Show right-click context menu for a component."""
        menu = tk.Menu(self, tearoff=0)
        menu.add_command(label=f"Properties: {comp.label}...",
                         command=lambda: ComponentOptionsDialog(self, comp))
        menu.add_separator()
        menu.add_command(label="Disable", command=lambda: self._disable_component(comp))
        menu.add_separator()

        # Quick position presets
        if self.active_video_idx >= 0:
            video = self.videos[self.active_video_idx]
            menu.add_command(label="Move to Top-Left",
                             command=lambda: self._move_component(comp, 16, 16))
            menu.add_command(label="Move to Top-Right",
                             command=lambda: self._move_component(comp, video.eff_width - comp.width - 16, 16))
            menu.add_command(label="Move to Bottom-Left",
                             command=lambda: self._move_component(comp, 16, video.eff_height - comp.height - 16))
            menu.add_command(label="Move to Bottom-Right",
                             command=lambda: self._move_component(
                                 comp, video.eff_width - comp.width - 16, video.eff_height - comp.height - 16))
            menu.add_command(label="Centre",
                             command=lambda: self._move_component(
                                 comp, (video.eff_width - comp.width) // 2, (video.eff_height - comp.height) // 2))

        menu.tk_popup(event.x_root, event.y_root)

    def _disable_component(self, comp: OverlayComponent):
        comp.enabled = False
        if comp.name in self.comp_vars:
            self.comp_vars[comp.name].set(False)
        self._redraw_components()

    def _move_component(self, comp: OverlayComponent, x: int, y: int):
        if self.snap_enabled.get():
            x = self._snap_value(x)
            y = self._snap_value(y)
        comp.x = x
        comp.y = y
        self._redraw_components()

    # -- Save / Export --

    def _load_layout_xml(self, path: Optional[str] = None):
        if not self.components:
            messagebox.showinfo("Info", "Add a video first to initialise components.")
            return
        if not path:
            path = filedialog.askopenfilename(
                title="Load Layout XML",
                initialdir=self.workdir,
                filetypes=[("XML files", "*.xml"), ("All files", "*.*")])
        if not path:
            return
        try:
            load_layout_xml(Path(path), self.components)
            # Sync global settings from loaded component state
            for comp in self.components:
                if comp.name == "moving_map" and comp.enabled:
                    self.map_size.set(comp.width)
                    self.map_size_label.config(text=str(comp.width))
                if comp.name in ("date_and_time", "date_only") and comp.enabled:
                    if "date_format" in comp.custom_props:
                        self.date_format.set(comp.custom_props["date_format"])
                if comp.name in ("date_and_time", "time_only") and comp.enabled:
                    if "time_format" in comp.custom_props:
                        self.time_format.set(comp.custom_props["time_format"])
            self._loaded_layout_path = str(path)
            self._rebuild_component_checkboxes()
            self._redraw_components()
            self.status_var.set(f"Layout loaded: {Path(path).name}")
        except Exception as e:
            messagebox.showerror("Error", f"Failed to load layout:\n{e}")

    def _save_layout_xml(self):
        if not self.components:
            messagebox.showinfo("Info", "No layout to save. Add a video first.")
            return
        path = filedialog.asksaveasfilename(
            title="Save Layout XML",
            initialdir=self.workdir,
            defaultextension=".xml",
            filetypes=[("XML files", "*.xml"), ("All files", "*.*")])
        if not path:
            return
        try:
            xml = generate_layout_xml(self.components, self._get_map_props(), self.route_ranges)
            Path(path).write_text(xml)
            self.status_var.set(f"Layout saved: {Path(path).name}")
        except Exception as e:
            messagebox.showerror("Error", f"Failed to save layout:\n{e}")

    def _export_shell_script(self):
        if not self._validate_for_encoding():
            return
        path = filedialog.asksaveasfilename(
            title="Export Shell Script",
            initialdir=self.workdir,
            defaultextension=".sh",
            filetypes=[("Shell scripts", "*.sh"), ("All files", "*.*")])
        if not path:
            return
        try:
            xml = generate_layout_xml(self.components, self._get_map_props(), self.route_ranges)

            gpu = self.gpu_profile if self.use_gpu.get() else None
            script = generate_shell_script(
                videos=self.videos,
                gpx_path=self.gpx_path,
                layout_xml=xml,
                speed_unit=self.speed_unit.get(),
                font=self.font_path,
                gpu_profile=gpu,
                dashboard_script=self._dashboard_script,
                map_style=self.map_style.get(),
                gpx_time_offset=self.gpx_time_offset.get(),
                sample_duration=self._get_sample_duration(),
            )
            Path(path).write_text(script)
            os.chmod(path, 0o755)
            self.status_var.set(f"Script exported: {Path(path).name}")
        except Exception as e:
            messagebox.showerror("Error", f"Failed to export script:\n{e}")

    def _get_sample_duration(self) -> Optional[float]:
        val = self.sample_duration.get().strip()
        if not val:
            return None
        try:
            return float(val)
        except ValueError:
            return None

    def _validate_for_encoding(self) -> bool:
        if not self.gpx_path:
            messagebox.showwarning("Missing GPX", "Please open a GPX file first.")
            return False
        if not self.videos:
            messagebox.showwarning("No Videos", "Please add at least one video.")
            return False
        if not self.components:
            messagebox.showwarning("No Layout", "No layout components configured.")
            return False
        return True

    # -- In-app encoding --

    def _run_encoding(self):
        if not self._validate_for_encoding():
            return

        # Save layout XML to temp file (cleaned up by EncodingDialog on close)
        import tempfile
        try:
            xml = generate_layout_xml(self.components, self._get_map_props(), self.route_ranges)
        except Exception as e:
            messagebox.showerror("Error", f"Failed to generate layout XML:\n{e}")
            return
        fd, tmp_path = tempfile.mkstemp(prefix="gopro-layout-", suffix=".xml")
        os.close(fd)
        xml_path = Path(tmp_path)
        xml_path.write_text(xml)

        EncodingDialog(self, self.videos, self.gpx_path, xml_path,
                       self.speed_unit.get(), self.font_path,
                       self.gpu_profile if self.use_gpu.get() else None,
                       self._dashboard_script,
                       self.map_style.get(),
                       self.gpx_time_offset.get(),
                       self._get_sample_duration())


# ---------------------------------------------------------------------------
# Component options dialog (right-click properties)
# ---------------------------------------------------------------------------

# Per-component configurable options
COMPONENT_OPTIONS = {
    "date_and_time": {
        "title": "Date + Time Display Options",
        "fields": [
            ("date_format", "Date Format", "combo_editable", "%Y/%m/%d", DATE_FORMATS),
            ("time_format", "Time Format", "combo_editable", "%H:%M:%S.%f", TIME_FORMATS),
        ],
        "has_datetime_preview": True,
    },
    "date_only": {
        "title": "Date Display Options",
        "fields": [
            ("date_format", "Date Format", "combo_editable", "%Y/%m/%d", DATE_FORMATS),
        ],
        "has_datetime_preview": True,
    },
    "time_only": {
        "title": "Time Display Options",
        "fields": [
            ("time_format", "Time Format", "combo_editable", "%H:%M:%S.%f", TIME_FORMATS),
        ],
        "has_datetime_preview": True,
    },
    "gps_info": {
        "title": "GPS Info Options",
        "fields": [
            ("text_size", "Font Size", "spinbox", 16, (8, 48, 2)),
        ],
    },
    "gps_coords": {
        "title": "GPS Coordinates Options",
        "fields": [
            ("text_size", "Font Size", "spinbox", 16, (8, 48, 2)),
        ],
    },
    "big_mph": {
        "title": "Speed Display Options",
        "fields": [
            ("speed_unit", "Speed Unit", "combo", "kph", ["kph", "mph", "knots"]),
            ("speed_font_size", "Speed Font Size", "spinbox", 160, (32, 400, 8)),
            ("text_size", "Label Font Size", "spinbox", 16, (8, 48, 2)),
            ("label_rgb", "Label Colour (R,G,B)", "colour_select", "255,255,255", []),
            ("value_rgb", "Value Colour (R,G,B)", "colour_select", "255,255,255", []),
            ("comp_font", "Font", "font_select", "", []),
        ],
    },
    "gradient": {
        "title": "Gradient Options",
        "fields": [
            ("icon_size", "Icon Size", "spinbox", 64, (16, 128, 4)),
            ("value_font_size", "Value Font Size", "spinbox", 32, (12, 96, 4)),
            ("text_size", "Label Font Size", "spinbox", 16, (8, 48, 2)),
            ("value_y", "Value Y Offset", "spinbox", 18, (0, 100, 2)),
            ("label_rgb", "Label Colour (R,G,B)", "colour_select", "255,255,255", []),
            ("value_rgb", "Value Colour (R,G,B)", "colour_select", "255,255,255", []),
        ],
    },
    "altitude": {
        "title": "Altitude Options",
        "fields": [
            ("icon_size", "Icon Size", "spinbox", 64, (16, 128, 4)),
            ("value_font_size", "Value Font Size", "spinbox", 32, (12, 96, 4)),
            ("text_size", "Label Font Size", "spinbox", 16, (8, 48, 2)),
            ("value_y", "Value Y Offset", "spinbox", 18, (0, 100, 2)),
            ("label_rgb", "Label Colour (R,G,B)", "colour_select", "255,255,255", []),
            ("value_rgb", "Value Colour (R,G,B)", "colour_select", "255,255,255", []),
        ],
    },
    "temperature": {
        "title": "Temperature Options",
        "fields": [
            ("icon_size", "Icon Size", "spinbox", 64, (16, 128, 4)),
            ("value_font_size", "Value Font Size", "spinbox", 32, (12, 96, 4)),
            ("text_size", "Label Font Size", "spinbox", 16, (8, 48, 2)),
            ("value_y", "Value Y Offset", "spinbox", 18, (0, 100, 2)),
        ],
    },
    "cadence": {
        "title": "Cadence Options",
        "fields": [
            ("icon_size", "Icon Size", "spinbox", 64, (16, 128, 4)),
            ("value_font_size", "Value Font Size", "spinbox", 32, (12, 96, 4)),
            ("text_size", "Label Font Size", "spinbox", 16, (8, 48, 2)),
            ("value_y", "Value Y Offset", "spinbox", 18, (0, 100, 2)),
        ],
    },
    "heartbeat": {
        "title": "Heart Rate Options",
        "fields": [
            ("icon_size", "Icon Size", "spinbox", 64, (16, 128, 4)),
            ("value_font_size", "Value Font Size", "spinbox", 32, (12, 96, 4)),
            ("text_size", "Label Font Size", "spinbox", 16, (8, 48, 2)),
            ("value_y", "Value Y Offset", "spinbox", 18, (0, 100, 2)),
        ],
    },
    "moving_map": {
        "title": "Moving Map Options",
        "fields": [
            ("map_zoom", "Zoom Level", "slider", 16, (8, 20)),
            ("map_corner_radius", "Rounded Corners (0=off)", "spinbox", 35, (0, 128, 5)),
            ("map_opacity", "Opacity", "slider_float", 0.7, (0.0, 1.0)),
            ("map_rotate", "Rotate Map to Heading", "checkbox", True, []),
            ("map_style", "Map Style", "combo_preview", "osm", MAP_STYLES),
        ],
    },
    "journey_map": {
        "title": "Journey Map Options",
        "fields": [
            ("map_corner_radius", "Rounded Corners (0=off)", "spinbox", 35, (0, 128, 5)),
            ("map_opacity", "Opacity", "slider_float", 0.7, (0.0, 1.0)),
            ("jm_track_colour", "Track Colour", "colour_select", "255,0,0", []),
            ("jm_line_width", "Track Width", "slider", 4, (1, 12)),
            ("jm_loc_colour", "Location Fill", "colour_select", "0,0,255", []),
            ("jm_loc_outline", "Location Outline", "colour_select", "0,0,0", []),
            ("jm_loc_size", "Location Marker Size", "slider", 6, (2, 20)),
            ("map_style", "Map Style", "combo_preview", "osm", MAP_STYLES),
        ],
    },
    "altitude": {
        "title": "Altitude Options",
        "fields": [
            ("units", "Units", "combo", "metres", ["metres", "feet"]),
        ],
    },
    "gradient_chart": {
        "title": "Altitude Chart Options",
        "fields": [
            ("show_title", "Show Title", "checkbox", False, []),
            ("comp_title", "Chart Title", "entry", "Altitude", []),
            ("title_size", "Title Font Size", "spinbox", 14, (8, 32, 2)),
            ("title_rgb", "Title Colour (R,G,B)", "colour_select", "255,255,255", []),
            ("chart_corner_radius", "Corner Radius", "spinbox", 0, (0, 64, 5)),
            ("chart_outline", "Border (R,G,B,A)", "colour_select", "0,0,0,0", []),
            ("chart_seconds", "Time Window (sec)", "spinbox", 300, (60, 1800, 30)),
            ("chart_filled", "Filled", "checkbox", True, []),
            ("chart_bg", "Background (R,G,B,A)", "colour_select", "0,0,0,170", []),
            ("chart_fill", "Fill Colour (R,G,B,A)", "colour_select", "91,113,146,170", []),
            ("chart_line", "Line Colour (R,G,B,A)", "colour_select", "255,255,255,170", []),
        ],
    },
    "battery_chart": {
        "title": "Battery Chart Options",
        "fields": [
            ("show_title", "Show Title", "checkbox", False, []),
            ("comp_title", "Chart Title", "entry", "Battery %", []),
            ("title_size", "Title Font Size", "spinbox", 14, (8, 32, 2)),
            ("title_rgb", "Title Colour (R,G,B)", "colour_select", "255,255,255", []),
        ],
    },
    "voltage_chart": {
        "title": "Voltage Chart Options",
        "fields": [
            ("show_title", "Show Title", "checkbox", False, []),
            ("comp_title", "Chart Title", "entry", "Voltage", []),
            ("title_size", "Title Font Size", "spinbox", 14, (8, 32, 2)),
            ("title_rgb", "Title Colour (R,G,B)", "colour_select", "255,255,255", []),
        ],
    },
    "power_chart": {
        "title": "Power Chart Options",
        "fields": [
            ("show_title", "Show Title", "checkbox", False, []),
            ("comp_title", "Chart Title", "entry", "Power", []),
            ("title_size", "Title Font Size", "spinbox", 14, (8, 32, 2)),
            ("title_rgb", "Title Colour (R,G,B)", "colour_select", "255,255,255", []),
        ],
    },
    "battery_value": {
        "title": "Battery % Display Options",
        "fields": [
            ("value_font_size", "Value Font Size", "spinbox", 32, (12, 96, 4)),
            ("text_size", "Label Font Size", "spinbox", 16, (8, 48, 2)),
            ("value_y", "Value Y Offset", "spinbox", 18, (0, 100, 2)),
            ("label_rgb", "Label Colour (R,G,B)", "colour_select", "255,255,255", []),
            ("value_rgb", "Value Colour (R,G,B)", "colour_select", "255,255,255", []),
        ],
    },
    "voltage_value": {
        "title": "Voltage Display Options",
        "fields": [
            ("value_font_size", "Value Font Size", "spinbox", 32, (12, 96, 4)),
            ("text_size", "Label Font Size", "spinbox", 16, (8, 48, 2)),
            ("value_y", "Value Y Offset", "spinbox", 18, (0, 100, 2)),
            ("label_rgb", "Label Colour (R,G,B)", "colour_select", "255,255,255", []),
            ("value_rgb", "Value Colour (R,G,B)", "colour_select", "255,255,255", []),
        ],
    },
    "current_value": {
        "title": "Current Display Options",
        "fields": [
            ("value_font_size", "Value Font Size", "spinbox", 32, (12, 96, 4)),
            ("text_size", "Label Font Size", "spinbox", 16, (8, 48, 2)),
            ("value_y", "Value Y Offset", "spinbox", 18, (0, 100, 2)),
            ("label_rgb", "Label Colour (R,G,B)", "colour_select", "255,255,255", []),
            ("value_rgb", "Value Colour (R,G,B)", "colour_select", "255,255,255", []),
        ],
    },
    "power_value": {
        "title": "Power Display Options",
        "fields": [
            ("value_font_size", "Value Font Size", "spinbox", 32, (12, 96, 4)),
            ("text_size", "Label Font Size", "spinbox", 16, (8, 48, 2)),
            ("value_y", "Value Y Offset", "spinbox", 18, (0, 100, 2)),
            ("label_rgb", "Label Colour (R,G,B)", "colour_select", "255,255,255", []),
            ("value_rgb", "Value Colour (R,G,B)", "colour_select", "255,255,255", []),
        ],
    },
    "avg_speed_value": {
        "title": "Avg Speed Display Options",
        "fields": [
            ("value_font_size", "Value Font Size", "spinbox", 32, (12, 96, 4)),
            ("text_size", "Label Font Size", "spinbox", 16, (8, 48, 2)),
            ("value_y", "Value Y Offset", "spinbox", 18, (0, 100, 2)),
            ("label_rgb", "Label Colour (R,G,B)", "colour_select", "255,255,255", []),
            ("value_rgb", "Value Colour (R,G,B)", "colour_select", "255,255,255", []),
        ],
    },
    "avg_speed_moving_value": {
        "title": "Avg Moving Speed Display Options",
        "fields": [
            ("value_font_size", "Value Font Size", "spinbox", 32, (12, 96, 4)),
            ("text_size", "Label Font Size", "spinbox", 16, (8, 48, 2)),
            ("value_y", "Value Y Offset", "spinbox", 18, (0, 100, 2)),
            ("label_rgb", "Label Colour (R,G,B)", "colour_select", "255,255,255", []),
            ("value_rgb", "Value Colour (R,G,B)", "colour_select", "255,255,255", []),
        ],
    },
    "avg_speed_gauge": {
        "title": "Avg Speed Gauge Options",
        "fields": [
            ("gauge_style", "Gauge Style", "combo", "cairo-gauge-round-annotated",
             ["cairo-gauge-round-annotated", "cairo-gauge-arc-annotated", "cairo-gauge-donut", "cairo-gauge-marker"]),
            ("gauge_size", "Size", "spinbox", 256, (64, 512, 16)),
            ("show_title", "Show Title", "checkbox", False, []),
            ("comp_title", "Gauge Title", "entry", "Avg Speed", []),
            ("title_size", "Title Font Size", "spinbox", 14, (8, 32, 2)),
            ("title_rgb", "Title Colour (R,G,B)", "colour_select", "255,255,255", []),
            ("gauge_max", "Max Value", "spinbox", 60, (10, 300, 5)),
            ("gauge_min", "Min Value", "spinbox", 0, (0, 0, 1)),
            ("gauge_start", "Start Angle (°)", "spinbox", 143, (0, 360, 5)),
            ("gauge_length", "Arc Length (°)", "spinbox", 254, (90, 360, 10)),
            ("gauge_sectors", "Sectors", "spinbox", 6, (2, 20, 1)),
        ],
    },
    "avg_speed_moving_gauge": {
        "title": "Avg Moving Speed Gauge Options",
        "fields": [
            ("gauge_style", "Gauge Style", "combo", "cairo-gauge-round-annotated",
             ["cairo-gauge-round-annotated", "cairo-gauge-arc-annotated", "cairo-gauge-donut", "cairo-gauge-marker"]),
            ("gauge_size", "Size", "spinbox", 256, (64, 512, 16)),
            ("show_title", "Show Title", "checkbox", False, []),
            ("comp_title", "Gauge Title", "entry", "Avg Moving", []),
            ("title_size", "Title Font Size", "spinbox", 14, (8, 32, 2)),
            ("title_rgb", "Title Colour (R,G,B)", "colour_select", "255,255,255", []),
            ("gauge_max", "Max Value", "spinbox", 60, (10, 300, 5)),
            ("gauge_min", "Min Value", "spinbox", 0, (0, 0, 1)),
            ("gauge_start", "Start Angle (°)", "spinbox", 143, (0, 360, 5)),
            ("gauge_length", "Arc Length (°)", "spinbox", 254, (90, 360, 10)),
            ("gauge_sectors", "Sectors", "spinbox", 6, (2, 20, 1)),
        ],
    },
    "avg_speed_bar": {
        "title": "Avg Speed Bar Options",
        "fields": [
            ("show_bar_scale", "Show Min/Max Scale", "checkbox", True, []),
            ("gauge_max", "Max Speed (kph)", "spinbox", 60, (20, 200, 10)),
            ("show_title", "Show Title", "checkbox", False, []),
            ("comp_title", "Bar Title", "entry", "Avg Speed", []),
            ("title_size", "Title Font Size", "spinbox", 14, (8, 32, 2)),
            ("title_rgb", "Title Colour (R,G,B)", "colour_select", "255,255,255", []),
        ],
    },
    "avg_speed_moving_bar": {
        "title": "Avg Moving Speed Bar Options",
        "fields": [
            ("show_bar_scale", "Show Min/Max Scale", "checkbox", True, []),
            ("gauge_max", "Max Speed (kph)", "spinbox", 60, (20, 200, 10)),
            ("show_title", "Show Title", "checkbox", False, []),
            ("comp_title", "Bar Title", "entry", "Avg Moving", []),
            ("title_size", "Title Font Size", "spinbox", 14, (8, 32, 2)),
            ("title_rgb", "Title Colour (R,G,B)", "colour_select", "255,255,255", []),
        ],
    },
    "speed_gauge": {
        "title": "Speed Gauge Options",
        "fields": [
            ("gauge_style", "Gauge Style", "combo", "cairo-gauge-round-annotated",
             ["cairo-gauge-round-annotated", "cairo-gauge-arc-annotated", "cairo-gauge-donut", "cairo-gauge-marker"]),
            ("gauge_size", "Size", "spinbox", 256, (64, 512, 16)),
            ("show_title", "Show Title", "checkbox", False, []),
            ("comp_title", "Gauge Title", "entry", "Speed", []),
            ("title_size", "Title Font Size", "spinbox", 14, (8, 32, 2)),
            ("title_rgb", "Title Colour (R,G,B)", "colour_select", "255,255,255", []),
            ("gauge_max", "Max Value", "spinbox", 60, (10, 300, 5)),
            ("gauge_min", "Min Value", "spinbox", 0, (-100, 100, 5)),
            ("gauge_start", "Start Angle (°)", "spinbox", 143, (0, 360, 5)),
            ("gauge_length", "Arc Length (°)", "spinbox", 254, (90, 360, 10)),
            ("gauge_sectors", "Sectors", "spinbox", 6, (2, 20, 1)),
            ("gauge_bg", "Background (R,G,B,A)", "colour_select", "255,255,255,150", []),
            ("gauge_needle", "Needle (R,G,B)", "colour_select", "255,0,0", []),
            ("gauge_tick", "Tick Marks (R,G,B)", "colour_select", "0,0,0", []),
            ("gauge_ann", "Annotations (R,G,B)", "colour_select", "0,0,0", []),
            ("gauge_fill", "Gauge Fill (R,G,B)", "colour_select", "0,191,255", []),
        ],
    },
    "battery_gauge": {
        "title": "Battery Gauge Options",
        "fields": [
            ("gauge_style", "Gauge Style", "combo", "cairo-gauge-donut",
             ["cairo-gauge-donut", "cairo-gauge-round-annotated", "cairo-gauge-arc-annotated", "cairo-gauge-marker"]),
            ("gauge_size", "Size", "spinbox", 256, (64, 512, 16)),
            ("show_title", "Show Title", "checkbox", False, []),
            ("comp_title", "Gauge Title", "entry", "Battery", []),
            ("title_size", "Title Font Size", "spinbox", 14, (8, 32, 2)),
            ("title_rgb", "Title Colour (R,G,B)", "colour_select", "255,255,255", []),
            ("gauge_max", "Max Value (%)", "spinbox", 100, (50, 100, 5)),
            ("gauge_min", "Min Value (%)", "spinbox", 0, (0, 50, 5)),
            ("gauge_start", "Start Angle (°)", "spinbox", 90, (0, 360, 5)),
            ("gauge_length", "Arc Length (°)", "spinbox", 270, (90, 360, 10)),
            ("gauge_sectors", "Sectors", "spinbox", 5, (2, 20, 1)),
            ("gauge_bg", "Background (R,G,B,A)", "colour_select", "255,255,255,150", []),
            ("gauge_needle", "Needle (R,G,B)", "colour_select", "255,0,0", []),
            ("gauge_tick", "Tick Marks (R,G,B)", "colour_select", "0,0,0", []),
            ("gauge_ann", "Annotations (R,G,B)", "colour_select", "0,0,0", []),
            ("gauge_fill", "Gauge Fill (R,G,B)", "colour_select", "0,191,255", []),
        ],
    },
    "power_gauge": {
        "title": "Power Gauge Options",
        "fields": [
            ("gauge_style", "Gauge Style", "combo", "cairo-gauge-arc-annotated",
             ["cairo-gauge-arc-annotated", "cairo-gauge-round-annotated", "cairo-gauge-donut", "cairo-gauge-marker"]),
            ("gauge_size", "Size", "spinbox", 256, (64, 512, 16)),
            ("show_title", "Show Title", "checkbox", False, []),
            ("comp_title", "Gauge Title", "entry", "Power", []),
            ("title_size", "Title Font Size", "spinbox", 14, (8, 32, 2)),
            ("title_rgb", "Title Colour (R,G,B)", "colour_select", "255,255,255", []),
            ("gauge_max", "Max Power (W)", "spinbox", 2000, (100, 10000, 100)),
            ("gauge_min", "Min Power (W)", "spinbox", -500, (-5000, 0, 100)),
            ("gauge_start", "Start Angle (°)", "spinbox", 150, (0, 360, 5)),
            ("gauge_length", "Arc Length (°)", "spinbox", 240, (90, 360, 10)),
            ("gauge_sectors", "Sectors", "spinbox", 6, (2, 20, 1)),
            ("gauge_bg", "Background (R,G,B,A)", "colour_select", "255,255,255,150", []),
            ("gauge_needle", "Needle (R,G,B)", "colour_select", "255,0,0", []),
            ("gauge_tick", "Tick Marks (R,G,B)", "colour_select", "0,0,0", []),
            ("gauge_ann", "Annotations (R,G,B)", "colour_select", "0,0,0", []),
            ("gauge_fill", "Gauge Fill (R,G,B)", "colour_select", "0,191,255", []),
            ("regen_colour", "Regen Zone (R,G,B,A)", "colour_select", "68,187,68,180", []),
        ],
    },
    "voltage_gauge": {
        "title": "Voltage Gauge Options",
        "fields": [
            ("gauge_style", "Gauge Style", "combo", "cairo-gauge-marker",
             ["cairo-gauge-marker", "cairo-gauge-round-annotated", "cairo-gauge-arc-annotated", "cairo-gauge-donut"]),
            ("gauge_size", "Size", "spinbox", 256, (64, 512, 16)),
            ("show_title", "Show Title", "checkbox", False, []),
            ("comp_title", "Gauge Title", "entry", "Voltage", []),
            ("title_size", "Title Font Size", "spinbox", 14, (8, 32, 2)),
            ("title_rgb", "Title Colour (R,G,B)", "colour_select", "255,255,255", []),
            ("gauge_max", "Max Voltage (V)", "spinbox", 100, (48, 150, 1)),
            ("gauge_min", "Min Voltage (V)", "spinbox", 60, (30, 100, 1)),
            ("gauge_start", "Start Angle (°)", "spinbox", 150, (0, 360, 5)),
            ("gauge_length", "Arc Length (°)", "spinbox", 240, (90, 360, 10)),
            ("gauge_sectors", "Sectors", "spinbox", 5, (2, 20, 1)),
            ("gauge_bg", "Background (R,G,B,A)", "colour_select", "255,255,255,150", []),
            ("gauge_needle", "Needle (R,G,B)", "colour_select", "255,0,0", []),
            ("gauge_tick", "Tick Marks (R,G,B)", "colour_select", "0,0,0", []),
            ("gauge_ann", "Annotations (R,G,B)", "colour_select", "0,0,0", []),
            ("gauge_fill", "Gauge Fill (R,G,B)", "colour_select", "0,191,255", []),
        ],
    },
    "compass_display": {
        "title": "Compass Options",
        "fields": [
            ("gauge_size", "Size", "spinbox", 256, (64, 512, 16)),
        ],
    },
    "compass_arrow_display": {
        "title": "Compass Arrow Options",
        "fields": [
            ("gauge_size", "Size", "spinbox", 256, (64, 512, 16)),
        ],
    },
    "circuit_map": {
        "title": "Circuit Map Options",
        "fields": [
            ("gauge_size", "Size", "spinbox", 256, (64, 512, 16)),
        ],
    },
    "speed_bar": {
        "title": "Speed Bar Options",
        "fields": [
            ("show_bar_scale", "Show Min/Max Scale", "checkbox", True, []),
            ("gauge_max", "Max Speed (kph)", "spinbox", 60, (20, 200, 10)),
            ("show_title", "Show Title", "checkbox", False, []),
            ("comp_title", "Bar Title", "entry", "Speed", []),
            ("title_size", "Title Font Size", "spinbox", 14, (8, 32, 2)),
            ("title_rgb", "Title Colour (R,G,B)", "colour_select", "255,255,255", []),
        ],
    },
    "battery_bar": {
        "title": "Battery Bar Options",
        "fields": [
            ("show_bar_scale", "Show Min/Max Scale", "checkbox", True, []),
            ("show_title", "Show Title", "checkbox", False, []),
            ("comp_title", "Bar Title", "entry", "Battery", []),
            ("title_size", "Title Font Size", "spinbox", 14, (8, 32, 2)),
            ("title_rgb", "Title Colour (R,G,B)", "colour_select", "255,255,255", []),
        ],
    },
    "moving_journey_map": {
        "title": "Moving Journey Map Options",
        "fields": [
            ("map_zoom", "Zoom Level", "slider", 16, (8, 20)),
            ("map_corner_radius", "Rounded Corners (0=off)", "spinbox", 35, (0, 128, 5)),
            ("map_opacity", "Opacity", "slider_float", 0.7, (0.0, 1.0)),
        ],
    },
    "asi_gauge": {
        "title": "Airspeed Indicator Options",
        "fields": [
            ("gauge_size", "Size", "spinbox", 256, (64, 512, 16)),
        ],
    },
    "msi_gauge": {
        "title": "Motor Speed Indicator Options",
        "fields": [
            ("gauge_size", "Size", "spinbox", 256, (64, 512, 16)),
            ("msi_green", "Green Zone Start", "spinbox", 0, (0, 200, 5)),
            ("msi_yellow", "Yellow Zone Start", "spinbox", 40, (0, 200, 5)),
            ("msi_end", "Max / Redline", "spinbox", 60, (10, 300, 5)),
            ("msi_rotate", "Rotation (°)", "spinbox", 180, (0, 360, 10)),
            ("msi_textsize", "Text Size", "spinbox", 16, (8, 32, 2)),
        ],
    },
    "heading_tape": {
        "title": "Heading Tape Options",
        "fields": [
            ("ht_visible_range", "Visible Range (°)", "spinbox", 90, (30, 360, 10)),
            ("ht_tick_interval", "Tick Interval (°)", "spinbox", 10, (5, 45, 5)),
            ("ht_label_interval", "Label Interval (°)", "spinbox", 30, (10, 90, 5)),
            ("ht_show_values", "Show Degree Values", "checkbox", True, []),
            ("ht_show_border", "Show Border", "checkbox", True, []),
            ("ht_font_size", "Font Size", "spinbox", 16, (8, 48, 2)),
            ("ht_bg", "Background (R,G,B)", "colour_select", "0,0,0", []),
            ("ht_fg", "Foreground (R,G,B)", "colour_select", "255,255,255", []),
            ("ht_marker", "Marker (R,G,B)", "colour_select", "255,0,0", []),
            ("ht_opacity", "Opacity (0-255)", "spinbox", 180, (0, 255, 10)),
        ],
    },
}

# Default options for any component
DEFAULT_COMPONENT_OPTIONS = {
    "title": "Component Options",
    "fields": [],
}


class ComponentOptionsDialog(tk.Toplevel):
    """Modal dialog for editing properties of a specific overlay component."""

    def __init__(self, parent: LayoutEditorApp, comp: OverlayComponent):
        super().__init__(parent)
        self.parent_app = parent
        self.comp = comp
        self.result_vars: dict[str, tk.Variable] = {}
        self._tile_photo: Optional[ImageTk.PhotoImage] = None  # prevent GC
        self._tile_raw_img: Optional[Image.Image] = None  # raw fetched tile for re-rendering

        opts = COMPONENT_OPTIONS.get(comp.name, DEFAULT_COMPONENT_OPTIONS)
        self.title(opts["title"])
        # Size dialog based on content
        field_count = len(opts.get("fields", []))
        has_map = any(wt == "combo_preview" for _, _, wt, _, _ in opts.get("fields", []))
        if has_map:
            height = 750 + max(0, (field_count - 4)) * 40
        else:
            height = 300 + field_count * 45
        self.geometry(f"520x{min(height, 950)}")
        self.resizable(True, True)
        self.transient(parent)
        self.grab_set()

        self._build_ui(opts)
        self.bind("<Return>", lambda e: self._apply())
        self.bind("<Escape>", lambda e: self.destroy())

    def _build_ui(self, opts: dict):
        # Scrollable content area
        outer = ttk.Frame(self)
        outer.pack(fill=tk.BOTH, expand=True)

        canvas = tk.Canvas(outer, highlightthickness=0)
        scrollbar = ttk.Scrollbar(outer, orient=tk.VERTICAL, command=canvas.yview)
        self._scroll_frame = ttk.Frame(canvas)

        self._scroll_frame.bind("<Configure>",
                                 lambda e: canvas.configure(scrollregion=canvas.bbox("all")))
        canvas.create_window((0, 0), window=self._scroll_frame, anchor=tk.NW)
        canvas.configure(yscrollcommand=scrollbar.set)

        canvas.pack(side=tk.LEFT, fill=tk.BOTH, expand=True)
        scrollbar.pack(side=tk.RIGHT, fill=tk.Y)

        # Enable mousewheel scrolling
        def _on_mousewheel(event):
            canvas.yview_scroll(-1 * (event.delta // 120 or (1 if event.num == 4 else -1)), "units")
        canvas.bind_all("<MouseWheel>", _on_mousewheel)
        canvas.bind_all("<Button-4>", _on_mousewheel)
        canvas.bind_all("<Button-5>", _on_mousewheel)
        self.bind("<Destroy>", lambda e: (canvas.unbind_all("<MouseWheel>"),
                                           canvas.unbind_all("<Button-4>"),
                                           canvas.unbind_all("<Button-5>")))

        container = self._scroll_frame

        # Component name header
        ttk.Label(container, text=self.comp.label,
                  font=("sans-serif", 13, "bold")).pack(padx=12, pady=(12, 4), anchor=tk.W)

        # Position fields (always present)
        pos_frame = ttk.LabelFrame(container, text="Position & Size")
        pos_frame.pack(fill=tk.X, padx=12, pady=4)

        row = ttk.Frame(pos_frame)
        row.pack(fill=tk.X, padx=8, pady=4)
        ttk.Label(row, text="X:").pack(side=tk.LEFT)
        self.x_var = tk.IntVar(value=self.comp.x)
        ttk.Spinbox(row, from_=0, to=9999, textvariable=self.x_var, width=6).pack(side=tk.LEFT, padx=4)
        ttk.Label(row, text="Y:").pack(side=tk.LEFT, padx=(8, 0))
        self.y_var = tk.IntVar(value=self.comp.y)
        ttk.Spinbox(row, from_=0, to=9999, textvariable=self.y_var, width=6).pack(side=tk.LEFT, padx=4)
        ttk.Label(row, text="W:").pack(side=tk.LEFT, padx=(8, 0))
        self.w_var = tk.IntVar(value=self.comp.width)
        ttk.Spinbox(row, from_=16, to=2048, textvariable=self.w_var, width=6).pack(side=tk.LEFT, padx=4)
        ttk.Label(row, text="H:").pack(side=tk.LEFT, padx=(8, 0))
        self.h_var = tk.IntVar(value=self.comp.height)
        ttk.Spinbox(row, from_=16, to=2048, textvariable=self.h_var, width=6).pack(side=tk.LEFT, padx=4)

        # Component-specific fields
        fields = opts.get("fields", [])
        if fields:
            specific_frame = ttk.LabelFrame(container, text="Options")
            specific_frame.pack(fill=tk.X, padx=12, pady=4)

            for field_id, label, widget_type, default, params in fields:
                field_row = ttk.Frame(specific_frame)
                field_row.pack(fill=tk.X, padx=8, pady=3)
                ttk.Label(field_row, text=f"{label}:").pack(side=tk.LEFT)

                current = self._get_current_value(field_id, default)

                if widget_type == "entry":
                    var = tk.StringVar(value=current)
                    ttk.Entry(field_row, textvariable=var, width=20).pack(side=tk.LEFT, padx=4)

                elif widget_type == "spinbox":
                    from_, to_, inc = params
                    var = tk.IntVar(value=current)
                    ttk.Spinbox(field_row, from_=from_, to=to_, increment=inc,
                                textvariable=var, width=8).pack(side=tk.LEFT, padx=4)
                    # Live-update tile preview when relevant fields change
                    if field_id in ("map_corner_radius", "jm_line_width", "jm_loc_size"):
                        var.trace_add("write", lambda *_: self._render_tile_preview())

                elif widget_type == "slider":
                    from_, to_ = params
                    var = tk.IntVar(value=int(current))
                    val_label = ttk.Label(field_row, text=str(int(current)), width=4)
                    val_label.pack(side=tk.RIGHT)
                    def _make_slider_cmd(vl):
                        def _cmd(v):
                            vl.config(text=str(int(float(v))))
                            if self._tile_raw_img:
                                self._render_tile_preview()
                        return _cmd
                    scale = ttk.Scale(field_row, from_=from_, to=to_, orient=tk.HORIZONTAL,
                                      variable=var, command=_make_slider_cmd(val_label))
                    scale.pack(side=tk.LEFT, fill=tk.X, expand=True, padx=4)

                elif widget_type == "slider_float":
                    from_, to_ = params
                    var = tk.DoubleVar(value=float(current))
                    val_label = ttk.Label(field_row, text=f"{float(current):.1f}", width=4)
                    val_label.pack(side=tk.RIGHT)
                    scale = ttk.Scale(field_row, from_=from_, to=to_, orient=tk.HORIZONTAL,
                                      variable=var,
                                      command=lambda v, vl=val_label: vl.config(text=f"{float(v):.1f}"))
                    scale.pack(side=tk.LEFT, fill=tk.X, expand=True, padx=4)

                elif widget_type == "checkbox":
                    current_bool = str(current).lower() in ("true", "1", "yes")
                    var = tk.StringVar(value="true" if current_bool else "false")
                    cb = ttk.Checkbutton(field_row, text="",
                                          command=lambda v=var: v.set("true" if v.get() == "false" else "false"))
                    if current_bool:
                        cb.state(["selected"])
                    else:
                        cb.state(["!selected"])
                    # Sync checkbox state with var
                    def _make_cb_cmd(v, c):
                        def _toggle():
                            if "selected" in c.state():
                                v.set("true")
                            else:
                                v.set("false")
                        return _toggle
                    cb.config(command=_make_cb_cmd(var, cb))
                    cb.pack(side=tk.LEFT, padx=4)

                elif widget_type == "combo":
                    var = tk.StringVar(value=current)
                    combo = ttk.Combobox(field_row, textvariable=var, values=params,
                                          state="readonly", width=24)
                    combo.pack(side=tk.LEFT, padx=4)

                elif widget_type == "combo_editable":
                    # Editable combobox — select from presets or type custom format
                    var = tk.StringVar(value=current)
                    combo = ttk.Combobox(field_row, textvariable=var, values=params, width=24)
                    combo.pack(side=tk.LEFT, padx=4)
                    # Inline preview of the format
                    preview = ttk.Label(field_row, text="", foreground="#888888")
                    preview.pack(side=tk.LEFT, padx=4)
                    combo.bind("<<ComboboxSelected>>",
                               lambda e, v=var, p=preview: self._update_format_preview(v, p))
                    combo.bind("<KeyRelease>",
                               lambda e, v=var, p=preview: self._update_format_preview(v, p))
                    self._update_format_preview(var, preview)

                elif widget_type == "combo_preview":
                    # Combo with tile image preview (for map styles)
                    var = tk.StringVar(value=current)
                    combo = ttk.Combobox(field_row, textvariable=var, values=params,
                                          state="readonly", width=24)
                    combo.pack(side=tk.LEFT, padx=4)

                    desc_label = ttk.Label(specific_frame, text="", wraplength=440,
                                           foreground="#888888")
                    desc_label.pack(padx=8, pady=2, anchor=tk.W)

                    self._tile_canvas = tk.Canvas(specific_frame, width=256, height=256,
                                                   bg="#2a2a2a", highlightthickness=1,
                                                   highlightbackground="#555555")
                    self._tile_canvas.pack(padx=8, pady=4)

                    # API key help area below the preview
                    self._api_help_frame = ttk.Frame(specific_frame)
                    self._api_help_frame.pack(fill=tk.X, padx=8, pady=(0, 4))

                    combo.bind("<<ComboboxSelected>>",
                               lambda e, v=var, dl=desc_label: self._on_map_style_change(v, dl))
                    self._on_map_style_change(var, desc_label)
                elif widget_type == "colour_select":
                    # Colour picker with swatch preview
                    var = tk.StringVar(value=current)
                    # Convert R,G,B[,A] string to hex for the swatch
                    try:
                        parts = [int(c.strip()) for c in str(current).split(",")]
                        hex_col = f"#{parts[0]:02x}{parts[1]:02x}{parts[2]:02x}"
                    except Exception:
                        hex_col = "#ffffff"
                    swatch = tk.Label(field_row, text="  ", bg=hex_col, width=3,
                                      relief=tk.RAISED, borderwidth=2)
                    swatch.pack(side=tk.LEFT, padx=4)
                    val_label = ttk.Label(field_row, text=str(current), width=12)
                    val_label.pack(side=tk.LEFT, padx=2)
                    ttk.Button(field_row, text="Choose...",
                               command=lambda v=var, sw=swatch, vl=val_label:
                                   self._pick_colour(v, sw, vl)).pack(side=tk.LEFT, padx=2)

                elif widget_type == "font_select":
                    # Font selector — shows current font name + browse button
                    current_font = current or self.parent_app.font_path
                    var = tk.StringVar(value=current_font)
                    font_label = ttk.Label(field_row,
                                           text=Path(current_font).name if current_font else "(global)",
                                           foreground="#4488ff", width=20, anchor=tk.W)
                    font_label.pack(side=tk.LEFT, padx=4)
                    ttk.Button(field_row, text="Select...",
                               command=lambda v=var, fl=font_label: self._pick_font(v, fl)
                               ).pack(side=tk.LEFT, padx=2)
                    ttk.Button(field_row, text="Reset",
                               command=lambda v=var, fl=font_label: self._reset_font(v, fl)
                               ).pack(side=tk.LEFT, padx=2)

                else:
                    var = tk.StringVar(value=str(current))
                    ttk.Entry(field_row, textvariable=var, width=12).pack(side=tk.LEFT, padx=4)

                self.result_vars[field_id] = var

            # Hint for date/time format fields only
            if any(fid in ("date_format", "time_format") and wt == "combo_editable"
                   for fid, _, wt, _, _ in fields):
                ttk.Label(specific_frame,
                          text="Codes: %Y=year %m=month %d=day %b=abbr month %B=full month\n"
                               "%H=24h %I=12h %M=min %S=sec %p=AM/PM %a=weekday",
                          foreground="#666666", font=("monospace", 8),
                          justify=tk.LEFT).pack(padx=8, pady=(0, 4), anchor=tk.W)

        # Date/time combined preview
        if opts.get("has_datetime_preview"):
            preview_frame = ttk.LabelFrame(container, text="Preview")
            preview_frame.pack(fill=tk.X, padx=12, pady=4)
            self.dt_preview_label = tk.Label(preview_frame, text="",
                                              font=("monospace", 14), bg="#2a2a2a", fg="white",
                                              height=2, anchor=tk.W, padx=8)
            self.dt_preview_label.pack(fill=tk.X, padx=8, pady=4)
            self._update_full_datetime_preview()

        # Buttons
        btn_frame = ttk.Frame(self)
        btn_frame.pack(pady=12)
        ttk.Button(btn_frame, text="Apply", command=self._apply).pack(side=tk.LEFT, padx=4)
        ttk.Button(btn_frame, text="Cancel", command=self.destroy).pack(side=tk.LEFT, padx=4)

    def _pick_colour(self, var: tk.StringVar, swatch: tk.Label, val_label: ttk.Label):
        """Open colour chooser dialog. Preserves alpha channel if present."""
        from tkinter import colorchooser
        parts = [int(c.strip()) for c in var.get().split(",")]
        try:
            initial = f"#{parts[0]:02x}{parts[1]:02x}{parts[2]:02x}"
        except Exception:
            initial = "#ffffff"
        alpha = parts[3] if len(parts) > 3 else None
        result = colorchooser.askcolor(color=initial, parent=self, title="Choose Colour")
        if result and result[0]:
            r, g, b = [int(c) for c in result[0]]
            if alpha is not None:
                rgb_str = f"{r},{g},{b},{alpha}"
            else:
                rgb_str = f"{r},{g},{b}"
            var.set(rgb_str)
            swatch.config(bg=result[1])
            val_label.config(text=rgb_str)
            # Re-render tile preview if we have one
            if self._tile_raw_img:
                self._render_tile_preview()

    def _pick_font(self, var: tk.StringVar, label: ttk.Label):
        """Open file dialog to pick a font file."""
        path = filedialog.askopenfilename(
            title="Select Font File",
            parent=self,
            filetypes=[("Font files", "*.ttf *.otf *.TTF *.OTF"), ("All files", "*.*")])
        if path and os.path.isfile(path):
            var.set(path)
            label.config(text=Path(path).name)

    def _reset_font(self, var: tk.StringVar, label: ttk.Label):
        """Reset to global font."""
        var.set("")
        label.config(text="(global)")

    def _update_format_preview(self, var: tk.StringVar, preview_label: ttk.Label):
        """Show a live example of the selected strftime format."""
        import datetime as _dt
        fmt = var.get()
        try:
            example = _dt.datetime(2026, 4, 3, 17, 40, 50, 123456).strftime(fmt)
            preview_label.config(text=example)
        except Exception:
            preview_label.config(text="(invalid)")
        if hasattr(self, "dt_preview_label"):
            self._update_full_datetime_preview()

    def _update_full_datetime_preview(self):
        """Update the combined date + time preview."""
        import datetime as _dt
        sample = _dt.datetime(2026, 4, 3, 17, 40, 50, 123456)
        date_fmt = self.result_vars.get("date_format")
        time_fmt = self.result_vars.get("time_format")
        lines = []
        if date_fmt:
            try:
                lines.append(sample.strftime(date_fmt.get()))
            except Exception:
                lines.append("(invalid date format)")
        if time_fmt:
            try:
                lines.append(sample.strftime(time_fmt.get()))
            except Exception:
                lines.append("(invalid time format)")
        self.dt_preview_label.config(text="\n".join(lines))

    def _on_map_style_change(self, var: tk.StringVar, desc_label: ttk.Label):
        """Update map style description, tile preview from GPX location, and API key help."""
        style = var.get()
        info = MAP_STYLE_INFO.get(style)

        # Clear help area
        if hasattr(self, "_api_help_frame"):
            for w in self._api_help_frame.winfo_children():
                w.destroy()

        if not info:
            desc_label.config(text="")
            return

        desc, access, _ = info

        # Build tile URL dynamically from GPX location
        loc = self.parent_app.gpx_location
        if loc:
            tile_url = _build_tile_url(style, loc[0], loc[1], zoom=14)
        else:
            # Fallback: generic location if no GPX loaded
            tile_url = _build_tile_url(style, -27.47, 153.02, zoom=13)

        key_note = ""
        if access.startswith("key:"):
            provider = access.split(":")[1]
            if tile_url:
                key_note = f" (API key: {provider})"
            else:
                key_note = f" (requires {provider} API key - no preview)"
        desc_label.config(text=f"{desc}{key_note}")

        if tile_url:
            self._fetch_tile_preview(tile_url)
            if access.startswith("key:"):
                provider = access.split(":")[1]
                ttk.Label(self._api_help_frame,
                          text=f"{provider} API key active",
                          foreground="#44aa44").pack(anchor=tk.W)
            if loc:
                ttk.Label(self._api_help_frame,
                          text=f"Preview centred on GPX start: {loc[0]:.4f}, {loc[1]:.4f}",
                          foreground="#888888").pack(anchor=tk.W)
        else:
            self._tile_canvas.delete("all")
            self._tile_canvas.create_text(128, 128,
                                           text="No preview available\n(API key required)",
                                           fill="#666666", font=("sans-serif", 10),
                                           justify=tk.CENTER)

            # Show help with clickable link
            provider = access.split(":")[1] if access.startswith("key:") else ""
            _API_KEY_HELP = {
                "thunderforest": {
                    "url": "https://www.thunderforest.com/pricing/",
                    "label": "thunderforest.com/pricing",
                    "env": "API_KEY_THUNDERFOREST",
                    "cli": "--thunderforest-key",
                },
                "geoapify": {
                    "url": "https://myprojects.geoapify.com/",
                    "label": "myprojects.geoapify.com",
                    "env": "API_KEY_GEOAPIFY",
                    "cli": "--geoapify-key",
                },
            }
            pinfo = _API_KEY_HELP.get(provider)
            if pinfo:
                ttk.Label(self._api_help_frame,
                          text=f"A free {provider} API key is required for this style.",
                          foreground="#cc8844").pack(anchor=tk.W, pady=(2, 0))

                ttk.Label(self._api_help_frame,
                          text="Get your free key at:").pack(anchor=tk.W)

                link = tk.Label(self._api_help_frame, text=pinfo["label"],
                                fg="#4488ff", cursor="hand2",
                                font=("sans-serif", 10, "underline"))
                link.pack(anchor=tk.W)
                link.bind("<Button-1>", lambda e, u=pinfo["url"]: _open_url(u))

                ttk.Label(self._api_help_frame,
                          text=f"Then set it via:\n"
                               f"  CLI:  {pinfo['cli']} YOUR_KEY\n"
                               f"  Env:  export {pinfo['env']}=YOUR_KEY\n"
                               f"  Settings panel: API Keys section",
                          foreground="#888888",
                          font=("monospace", 8),
                          justify=tk.LEFT).pack(anchor=tk.W, pady=(2, 0))

    def _fetch_tile_preview(self, url: str):
        """Fetch a map tile in a background thread and display it."""
        self._tile_canvas.delete("all")
        self._tile_canvas.create_text(128, 128, text="Loading...", fill="#666666")

        def _fetch():
            import urllib.request
            try:
                req = urllib.request.Request(url, headers={"User-Agent": "gopro-layout-editor/1.0"})
                with urllib.request.urlopen(req, timeout=10) as resp:
                    data = resp.read()
                img = Image.open(io.BytesIO(data)).convert("RGBA")
                return img
            except Exception:
                return None

        def _on_fetched(img):
            if img:
                self._tile_raw_img = img.resize((256, 256), Image.Resampling.LANCZOS)
                self._render_tile_preview()
            else:
                self._tile_canvas.delete("all")
                self._tile_canvas.create_text(128, 128, text="Failed to load",
                                               fill="#cc4444", font=("sans-serif", 10))

        def _run():
            result = _fetch()
            self.after(0, lambda: _on_fetched(result))

        threading.Thread(target=_run, daemon=True).start()

    def _render_tile_preview(self):
        """Render the cached tile with current settings, route, and marker."""
        if not self._tile_raw_img:
            return
        self._tile_canvas.delete("all")
        img = self._tile_raw_img.copy()

        from PIL import ImageDraw as PilDraw

        # Draw route and marker for journey_map, or position marker for moving_map
        if self.comp.name == "journey_map" and self.parent_app.gpx_path:
            self._draw_route_on_tile(img)
        elif self.comp.name == "moving_map":
            self._draw_position_marker(img)

        # Read corner radius from dialog field if present, else from app
        cr = 0
        if "map_corner_radius" in self.result_vars:
            try:
                cr = int(self.result_vars["map_corner_radius"].get())
            except (ValueError, TypeError):
                cr = self.parent_app.map_corner_radius.get()
        else:
            cr = self.parent_app.map_corner_radius.get()

        if cr > 0:
            img = _apply_rounded_corners(img, cr)

        bordered = img.copy()
        draw = PilDraw.Draw(bordered)
        draw.rounded_rectangle([0, 0, 255, 255], radius=cr,
                                outline=(180, 180, 180), width=3)
        self._tile_photo = ImageTk.PhotoImage(bordered)
        self._tile_canvas.create_image(0, 0, anchor=tk.NW, image=self._tile_photo)

    def _draw_route_on_tile(self, img: Image.Image):
        """Draw GPX route and location marker on the tile preview image."""
        from PIL import ImageDraw as PilDraw
        points = _parse_gpx_points(self.parent_app.gpx_path)
        if len(points) < 2:
            return

        # Get current colours from dialog fields
        def _get_rgb(field_id, default):
            if field_id in self.result_vars:
                try:
                    parts = [int(c.strip()) for c in self.result_vars[field_id].get().split(",")]
                    return tuple(parts[:3])
                except Exception:
                    pass
            return default

        def _get_int(field_id, default):
            if field_id in self.result_vars:
                try:
                    return int(self.result_vars[field_id].get())
                except Exception:
                    pass
            return default

        track_colour = _get_rgb("jm_track_colour", (255, 0, 0))
        line_width = _get_int("jm_line_width", 4)
        loc_colour = _get_rgb("jm_loc_colour", (0, 0, 255))
        loc_outline = _get_rgb("jm_loc_outline", (0, 0, 0))
        loc_size = _get_int("jm_loc_size", 6)

        # Find bounding box of all points to fit them into the tile
        lats = [p[0] for p in points]
        lons = [p[1] for p in points]
        min_lat, max_lat = min(lats), max(lats)
        min_lon, max_lon = min(lons), max(lons)

        # Add padding
        lat_range = max_lat - min_lat or 0.001
        lon_range = max_lon - min_lon or 0.001
        pad = 0.1
        min_lat -= lat_range * pad
        max_lat += lat_range * pad
        min_lon -= lon_range * pad
        max_lon += lon_range * pad
        lat_range = max_lat - min_lat
        lon_range = max_lon - min_lon

        # Convert lat/lon to pixel coords on the 256x256 tile
        def to_px(lat, lon):
            x = int((lon - min_lon) / lon_range * 240 + 8)
            y = int((max_lat - lat) / lat_range * 240 + 8)
            return (x, y)

        # Draw track line
        draw = PilDraw.Draw(img)
        pixel_points = [to_px(lat, lon) for lat, lon, _ in points]
        if len(pixel_points) >= 2:
            draw.line(pixel_points, fill=track_colour, width=line_width)

        # Draw current location marker (at video start position)
        loc = self.parent_app.gpx_location
        if loc:
            mx, my = to_px(loc[0], loc[1])
            r = loc_size
            draw.ellipse([mx - r, my - r, mx + r, my + r],
                         fill=loc_colour, outline=loc_outline, width=1)


    def _draw_position_marker(self, img: Image.Image):
        """Draw a centre position marker on the moving_map tile preview."""
        from PIL import ImageDraw as PilDraw
        draw = PilDraw.Draw(img)
        cx, cy = 128, 128
        # Blue dot with black outline — matches the default moving_map marker
        r = 6
        draw.ellipse([cx - r, cy - r, cx + r, cy + r],
                     fill=(0, 0, 255), outline=(0, 0, 0), width=2)
        # Crosshair lines
        draw.line([cx - 12, cy, cx - r - 2, cy], fill=(0, 0, 255, 180), width=1)
        draw.line([cx + r + 2, cy, cx + 12, cy], fill=(0, 0, 255, 180), width=1)
        draw.line([cx, cy - 12, cx, cy - r - 2], fill=(0, 0, 255, 180), width=1)
        draw.line([cx, cy + r + 2, cx, cy + 12], fill=(0, 0, 255, 180), width=1)

    # Map field_id to the app's tk variable name
    _APP_VAR_MAP = {
        "map_size": "map_size", "map_zoom": "map_zoom",
        "map_corner_radius": "map_corner_radius", "map_opacity": "map_opacity",
        "map_rotate": "map_rotate", "map_style": "map_style",
        "jm_track_colour": "jm_track_colour", "jm_line_width": "jm_line_width",
        "jm_loc_colour": "jm_loc_colour", "jm_loc_outline": "jm_loc_outline",
        "jm_loc_size": "jm_loc_size",
        "date_format": "date_format", "time_format": "time_format",
        "speed_unit": "speed_unit",
    }

    def _get_current_value(self, field_id: str, default):
        """Get current value — check component custom_props first, then app state."""
        if field_id in self.comp.custom_props:
            return self.comp.custom_props[field_id]
        if field_id == "chart_width":
            return self.comp.width
        attr = self._APP_VAR_MAP.get(field_id)
        if attr and hasattr(self.parent_app, attr):
            return getattr(self.parent_app, attr).get()
        return default

    def _apply(self):
        self.comp.x = self.x_var.get()
        self.comp.y = self.y_var.get()
        self.comp.width = self.w_var.get()
        self.comp.height = self.h_var.get()

        # Sync W/H to chart dimensions
        if self.comp.name == "gradient_chart":
            self.comp.custom_props["chart_width"] = self.comp.width
            self.comp.custom_props["chart_height"] = self.comp.height

        # Sync W/H to map_size for map components (maps are always square)
        if self.comp.name in ("moving_map", "journey_map", "moving_journey_map"):
            size = min(self.comp.width, self.comp.height)
            self.comp.width = size
            self.comp.height = size
            self.parent_app.map_size.set(size)
            self.parent_app.map_size_label.config(text=str(size))
            for c in self.parent_app.components:
                if c.name in ("moving_map", "journey_map", "moving_journey_map"):
                    c.width = size
                    c.height = size

        # Sync W/H for square gauge components
        _gauge_names = {"speed_gauge", "battery_gauge", "power_gauge", "voltage_gauge",
                        "compass_display", "compass_arrow_display", "asi_gauge", "msi_gauge",
                        "circuit_map"}
        if self.comp.name in _gauge_names:
            size = min(self.comp.width, self.comp.height)
            self.comp.width = size
            self.comp.height = size
            self.comp.custom_props["gauge_size"] = size

        for field_id, var in self.result_vars.items():
            value = var.get()

            # Map size updates both map components' bounding boxes
            if field_id == "map_size":
                size = int(value)
                self.parent_app.map_size.set(size)
                self.parent_app.map_size_label.config(text=str(size))
                for c in self.parent_app.components:
                    if c.name in ("moving_map", "journey_map"):
                        c.width = size
                        c.height = size
            elif field_id == "map_zoom":
                self.parent_app.map_zoom.set(int(value))
                self.parent_app.map_zoom_label.config(text=str(int(value)))
            elif field_id == "chart_width":
                self.comp.width = int(value)
            elif field_id in ("date_format", "time_format"):
                self.comp.custom_props[field_id] = value
                getattr(self.parent_app, field_id).set(value)
            else:
                # Try updating the app-level variable if mapped
                attr = self._APP_VAR_MAP.get(field_id)
                if attr and hasattr(self.parent_app, attr):
                    app_var = getattr(self.parent_app, attr)
                    try:
                        app_var.set(type(app_var.get())(value))
                    except (ValueError, TypeError):
                        app_var.set(value)
                else:
                    # Store as per-component custom property
                    try:
                        self.comp.custom_props[field_id] = int(value)
                    except (ValueError, TypeError):
                        self.comp.custom_props[field_id] = value

        self.parent_app._redraw_components()
        self.destroy()


# ---------------------------------------------------------------------------
# API Keys dialog
# ---------------------------------------------------------------------------
# GPX / Video time sync dialog
# ---------------------------------------------------------------------------

class SyncDialog(tk.Toplevel):
    """Modal dialog for visually aligning video and GPX altitude data.

    The red marker on the altitude chart tracks the video scrub position
    automatically (using the current offset). A fine-tune slider adjusts
    the offset so the marker can be shifted to align altitude features
    with what is visible in the video frame.
    """

    CHART_W = 800
    CHART_H = 200
    PREVIEW_H = 360

    def __init__(self, parent: LayoutEditorApp, video: VideoEntry,
                 gpx_path: Path, offset_var: tk.DoubleVar):
        super().__init__(parent)
        self.title("GPX / Video Time Sync")
        self.transient(parent)
        self.grab_set()
        self.parent_app = parent

        self.video = video
        self.gpx_path = gpx_path
        self.offset_var = offset_var

        # Parse GPX elevation data
        self.elev_points = _parse_gpx_elevations(gpx_path)
        if not self.elev_points:
            messagebox.showinfo("Error", "No elevation data found in GPX file.", parent=self)
            self.destroy()
            return

        # Calculate video start relative to GPX start
        self.gpx_start = _gpx_start_time(gpx_path)
        self.video_start = self._parse_video_start()
        self.gpx_duration = self.elev_points[-1][0] if self.elev_points else 0
        self.video_duration = video.duration_seconds

        # Base offset: where the video sits on the GPX timeline without adjustment
        if self.gpx_start and self.video_start:
            self.video_gpx_base = (self.video_start - self.gpx_start).total_seconds()
        else:
            self.video_gpx_base = 0.0

        # Current video scrub position (seconds into video)
        self.video_pos = 0.0

        # Working offset (the value being fine-tuned)
        self.working_offset = offset_var.get()

        # Chart zoom: visible window duration in seconds
        self.chart_window_secs = 600.0  # default 10 minutes

        self._frame_queue = queue.Queue()
        self._scrub_after = None
        self._resize_after = None
        # Frame extraction sequencing: latest request wins
        self._frame_seq = 0
        self._frame_thread_running = False
        self._pending_time: Optional[float] = None
        self._seq_lock = threading.Lock()

        self._build_ui()
        self._update_zoom_label()
        self._draw_chart_static()
        self._update_marker()
        self._extract_frame(0.0)

    def _parse_video_start(self):
        from datetime import datetime, timezone
        try:
            ts = self.video.creation_time.replace("Z", "+00:00")
            dt = datetime.fromisoformat(ts)
            if dt.tzinfo is None:
                dt = dt.replace(tzinfo=timezone.utc)
            return dt
        except Exception:
            return None

    def _gpx_time_for_video_pos(self) -> float:
        """GPX timeline position (seconds) for the current video frame."""
        return self.video_gpx_base + self.working_offset + self.video_pos

    # Scrolling window: show this many seconds of the GPX profile centred on the marker
    def _build_ui(self):
        self.geometry(f"{self.CHART_W + 40}x{self.PREVIEW_H + self.CHART_H + 320}")
        self.resizable(True, True)
        self.minsize(600, 500)

        self._aspect = self.video.eff_width / self.video.eff_height

        # -- Video preview (resizable) --
        vid_frame = ttk.LabelFrame(self, text="Video")
        vid_frame.pack(fill=tk.BOTH, expand=True, padx=8, pady=4)

        self.preview_w, self.preview_h = self._calc_preview_size()
        self.video_canvas = tk.Canvas(vid_frame, bg="black")
        self.video_canvas.pack(fill=tk.BOTH, expand=True, pady=4)
        self.video_photo = None

        vid_slider_frame = ttk.Frame(vid_frame)
        vid_slider_frame.pack(fill=tk.X, padx=8, pady=2)
        self.video_slider = ttk.Scale(vid_slider_frame, from_=0,
                                       to=self.video_duration,
                                       orient=tk.HORIZONTAL,
                                       command=self._on_video_scrub)
        self.video_slider.pack(side=tk.LEFT, fill=tk.X, expand=True)
        self.video_time_label = ttk.Label(vid_slider_frame, text="0:00", width=10)
        self.video_time_label.pack(side=tk.RIGHT, padx=4)

        # Playback controls
        play_frame = ttk.Frame(vid_frame)
        play_frame.pack(fill=tk.X, padx=8, pady=2)
        self._play_speed = 0.0  # 0 = paused
        self._play_after_id = None

        ttk.Button(play_frame, text="<<", width=3,
                    command=lambda: self._scrub_video(-30)).pack(side=tk.LEFT, padx=1)
        ttk.Button(play_frame, text="<", width=3,
                    command=lambda: self._scrub_video(-5)).pack(side=tk.LEFT, padx=1)
        self._play_btn = ttk.Button(play_frame, text="Play 1x", width=7,
                                     command=self._toggle_play)
        self._play_btn.pack(side=tk.LEFT, padx=4)
        ttk.Button(play_frame, text="2x", width=3,
                    command=lambda: self._set_play_speed(2.0)).pack(side=tk.LEFT, padx=1)
        ttk.Button(play_frame, text="4x", width=3,
                    command=lambda: self._set_play_speed(4.0)).pack(side=tk.LEFT, padx=1)
        ttk.Button(play_frame, text="8x", width=3,
                    command=lambda: self._set_play_speed(8.0)).pack(side=tk.LEFT, padx=1)
        ttk.Button(play_frame, text=">", width=3,
                    command=lambda: self._scrub_video(5)).pack(side=tk.LEFT, padx=1)
        ttk.Button(play_frame, text=">>", width=3,
                    command=lambda: self._scrub_video(30)).pack(side=tk.LEFT, padx=1)

        # -- Altitude chart (scrolling window with scrollbar) --
        chart_frame = ttk.LabelFrame(self, text="GPX Altitude Profile  (red marker = current video position)")
        chart_frame.pack(fill=tk.X, padx=8, pady=4)

        self.chart_canvas = tk.Canvas(chart_frame, height=self.CHART_H, bg="#1a1a2e")
        self.chart_canvas.pack(fill=tk.X, padx=8, pady=(4, 0))

        # Scrollbar to pan the visible window along the GPX timeline
        self._chart_scroll_var = tk.DoubleVar(value=0.0)
        self.chart_scrollbar = ttk.Scale(chart_frame, from_=0, to=max(0, self.gpx_duration - self.chart_window_secs),
                                          orient=tk.HORIZONTAL, variable=self._chart_scroll_var,
                                          command=self._on_chart_scroll)
        self.chart_scrollbar.pack(fill=tk.X, padx=8, pady=(0, 4))

        # Chart zoom controls
        zoom_row = ttk.Frame(chart_frame)
        zoom_row.pack(fill=tk.X, padx=8, pady=(0, 4))
        ttk.Label(zoom_row, text="Zoom:").pack(side=tk.LEFT, padx=(0, 4))
        ttk.Button(zoom_row, text="-", width=3,
                    command=self._zoom_out).pack(side=tk.LEFT, padx=1)
        ttk.Button(zoom_row, text="+", width=3,
                    command=self._zoom_in).pack(side=tk.LEFT, padx=1)
        self._zoom_label = ttk.Label(zoom_row, text="", width=14)
        self._zoom_label.pack(side=tk.LEFT, padx=4)
        # Preset window durations (seconds)
        for label, secs in [("30s", 30), ("2m", 120), ("10m", 600), ("30m", 1800), ("All", 0)]:
            ttk.Button(zoom_row, text=label, width=4,
                        command=lambda s=secs: self._set_zoom(s)).pack(side=tk.LEFT, padx=1)
        # Track whether user is manually scrolling vs auto-follow
        self._chart_manual_scroll = False

        # -- Offset fine-tune --
        tune_frame = ttk.LabelFrame(self, text="Fine-tune Offset")
        tune_frame.pack(fill=tk.X, padx=8, pady=4)

        slider_row = ttk.Frame(tune_frame)
        slider_row.pack(fill=tk.X, padx=8, pady=4)

        ttk.Button(slider_row, text="-1s", width=4,
                    command=lambda: self._nudge_offset(-1)).pack(side=tk.LEFT, padx=2)
        ttk.Button(slider_row, text="-0.1s", width=5,
                    command=lambda: self._nudge_offset(-0.1)).pack(side=tk.LEFT, padx=2)

        self.offset_slider = ttk.Scale(slider_row, from_=-3600, to=3600,
                                        orient=tk.HORIZONTAL,
                                        command=self._on_offset_slider)
        self.offset_slider.set(self.working_offset)
        self.offset_slider.pack(side=tk.LEFT, fill=tk.X, expand=True, padx=4)

        ttk.Button(slider_row, text="+0.1s", width=5,
                    command=lambda: self._nudge_offset(0.1)).pack(side=tk.LEFT, padx=2)
        ttk.Button(slider_row, text="+1s", width=4,
                    command=lambda: self._nudge_offset(1)).pack(side=tk.LEFT, padx=2)

        entry_row = ttk.Frame(tune_frame)
        entry_row.pack(fill=tk.X, padx=8, pady=(0, 4))

        ttk.Label(entry_row, text="Offset (s):").pack(side=tk.LEFT, padx=4)
        self.offset_entry = ttk.Entry(entry_row, width=8)
        self.offset_entry.insert(0, str(self.working_offset))
        self.offset_entry.pack(side=tk.LEFT, padx=4)
        self.offset_entry.bind("<Return>", self._on_offset_entry)

        self.offset_display = ttk.Label(entry_row, text="", font=("monospace", 11, "bold"))
        self.offset_display.pack(side=tk.LEFT, padx=8)

        self.alt_display = ttk.Label(entry_row, text="", font=("monospace", 10))
        self.alt_display.pack(side=tk.RIGHT, padx=8)

        # -- Buttons --
        btn_frame = ttk.Frame(self)
        btn_frame.pack(fill=tk.X, padx=8, pady=8)
        ttk.Button(btn_frame, text="Apply", command=self._apply).pack(side=tk.RIGHT, padx=4)
        ttk.Button(btn_frame, text="Cancel", command=self.destroy).pack(side=tk.RIGHT, padx=4)
        ttk.Button(btn_frame, text="Reset to 0", command=self._reset_offset).pack(side=tk.LEFT, padx=4)

        # Keyboard bindings for scrubbing and playback
        self.bind("<space>", lambda e: self._toggle_play())
        self.bind("<Left>", lambda e: self._scrub_video(-5))
        self.bind("<Right>", lambda e: self._scrub_video(5))
        self.bind("<Shift-Left>", lambda e: self._scrub_video(-0.5))
        self.bind("<Shift-Right>", lambda e: self._scrub_video(0.5))
        self.bind("<Up>", lambda e: self._nudge_offset(1))
        self.bind("<Down>", lambda e: self._nudge_offset(-1))
        self.bind("<Shift-Up>", lambda e: self._nudge_offset(0.1))
        self.bind("<Shift-Down>", lambda e: self._nudge_offset(-0.1))
        self.bind("+", lambda e: self._zoom_in())
        self.bind("=", lambda e: self._zoom_in())
        self.bind("-", lambda e: self._zoom_out())

        # Resize handler
        self.bind("<Configure>", self._on_resize)
        self.bind("<Escape>", lambda e: self.destroy())
        self.bind("<Return>", lambda e: self._apply())
        self._last_resize_w = 0
        self.focus_set()

    def _calc_preview_size(self):
        if self._aspect >= 1:
            pw = min(self.CHART_W, self.video.eff_width)
            ph = int(pw / self._aspect)
        else:
            ph = min(self.PREVIEW_H, self.video.eff_height)
            pw = int(ph * self._aspect)
        return pw, ph

    def _on_resize(self, event):
        if event.widget is not self:
            return
        new_w = event.width
        if abs(new_w - self._last_resize_w) < 20:
            return
        self._last_resize_w = new_w
        chart_w = self.chart_canvas.winfo_width()
        if chart_w > 50:
            self.CHART_W = chart_w
        self._update_marker()
        # Re-extract frame at new fitted size (debounced)
        if self._resize_after:
            self.after_cancel(self._resize_after)
        self._resize_after = self.after(300, lambda: self._extract_frame(self.video_pos))

    def _on_chart_scroll(self, value):
        """User manually scrolled the chart scrollbar."""
        self._chart_manual_scroll = True
        self._update_marker()

    def _zoom_in(self):
        self._set_zoom(self.chart_window_secs / 2)

    def _zoom_out(self):
        self._set_zoom(self.chart_window_secs * 2)

    def _set_zoom(self, secs: float):
        """Set the visible chart window duration. 0 means show the whole GPX."""
        if secs <= 0 or secs > self.gpx_duration:
            secs = self.gpx_duration
        secs = max(5.0, min(secs, self.gpx_duration))
        self.chart_window_secs = secs
        # Update scrollbar range
        max_scroll = max(0, self.gpx_duration - secs)
        self.chart_scrollbar.config(to=max_scroll)
        # Re-centre on marker after zoom
        self._chart_manual_scroll = False
        self._update_zoom_label()
        self._update_marker()

    def _update_zoom_label(self):
        s = self.chart_window_secs
        if s >= 3600:
            txt = f"{s/3600:.1f}h window"
        elif s >= 60:
            txt = f"{s/60:.1f}m window"
        else:
            txt = f"{s:.0f}s window"
        self._zoom_label.config(text=txt)

    # -- Chart drawing --

    def _chart_window(self) -> tuple[float, float]:
        """Return (win_start, win_end) in GPX seconds for the visible chart window.

        When the user is scrubbing the video, auto-centres on the marker.
        When the user drags the chart scrollbar, uses the scrollbar position.
        """
        if self._chart_manual_scroll:
            win_start = float(self._chart_scroll_var.get())
        else:
            # Auto-centre on marker
            centre = self._gpx_time_for_video_pos()
            win_start = centre - self.chart_window_secs / 2

        win_start = max(0, min(win_start, max(0, self.gpx_duration - self.chart_window_secs)))
        win_end = min(win_start + self.chart_window_secs, self.gpx_duration)

        # Keep scrollbar in sync
        self._chart_scroll_var.set(win_start)
        return win_start, win_end

    def _draw_chart_static(self):
        """Draw the altitude profile and grid for the visible window."""
        c = self.chart_canvas
        c.delete("all")
        w, h = self.CHART_W, self.CHART_H

        if not self.elev_points:
            return

        margin_l, margin_r, margin_t, margin_b = 50, 10, 10, 25
        plot_w = w - margin_l - margin_r
        plot_h = h - margin_t - margin_b

        win_start, win_end = self._chart_window()
        win_dur = max(win_end - win_start, 1)

        # Elevation range for visible window
        visible_elevs = [e for t, e in self.elev_points if win_start <= t <= win_end]
        if not visible_elevs:
            visible_elevs = [e for _, e in self.elev_points]
        min_e, max_e = min(visible_elevs), max(visible_elevs)
        range_e = max(max_e - min_e, 1)
        # Add 5% padding
        min_e -= range_e * 0.05
        max_e += range_e * 0.05
        range_e = max_e - min_e

        # Store geometry for marker positioning
        self._chart_margin_l = margin_l
        self._chart_plot_w = plot_w
        self._chart_margin_t = margin_t
        self._chart_plot_h = plot_h
        self._chart_win_start = win_start
        self._chart_win_dur = win_dur
        self._chart_min_e = min_e
        self._chart_range_e = range_e

        def tx(t):
            return margin_l + ((t - win_start) / win_dur) * plot_w

        def ty(e):
            return margin_t + plot_h - ((e - min_e) / range_e) * plot_h

        # Grid lines with elevation labels
        for i in range(5):
            y = margin_t + i * plot_h / 4
            e = max_e - i * range_e / 4
            c.create_line(margin_l, y, w - margin_r, y, fill="#333355", dash=(2, 4))
            c.create_text(margin_l - 4, y, text=f"{e:.0f}m", anchor=tk.E,
                         fill="#8888aa", font=("sans-serif", 7))

        # Time labels
        for i in range(5):
            t = win_start + i * win_dur / 4
            x = tx(t)
            total_secs = int(t)
            mins = total_secs // 60
            secs = total_secs % 60
            c.create_text(x, h - 5, text=f"{mins}:{secs:02d}", anchor=tk.S,
                         fill="#8888aa", font=("sans-serif", 7))

        # Altitude profile for visible window — downsample for performance
        visible_pts = [(t, e) for t, e in self.elev_points if win_start <= t <= win_end]
        step = max(1, len(visible_pts) // plot_w)
        coords = []
        for i in range(0, len(visible_pts), step):
            t, e = visible_pts[i]
            coords.extend([tx(t), ty(e)])
        if len(coords) >= 4:
            c.create_line(coords, fill="#5b7192", width=2, smooth=True)

    def _update_marker(self):
        """Redraw chart (scrolling window follows marker) and overlay the marker."""
        # Redraw static chart centred on new position
        self._draw_chart_static()

        c = self.chart_canvas
        gpx_t = self._gpx_time_for_video_pos()

        win_start = self._chart_win_start
        win_dur = self._chart_win_dur

        def tx(t):
            return self._chart_margin_l + ((t - win_start) / win_dur) * self._chart_plot_w

        # Video window overlay
        vid_start = self.video_gpx_base + self.working_offset
        vid_end = vid_start + self.video_duration
        vx1 = max(self._chart_margin_l, tx(vid_start))
        vx2 = min(self._chart_margin_l + self._chart_plot_w, tx(vid_end))
        if vx2 > vx1:
            c.create_rectangle(vx1, self._chart_margin_t, vx2,
                              self._chart_margin_t + self._chart_plot_h,
                              fill="#2244aa", stipple="gray12", outline="#4466cc",
                              tags="video_window")
            c.tag_lower("video_window")

        # Red marker line at current GPX position
        if win_start <= gpx_t <= win_start + win_dur:
            x = tx(gpx_t)
            c.create_line(x, self._chart_margin_t, x,
                         self._chart_margin_t + self._chart_plot_h,
                         fill="#ff4444", width=2, tags="marker")

            alt = self._alt_at_time(gpx_t)
            if alt is not None:
                y = self._chart_margin_t + self._chart_plot_h - \
                    ((alt - self._chart_min_e) / self._chart_range_e) * self._chart_plot_h
                c.create_oval(x - 5, y - 5, x + 5, y + 5, fill="#ff4444",
                             outline="white", width=1, tags="marker")
                self.alt_display.config(text=f"Alt: {alt:.1f}m  |  GPX pos: {self._fmt_time(gpx_t)}")
            else:
                self.alt_display.config(text=f"GPX pos: {self._fmt_time(gpx_t)}")
        else:
            self.alt_display.config(text="(marker outside GPX range)")

        self.offset_display.config(text=f"Offset: {self.working_offset:+.1f}s")

    def _alt_at_time(self, t: float) -> Optional[float]:
        """Interpolate altitude at a given GPX time."""
        if not self.elev_points:
            return None
        if t <= self.elev_points[0][0]:
            return self.elev_points[0][1]
        if t >= self.elev_points[-1][0]:
            return self.elev_points[-1][1]
        lo, hi = 0, len(self.elev_points) - 1
        while hi - lo > 1:
            mid = (lo + hi) // 2
            if self.elev_points[mid][0] <= t:
                lo = mid
            else:
                hi = mid
        t0, e0 = self.elev_points[lo]
        t1, e1 = self.elev_points[hi]
        if t1 == t0:
            return e0
        frac = (t - t0) / (t1 - t0)
        return e0 + frac * (e1 - e0)

    # -- Event handlers --

    def _scrub_video(self, delta: float):
        """Scrub the video by delta seconds."""
        self._stop_play()
        new_t = max(0, min(self.video_duration, self.video_pos + delta))
        self.video_slider.set(new_t)
        # Slider callback handles the rest

    def _toggle_play(self):
        """Toggle play/pause at 1x speed."""
        if self._play_speed > 0:
            self._stop_play()
        else:
            self._set_play_speed(1.0)

    def _set_play_speed(self, speed: float):
        """Start or change playback speed."""
        self._play_speed = speed
        self._play_btn.config(text=f"Pause" if speed > 0 else "Play 1x")
        if self._play_after_id is None:
            self._play_tick()

    def _stop_play(self):
        self._play_speed = 0.0
        self._play_btn.config(text="Play 1x")
        if self._play_after_id:
            self.after_cancel(self._play_after_id)
            self._play_after_id = None

    def _play_tick(self):
        """Advance the video by one tick at the current play speed."""
        if self._play_speed <= 0:
            self._play_after_id = None
            return
        # Advance by speed * tick_interval
        tick_ms = 200  # update every 200ms
        delta = self._play_speed * (tick_ms / 1000.0)
        new_t = self.video_pos + delta
        if new_t >= self.video_duration:
            self._stop_play()
            return
        self.video_slider.set(new_t)
        self._play_after_id = self.after(tick_ms, self._play_tick)

    def _on_video_scrub(self, value):
        t = float(value)
        self.video_pos = t
        self.video_time_label.config(text=self._fmt_time(t))

        # Video scrub resets manual scroll so chart follows
        self._chart_manual_scroll = False

        # Request frame immediately; worker coalesces rapid requests
        self._extract_frame(t)

        # Marker follows the video position
        self._update_marker()

    def _on_offset_slider(self, value):
        self.working_offset = round(float(value), 1)
        self.offset_entry.delete(0, tk.END)
        self.offset_entry.insert(0, str(self.working_offset))
        self._chart_manual_scroll = False
        self._update_marker()

    def _on_offset_entry(self, event=None):
        try:
            v = float(self.offset_entry.get())
            self.working_offset = v
            self.offset_slider.set(v)
            self._update_marker()
        except ValueError:
            pass

    def _nudge_offset(self, delta: float):
        self.working_offset = round(self.working_offset + delta, 1)
        self.offset_slider.set(self.working_offset)
        self.offset_entry.delete(0, tk.END)
        self.offset_entry.insert(0, str(self.working_offset))
        self._chart_manual_scroll = False  # re-centre chart on marker
        self._update_marker()

    def _reset_offset(self):
        self._nudge_offset(-self.working_offset)

    def _fit_size(self) -> tuple[int, int]:
        """Calculate frame size that fits the canvas while preserving aspect ratio."""
        cw = self.video_canvas.winfo_width()
        ch = self.video_canvas.winfo_height()
        if cw < 50 or ch < 50:
            return self.preview_w, self.preview_h
        # Fit within canvas bounds
        if cw / ch > self._aspect:
            # Canvas is wider than video — height-limited
            ph = ch
            pw = int(ch * self._aspect)
        else:
            # Canvas is taller than video — width-limited
            pw = cw
            ph = int(cw / self._aspect)
        return max(2, pw), max(2, ph)

    def _extract_frame(self, time_seconds: float):
        """Request a frame at the given time. Always renders the most recent request."""
        with self._seq_lock:
            self._pending_time = time_seconds
            if self._frame_thread_running:
                # Worker will pick up the new time when current extraction finishes
                return
            self._frame_thread_running = True

        threading.Thread(target=self._frame_worker, daemon=True).start()
        self._poll_frame()

    def _frame_worker(self):
        """Single worker: pulls latest pending time and extracts until idle."""
        while True:
            with self._seq_lock:
                t = self._pending_time
                if t is None:
                    self._frame_thread_running = False
                    return
                self._pending_time = None

            pw, ph = self._fit_size()
            img = extract_frame(self.video.path, t, pw, ph)
            self._frame_queue.put(img)

    def _poll_frame(self):
        drained = None
        # Drain the queue and keep only the most recent frame
        try:
            while True:
                drained = self._frame_queue.get_nowait()
        except queue.Empty:
            pass

        if drained is not None:
            self.video_photo = ImageTk.PhotoImage(drained)
            self.video_canvas.delete("all")
            cx = self.video_canvas.winfo_width() // 2
            cy = self.video_canvas.winfo_height() // 2
            self.video_canvas.create_image(cx, cy, anchor=tk.CENTER, image=self.video_photo)

        # Keep polling while worker is running or more pending
        with self._seq_lock:
            still_active = self._frame_thread_running or self._pending_time is not None
        if still_active:
            self.after(30, self._poll_frame)

    def _apply(self):
        self._stop_play()
        self.offset_var.set(round(self.working_offset, 1))
        self.parent_app.gpx_offset_entry.delete(0, tk.END)
        self.parent_app.gpx_offset_entry.insert(0, str(round(self.working_offset, 1)))
        self.destroy()

    @staticmethod
    def _fmt_time(seconds: float) -> str:
        m = int(abs(seconds)) // 60
        s = abs(seconds) % 60
        sign = "-" if seconds < 0 else ""
        return f"{sign}{m}:{int(s):02d}.{int(s * 10) % 10}"


# ---------------------------------------------------------------------------

class ApiKeysDialog(tk.Toplevel):
    """Modal dialog for viewing and editing map API keys."""

    def __init__(self, parent: LayoutEditorApp):
        super().__init__(parent)
        self.parent_app = parent
        self.title("API Keys")
        self.geometry("500x450")
        self.resizable(True, True)
        self.transient(parent)
        self.grab_set()

        self._build_ui()
        self.bind("<Return>", lambda e: self._apply())
        self.bind("<Escape>", lambda e: self.destroy())

    def _build_ui(self):
        ttk.Label(self, text="Map Tile API Keys",
                  font=("sans-serif", 13, "bold")).pack(padx=16, pady=(16, 8), anchor=tk.W)

        ttk.Label(self, text="API keys are needed for Thunderforest (tf-*) and Geoapify (geo-*) map styles.\n"
                             "The osm and cyclosm styles work without any key.",
                  foreground="#888888", wraplength=460,
                  justify=tk.LEFT).pack(padx=16, pady=(0, 8), anchor=tk.W)

        # Thunderforest
        tf_frame = ttk.LabelFrame(self, text="Thunderforest (tf-* styles)")
        tf_frame.pack(fill=tk.X, padx=16, pady=4)

        self.tf_var = tk.StringVar(value=os.environ.get("API_KEY_THUNDERFOREST", ""))
        ttk.Entry(tf_frame, textvariable=self.tf_var, width=50).pack(fill=tk.X, padx=8, pady=4)

        tf_link_row = ttk.Frame(tf_frame)
        tf_link_row.pack(fill=tk.X, padx=8, pady=(0, 4))
        ttk.Label(tf_link_row, text="Get a free key:").pack(side=tk.LEFT)
        tf_link = tk.Label(tf_link_row, text="thunderforest.com/pricing",
                           fg="#4488ff", cursor="hand2",
                           font=("sans-serif", 10, "underline"))
        tf_link.pack(side=tk.LEFT, padx=4)
        tf_link.bind("<Button-1>", lambda e: _open_url("https://www.thunderforest.com/pricing/"))
        ttk.Label(tf_frame, text="Env var: API_KEY_THUNDERFOREST",
                  foreground="#888888", font=("monospace", 9)).pack(padx=8, pady=(0, 4), anchor=tk.W)

        # Geoapify
        geo_frame = ttk.LabelFrame(self, text="Geoapify (geo-* styles)")
        geo_frame.pack(fill=tk.X, padx=16, pady=4)

        self.geo_var = tk.StringVar(value=os.environ.get("API_KEY_GEOAPIFY", ""))
        ttk.Entry(geo_frame, textvariable=self.geo_var, width=50).pack(fill=tk.X, padx=8, pady=4)

        geo_link_row = ttk.Frame(geo_frame)
        geo_link_row.pack(fill=tk.X, padx=8, pady=(0, 4))
        ttk.Label(geo_link_row, text="Get a free key:").pack(side=tk.LEFT)
        geo_link = tk.Label(geo_link_row, text="myprojects.geoapify.com",
                            fg="#4488ff", cursor="hand2",
                            font=("sans-serif", 10, "underline"))
        geo_link.pack(side=tk.LEFT, padx=4)
        geo_link.bind("<Button-1>", lambda e: _open_url("https://myprojects.geoapify.com/"))
        ttk.Label(geo_frame, text="Env var: API_KEY_GEOAPIFY",
                  foreground="#888888", font=("monospace", 9)).pack(padx=8, pady=(0, 4), anchor=tk.W)

        # Buttons
        btn_frame = ttk.Frame(self)
        btn_frame.pack(pady=12)
        ttk.Button(btn_frame, text="Apply", command=self._apply).pack(side=tk.LEFT, padx=4)
        ttk.Button(btn_frame, text="Cancel", command=self.destroy).pack(side=tk.LEFT, padx=4)

    def _apply(self):
        tf = self.tf_var.get().strip()
        geo = self.geo_var.get().strip()
        if tf:
            os.environ["API_KEY_THUNDERFOREST"] = tf
        elif "API_KEY_THUNDERFOREST" in os.environ:
            del os.environ["API_KEY_THUNDERFOREST"]
        if geo:
            os.environ["API_KEY_GEOAPIFY"] = geo
        elif "API_KEY_GEOAPIFY" in os.environ:
            del os.environ["API_KEY_GEOAPIFY"]
        self.parent_app.status_var.set("API keys updated.")
        self.destroy()


# ---------------------------------------------------------------------------
# Font selector dialog
# ---------------------------------------------------------------------------

class FontSelectorDialog(tk.Toplevel):
    """Modal dialog for selecting a font from system fonts or browsing for a file."""

    def __init__(self, parent: LayoutEditorApp):
        super().__init__(parent)
        self.parent_app = parent
        self.title("Select Font")
        self.geometry("500x450")
        self.transient(parent)
        self.grab_set()

        self.selected_path: Optional[str] = None

        self._build_ui()
        self._load_fonts()
        self.bind("<Return>", lambda e: self._apply())
        self.bind("<Escape>", lambda e: self.destroy())

    def _build_ui(self):
        ttk.Label(self, text="Select a font for the overlay:",
                  font=("sans-serif", 11)).pack(padx=12, pady=(12, 4), anchor=tk.W)

        # Search/filter
        filter_frame = ttk.Frame(self)
        filter_frame.pack(fill=tk.X, padx=12, pady=4)
        ttk.Label(filter_frame, text="Filter:").pack(side=tk.LEFT)
        self.filter_var = tk.StringVar()
        self.filter_var.trace_add("write", self._on_filter)
        filter_entry = ttk.Entry(filter_frame, textvariable=self.filter_var)
        filter_entry.pack(side=tk.LEFT, fill=tk.X, expand=True, padx=4)
        filter_entry.focus_set()

        # Buttons — pack at bottom first so they're always visible
        btn_frame = ttk.Frame(self)
        btn_frame.pack(side=tk.BOTTOM, pady=8)
        ttk.Button(btn_frame, text="Browse File...", command=self._browse).pack(side=tk.LEFT, padx=4)
        ttk.Button(btn_frame, text="Apply", command=self._apply).pack(side=tk.LEFT, padx=4)
        ttk.Button(btn_frame, text="Cancel", command=self.destroy).pack(side=tk.LEFT, padx=4)

        # Preview — pack before listbox so it stays visible
        self.preview_label = tk.Label(self, text="0123456789 km/h",
                                       font=("sans-serif", 20), bg="#2a2a2a", fg="white",
                                       height=2)
        self.preview_label.pack(side=tk.BOTTOM, fill=tk.X, padx=12, pady=4)

        # Current selection display
        self.selection_label = ttk.Label(self, text=f"Current: {self.parent_app.font_path}",
                                          wraplength=460)
        self.selection_label.pack(side=tk.BOTTOM, padx=12, pady=4, anchor=tk.W)

        # Font list — fills remaining space
        list_frame = ttk.Frame(self)
        list_frame.pack(fill=tk.BOTH, expand=True, padx=12, pady=4)
        scrollbar = ttk.Scrollbar(list_frame)
        scrollbar.pack(side=tk.RIGHT, fill=tk.Y)
        self.font_listbox = tk.Listbox(list_frame, yscrollcommand=scrollbar.set,
                                        font=("monospace", 10))
        self.font_listbox.pack(fill=tk.BOTH, expand=True)
        scrollbar.config(command=self.font_listbox.yview)
        self.font_listbox.bind("<<ListboxSelect>>", self._on_select)
        self.font_listbox.bind("<Double-1>", lambda e: self._apply())

        self.all_fonts: list[tuple[str, str]] = []  # (display_name, path)

    def _load_fonts(self):
        """Load system fonts using fc-list."""
        self.all_fonts = []
        try:
            result = subprocess.run(
                ["fc-list", "--format=%{file}\t%{family}\t%{style}\n"],
                capture_output=True, text=True, timeout=10)
            seen_paths = set()
            for line in sorted(result.stdout.strip().split("\n")):
                parts = line.split("\t")
                if len(parts) >= 2:
                    fpath = parts[0].strip()
                    family = parts[1].strip().split(",")[0]
                    style = parts[2].strip().split(",")[0] if len(parts) > 2 else ""
                    # Only show .ttf and .otf files
                    if fpath.lower().endswith((".ttf", ".otf")) and fpath not in seen_paths:
                        seen_paths.add(fpath)
                        display = f"{family} {style}".strip()
                        self.all_fonts.append((display, fpath))

            self.all_fonts.sort(key=lambda x: x[0].lower())
        except (FileNotFoundError, subprocess.TimeoutExpired):
            pass

        self._populate_list(self.all_fonts)

    def _populate_list(self, fonts: list[tuple[str, str]]):
        self.font_listbox.delete(0, tk.END)
        self._filtered_fonts = fonts
        for display, path in fonts:
            self.font_listbox.insert(tk.END, display)

    def _on_filter(self, *args):
        query = self.filter_var.get().lower()
        if not query:
            self._populate_list(self.all_fonts)
        else:
            filtered = [(d, p) for d, p in self.all_fonts if query in d.lower()]
            self._populate_list(filtered)

    def _on_select(self, event):
        sel = self.font_listbox.curselection()
        if sel and sel[0] < len(self._filtered_fonts):
            display, path = self._filtered_fonts[sel[0]]
            self.selected_path = path
            self.selection_label.config(text=f"Selected: {display}\n{path}")
            # Try to update preview with the selected font
            try:
                from tkinter.font import Font as TkFont
                preview_font = TkFont(family=display.split()[0], size=20)
                self.preview_label.config(font=preview_font)
            except Exception:
                pass

    def _browse(self):
        path = filedialog.askopenfilename(
            title="Select Font File",
            filetypes=[("Font files", "*.ttf *.otf *.TTF *.OTF"), ("All files", "*.*")])
        if path:
            self.selected_path = path
            self.selection_label.config(text=f"Selected: {Path(path).name}\n{path}")

    def _apply(self):
        if self.selected_path and os.path.isfile(self.selected_path):
            self.parent_app.font_path = self.selected_path
            self.parent_app.font_label.config(text=Path(self.selected_path).name)
            self.parent_app.status_var.set(f"Font: {Path(self.selected_path).name}")
            self.destroy()
        else:
            messagebox.showwarning("Invalid Font", "Please select a valid font file.", parent=self)


# ---------------------------------------------------------------------------
# Encoding progress dialog
# ---------------------------------------------------------------------------

class EncodingDialog(tk.Toplevel):

    def __init__(self, parent, videos: list[VideoEntry], gpx_path: Path,
                 layout_xml_path: Path, speed_unit: str, font: str,
                 gpu_profile: Optional[str], dashboard_script: Path,
                 map_style: str = "osm", gpx_time_offset: float = 0.0,
                 sample_duration: Optional[float] = None):
        super().__init__(parent)
        self.title("Encoding Progress")
        self.geometry("1800x800")
        self.transient(parent)

        self.videos = videos
        self.gpx_path = gpx_path
        self.layout_xml_path = layout_xml_path
        self.speed_unit = speed_unit
        self.font = font
        self.gpu_profile = gpu_profile
        self.dashboard_script = dashboard_script
        self.map_style = map_style
        self.gpx_time_offset = gpx_time_offset
        self.sample_duration = sample_duration
        self.process: Optional[subprocess.Popen] = None
        self.cancelled = False

        self._build_ui()
        self.bind("<Escape>", lambda e: self._cancel())
        self.after(100, self._start_encoding)

    def _build_ui(self):
        ttk.Label(self, text="Encoding Videos", font=("sans-serif", 14, "bold")).pack(pady=8)

        self.overall_label = ttk.Label(self, text="Preparing...")
        self.overall_progress = ttk.Progressbar(self, mode="determinate")
        if len(self.videos) > 1:
            self.overall_label.pack(padx=16, pady=4, anchor=tk.W)
            self.overall_progress.pack(fill=tk.X, padx=16, pady=4)

        self.current_label = ttk.Label(self, text="")
        self.current_label.pack(padx=16, pady=4, anchor=tk.W)

        self.current_progress = ttk.Progressbar(self, mode="determinate")
        self.current_progress.pack(fill=tk.X, padx=16, pady=4)

        # Log output
        self.log_text = tk.Text(self, height=10, state=tk.DISABLED, bg="#1a1a1a", fg="#cccccc",
                                font=("monospace", 9))
        self.log_text.pack(fill=tk.BOTH, expand=True, padx=16, pady=8)

        btn_frame = ttk.Frame(self)
        btn_frame.pack(pady=8)
        self.cancel_btn = ttk.Button(btn_frame, text="Cancel", command=self._cancel)
        self.cancel_btn.pack(side=tk.LEFT, padx=8)
        self.close_btn = ttk.Button(btn_frame, text="Close", command=self.destroy, state=tk.DISABLED)
        self.close_btn.pack(side=tk.LEFT, padx=8)

    def _log(self, msg: str):
        self.log_text.config(state=tk.NORMAL)
        self.log_text.insert(tk.END, msg + "\n")
        self.log_text.see(tk.END)
        self.log_text.config(state=tk.DISABLED)

    def _start_encoding(self):
        self.overall_progress.config(maximum=len(self.videos))
        threading.Thread(target=self._encode_all, daemon=True).start()

    def _encode_all(self):
        for i, video in enumerate(self.videos):
            if self.cancelled:
                self._ui_update(lambda: self._log("Cancelled."))
                break

            self._ui_update(lambda v=video, idx=i: self._update_labels(v, idx))

            success = self._encode_one(video)

            self._ui_update(lambda idx=i: self.overall_progress.config(value=idx + 1))

            if not success and not self.cancelled:
                self._ui_update(lambda v=video: self._log(f"FAILED: {v.basename}"))

        self._ui_update(self._encoding_complete)

    def _update_labels(self, video: VideoEntry, idx: int):
        self.overall_label.config(text=f"Video {idx + 1} of {len(self.videos)}")
        self.current_label.config(text=f"Processing: {video.basename}")
        self.current_progress.config(value=0)
        self._log(f"Starting: {video.basename} ({video.eff_width}x{video.eff_height})")

    def _encode_one(self, video: VideoEntry) -> bool:
        out_path = video.path.parent / (video.path.stem + "_overlay" + video.path.suffix)

        gpx_arg = "--xlsx" if self.gpx_path.suffix.lower() == ".xlsx" else "--gpx"
        cmd = [
            sys.executable, str(self.dashboard_script),
            "--font", self.font,
            gpx_arg, str(self.gpx_path),
            "--use-gpx-only",
            "--video-time-start", "video-created",
            "--overlay-size", f"{video.eff_width}x{video.eff_height}",
            "--units-speed", self.speed_unit,
            "--layout", "xml",
            "--layout-xml", str(self.layout_xml_path),
            "--map-style", self.map_style,
        ]
        if self.gpx_time_offset:
            cmd.extend(["--gpx-time-offset", str(self.gpx_time_offset)])
        if self.sample_duration:
            cmd.extend(["--sample-duration", str(self.sample_duration)])
        if self.gpu_profile:
            cmd.extend(["--profile", self.gpu_profile])
        cmd.extend([str(video.path), str(out_path)])

        try:
            self.process = subprocess.Popen(
                cmd, stdout=subprocess.PIPE, stderr=subprocess.STDOUT,
                text=True, bufsize=1)

            total_frames = None
            for line in self.process.stdout:
                if self.cancelled:
                    self.process.terminate()
                    return False

                line = line.strip()

                # Parse "Render: N [ P%]" lines
                if line.startswith("Render:"):
                    try:
                        parts = line.split("[")
                        if len(parts) >= 2:
                            pct_str = parts[1].split("%")[0].strip()
                            pct = int(pct_str)
                            self._ui_update(
                                lambda p=pct: self.current_progress.config(value=p, maximum=100))
                    except (ValueError, IndexError):
                        pass
                elif "Timeseries has" in line:
                    # "Timeseries has N data points"
                    try:
                        n = int(line.split("has")[1].split("data")[0].strip())
                        total_frames = n
                    except (ValueError, IndexError):
                        pass
                elif line and not line.startswith("frame="):
                    self._ui_update(lambda l=line: self._log(l))

            self.process.wait()
            return self.process.returncode == 0

        except Exception as e:
            self._ui_update(lambda: self._log(f"Error: {e}"))
            return False

    def _cancel(self):
        self.cancelled = True
        if self.process:
            self.process.terminate()
        self.cancel_btn.config(state=tk.DISABLED)

    def _encoding_complete(self):
        self._cleanup_layout_xml()
        self.cancel_btn.config(state=tk.DISABLED)
        self.close_btn.config(state=tk.NORMAL)
        if not self.cancelled:
            self._log("All encoding complete.")
            self.current_label.config(text="Complete!")
        else:
            self.current_label.config(text="Cancelled.")

    def _cleanup_layout_xml(self):
        try:
            if self.layout_xml_path and self.layout_xml_path.exists():
                self.layout_xml_path.unlink()
        except Exception:
            pass

    def _ui_update(self, func):
        """Schedule a function to run on the main UI thread."""
        self.after(0, func)


# ---------------------------------------------------------------------------
# Helpers
# ---------------------------------------------------------------------------

def _apply_rounded_corners(img: Image.Image, radius: int) -> Image.Image:
    """Apply rounded corners to a PIL image using an alpha mask."""
    from PIL import ImageDraw as PilDraw
    mask = Image.new("L", img.size, 0)
    draw = PilDraw.Draw(mask)
    draw.rounded_rectangle([0, 0, img.size[0] - 1, img.size[1] - 1],
                           radius=radius, fill=255)
    img = img.copy()
    img.putalpha(mask)
    return img


def _open_url(url: str):
    """Open a URL in the default browser."""
    import webbrowser
    webbrowser.open(url)


def _format_time(seconds: float) -> str:
    h = int(seconds // 3600)
    m = int((seconds % 3600) // 60)
    s = int(seconds % 60)
    return f"{h:02d}:{m:02d}:{s:02d}"


# ---------------------------------------------------------------------------
# Entry point
# ---------------------------------------------------------------------------

def _load_dotenv(path: str):
    """Load key=value pairs from a .env file into os.environ."""
    try:
        with open(path) as f:
            for line in f:
                line = line.strip()
                if not line or line.startswith("#"):
                    continue
                if "=" in line:
                    key, _, value = line.partition("=")
                    key = key.strip()
                    value = value.strip().strip("'\"")
                    os.environ[key] = value
    except FileNotFoundError:
        print(f"Warning: .env file not found: {path}")


if __name__ == "__main__":
    import argparse
    parser = argparse.ArgumentParser(description="GoPro Dashboard Overlay — Visual Layout Editor")
    parser.add_argument("workdir", nargs="?", default=None,
                        help="Working folder — sets initial directory for file dialogs (or set GOPRO_WORKDIR env var)")
    parser.add_argument("--videos", nargs="+", default=[], help="Video files to add on startup")
    parser.add_argument("--thunderforest-key", help="Thunderforest API key for tf-* map styles")
    parser.add_argument("--geoapify-key", help="Geoapify API key for geo-* map styles")
    parser.add_argument("--env-file", help="Path to .env file with API keys (API_KEY_THUNDERFOREST, API_KEY_GEOAPIFY)")
    parser.add_argument("--gpx", help="GPX/FIT route file to open on startup")
    parser.add_argument("--xlsx", help="EUC World XLSX file to open as route on startup")
    parser.add_argument("--layout", help="Load layout XML file (or set GOPRO_LAYOUT_XML env var)")
    cli_args = parser.parse_args()

    # Load .env file if specified
    if cli_args.env_file:
        _load_dotenv(cli_args.env_file)

    # Set API keys from CLI args into environment
    if cli_args.thunderforest_key:
        os.environ["API_KEY_THUNDERFOREST"] = cli_args.thunderforest_key
    if cli_args.geoapify_key:
        os.environ["API_KEY_GEOAPIFY"] = cli_args.geoapify_key

    app = LayoutEditorApp()

    # Set working directory
    if cli_args.workdir:
        app.workdir = os.path.abspath(cli_args.workdir)
    elif os.environ.get("GOPRO_WORKDIR"):
        app.workdir = os.path.abspath(os.environ["GOPRO_WORKDIR"])

    # Load route file from CLI args (--gpx or --xlsx)
    route_file = cli_args.xlsx or cli_args.gpx
    if route_file:
        app.gpx_path = Path(route_file)
        suffix = app.gpx_path.suffix.lower()
        type_label = {".gpx": "GPX", ".fit": "FIT", ".xlsx": "XLSX"}.get(suffix, suffix.upper())
        app.gpx_label.config(text=f"[{type_label}] {app.gpx_path.name}")
        if suffix != ".xlsx":
            loc = _extract_gpx_first_point(app.gpx_path)
            if loc:
                app.gpx_location = loc

    for v in (cli_args.videos or []):
        try:
            entry = analyse_video(Path(v))
            app.videos.append(entry)
            app.video_listbox.insert(tk.END, entry.basename)
        except Exception as e:
            print(f"Warning: failed to load {v}: {e}")
    if app.videos:
        app.video_listbox.selection_set(0)
        app._select_video(0)

    # Load layout XML from CLI arg or env var
    layout_xml = cli_args.layout or os.environ.get("GOPRO_LAYOUT_XML")
    if layout_xml:
        app._pending_layout_xml = layout_xml
        if app.components:
            app._load_layout_xml(layout_xml)
            app._pending_layout_xml = None

    app.mainloop()
