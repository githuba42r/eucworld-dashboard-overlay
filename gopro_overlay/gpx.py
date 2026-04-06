import collections
import gzip
import re
from pathlib import Path
from typing import List

import gpxpy

from .gpmf import GPSFix
from .point import Point
from .timeseries import Timeseries, Entry

GPX = collections.namedtuple("GPX", "time lat lon alt hr cad atemp power speed battery voltage current")


def _preprocess_gpx(xml_str: str) -> str:
    """Move non-standard <speed> elements (e.g. from EUC World) into
    TrackPointExtension so gpxpy can parse them.

    EUC World exports <speed> as a direct child of <trkpt>, which is not
    valid GPX 1.1.  This rewrites each occurrence into the standard
    Garmin TrackPointExtension namespace.
    """
    tpx_ns = "http://www.garmin.com/xmlschemas/TrackPointExtension/v1"

    # Match <speed>VALUE</speed> that is a direct child of <trkpt> (not inside <extensions>).
    # Use a simple regex that handles both namespaced and non-namespaced variants.
    def _move_speed(match):
        # EUC World exports speed in km/h; GPX standard expects m/s
        speed_kph = float(match.group(1))
        speed_mps = speed_kph / 3.6
        ext_block = (
            f'<extensions>'
            f'<gpxtpx:TrackPointExtension xmlns:gpxtpx="{tpx_ns}">'
            f'<gpxtpx:speed>{speed_mps:.2f}</gpxtpx:speed>'
            f'</gpxtpx:TrackPointExtension>'
            f'</extensions>'
        )
        return ext_block

    # Remove <speed> from trkpt body and add it to extensions.
    # Handle both with and without GPX namespace prefix.
    # First check if this GPX has speed as direct child (not in extensions)
    if re.search(r'<(?:[a-z]+:)?trkpt[^>]*>.*?<(?:[a-z]+:)?speed>', xml_str[:5000], re.DOTALL):
        # Extract and relocate speed elements
        # Pattern: match <speed>...</speed> (with optional namespace prefix)
        ns_pattern = r'<(?:{[^}]+})?speed>([^<]+)</(?:{[^}]+})?speed>\s*'
        plain_pattern = r'<speed>([^<]+)</speed>\s*'

        # Only process if there are no existing <extensions> blocks with speed
        if 'TrackPointExtension' not in xml_str:
            # Replace each <speed>val</speed> with extension block
            xml_str = re.sub(plain_pattern, _move_speed, xml_str)
            # Handle namespace-prefixed version too
            xml_str = re.sub(ns_pattern, _move_speed, xml_str)

    return xml_str


def fudge(gpx):
    for track in gpx.tracks:
        for segment in track.segments:
            for point in segment.points:
                data = {
                    "time": point.time,
                    "lat": point.latitude,
                    "lon": point.longitude,
                    "alt": point.elevation,
                    "atemp": None,
                    "hr": None,
                    "cad": None,
                    "power": None,
                    "speed": None,
                    "battery": None,
                    "voltage": None,
                    "current": None,
                }
                for extension in point.extensions:
                    for element in extension.iter():
                        tag = element.tag[element.tag.find("}") + 1:]
                        if tag in ("atemp", "hr", "cad", "power", "speed"):
                            data[tag] = float(element.text)
                        elif tag in ("battery", "voltage", "current"):
                            data[tag] = float(element.text)
                yield GPX(**data)


def with_unit(gpx, units):
    return GPX(
        gpx.time,
        gpx.lat,
        gpx.lon,
        units.Quantity(gpx.alt, units.m) if gpx.alt is not None else None,
        units.Quantity(gpx.hr, units.bpm) if gpx.hr is not None else None,
        units.Quantity(gpx.cad, units.rpm) if gpx.cad is not None else None,
        units.Quantity(gpx.atemp, units.celsius) if gpx.atemp is not None else None,
        units.Quantity(gpx.power, units.watt) if gpx.power is not None else None,
        units.Quantity(gpx.speed, units.mps) if gpx.speed is not None else None,
        units.Quantity(gpx.battery, units.percent) if gpx.battery is not None else None,
        units.Quantity(gpx.voltage, units.volt) if gpx.voltage is not None else None,
        units.Quantity(gpx.current, units.ampere) if gpx.current is not None else None,
    )


def load(filepath: Path, units):
    if filepath.suffix == ".gz":
        with gzip.open(filepath, 'rb') as gpx_file:
            xml_str = gpx_file.read().decode('utf-8')
    else:
        with filepath.open('r') as gpx_file:
            xml_str = gpx_file.read()
    xml_str = _preprocess_gpx(xml_str)
    return load_xml(xml_str, units)


def load_xml(file_or_str, units) -> List[GPX]:
    gpx = gpxpy.parse(file_or_str)

    return [with_unit(p, units) for p in fudge(gpx)]


def gpx_to_timeseries(gpx: List[GPX], units):
    gpx_timeseries = Timeseries()

    points = [
        Entry(
            point.time,
            point=Point(point.lat, point.lon),
            alt=point.alt,
            hr=point.hr,
            cad=point.cad,
            atemp=point.atemp,
            power=point.power,
            speed=point.speed,
            battery=point.battery,
            voltage=point.voltage,
            current=point.current,
            packet=units.Quantity(index),
            packet_index=units.Quantity(0),
            # we should set the gps fix or Journey.accept() will skip the point:
            gpsfix=GPSFix.LOCK_3D.value,
            gpslock=units.Quantity(GPSFix.LOCK_3D.value)
        )
        for index, point in enumerate(gpx)
    ]

    gpx_timeseries.add(*points)

    return gpx_timeseries


def load_timeseries(filepath: Path, units) -> Timeseries:
    return gpx_to_timeseries(load(filepath, units), units)


def load_xlsx(filepath: Path, units) -> List[GPX]:
    """Load EUC World XLSX export and return GPX-compatible point list."""
    try:
        import openpyxl
    except ImportError:
        raise ImportError("openpyxl is required for XLSX support: pip install openpyxl")

    from datetime import datetime as dt_cls, timezone as tz_cls

    wb = openpyxl.load_workbook(filepath, read_only=True)
    ws = wb.active

    headers = []
    for row in ws.iter_rows(min_row=1, max_row=1, values_only=True):
        headers = [h.strip() if h else "" for h in row]

    col = {h: i for i, h in enumerate(headers)}

    lat_col = col.get("GPS Latitude [°]")
    lon_col = col.get("GPS Longitude [°]")
    time_col = col.get("Date & Time")
    if lat_col is None or lon_col is None or time_col is None:
        raise IOError(f"XLSX missing required columns (GPS Latitude, GPS Longitude, Date & Time). Found: {headers}")

    alt_col = col.get("GPS Altitude [m]")
    speed_col = col.get("Speed [km/h]")
    gps_speed_col = col.get("GPS Speed [km/h]")
    battery_col = col.get("Battery [%]")
    voltage_col = col.get("Voltage [V]")
    current_col = col.get("Current [A]")
    power_col = col.get("Power [W]")
    temp_col = col.get("Temperature [°C]")

    local_tz = dt_cls.now().astimezone().tzinfo

    def safe_float(row, idx):
        if idx is not None and idx < len(row) and row[idx] is not None:
            try:
                return float(row[idx])
            except (ValueError, TypeError):
                pass
        return None

    points = []
    for row in ws.iter_rows(min_row=2, values_only=True):
        dt_val = row[time_col]
        lat = row[lat_col]
        lon = row[lon_col]
        if dt_val is None or lat is None or lon is None:
            continue

        if isinstance(dt_val, dt_cls):
            point_dt = dt_val
        else:
            point_dt = dt_cls.strptime(str(dt_val), "%Y-%m-%d %H:%M:%S")
        if point_dt.tzinfo is None:
            point_dt = point_dt.replace(tzinfo=local_tz)
        point_dt = point_dt.astimezone(tz_cls.utc)

        alt = safe_float(row, alt_col)
        spd = safe_float(row, speed_col) or safe_float(row, gps_speed_col)
        spd_mps = spd / 3.6 if spd is not None else None
        battery = safe_float(row, battery_col)
        voltage = safe_float(row, voltage_col)
        current = safe_float(row, current_col)
        pwr = safe_float(row, power_col)
        temp = safe_float(row, temp_col)

        points.append(GPX(
            time=point_dt,
            lat=float(lat),
            lon=float(lon),
            alt=alt,
            hr=None,
            cad=None,
            atemp=temp,
            power=pwr,
            speed=spd_mps,
            battery=battery,
            voltage=voltage,
            current=current,
        ))

    wb.close()

    return [with_unit(p, units) for p in points]


def load_xlsx_timeseries(filepath: Path, units) -> Timeseries:
    return gpx_to_timeseries(load_xlsx(filepath, units), units)
